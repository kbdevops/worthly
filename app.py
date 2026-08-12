"""
Net Worth Tracker
-----------------
Flask application with SQLite backend, multi-currency support (AUD base),
and detailed portfolio analytics. All data stored in prices.db.
"""

from flask import Flask, render_template, request, jsonify, send_from_directory
import yfinance as yf
import pandas as pd
import requests
import csv
import io
import xml.etree.ElementTree as ET
import json
import os
import sqlite3
import time
import threading
import uuid
from datetime import datetime, date, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from apscheduler.schedulers.background import BackgroundScheduler
from werkzeug.security import generate_password_hash, check_password_hash
from flask_jwt_extended import (
    JWTManager, create_access_token, jwt_required, get_jwt_identity,
)

app = Flask(__name__, template_folder="templates", static_folder="static")

# SECURITY: JWT_SECRET_KEY must be set via env var in any real deployment — the
# fallback here is only so the app doesn't hard-crash on a fresh dev checkout.
# A default secret means anyone can forge valid tokens; this is not safe to ship
# as-is to a real multi-user deployment without setting this explicitly.
app.config["JWT_SECRET_KEY"] = os.environ.get("JWT_SECRET_KEY", "dev-only-insecure-change-me")
# Long enough that a phone doesn't ask for a password every visit, short enough that a
# leaked token expires on its own. There is no revocation list, so this window IS the
# containment. Override with JWT_DAYS.
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(days=int(os.environ.get("JWT_DAYS", "7")))
jwt = JWTManager(app)

# ── Exposure hardening ────────────────────────────────────────────────────────
# Registration is CLOSED by default. Anyone who reached an open /api/register on an
# internet-facing instance could create an account inside the same SQLite file as the
# owner's data, where separation is only application-level user_id filtering. The one
# exception is bootstrapping a brand-new instance that has no users yet.
ALLOW_REGISTRATION = os.environ.get("ALLOW_REGISTRATION", "").lower() in ("1", "true", "yes")
MIN_PASSWORD_LEN = int(os.environ.get("MIN_PASSWORD_LEN", "10"))
# True when this instance is reachable from outside the LAN; tightens a few defaults.
PUBLIC_MODE = os.environ.get("PUBLIC_MODE", "").lower() in ("1", "true", "yes")

# Per-IP throttle for the credential endpoints. In-process and best-effort — it exists so
# an exposed login can't be brute-forced at machine speed, not as a substitute for an
# edge rate limiter. Single worker (see deploy/entrypoint.sh) keeps this counter coherent.
_AUTH_HITS = {}
_AUTH_LOCK = threading.Lock()
AUTH_MAX_ATTEMPTS = int(os.environ.get("AUTH_MAX_ATTEMPTS", "10"))
AUTH_WINDOW_SECONDS = int(os.environ.get("AUTH_WINDOW_SECONDS", "900"))


def _client_ip():
    """Caller's IP, honouring one proxy hop.

    Behind Traefik every request appears to come from the ingress, so throttling on
    remote_addr alone would rate-limit the whole internet as a single client. Only the
    left-most X-Forwarded-For entry is used, and only when TRUST_PROXY is set — the
    header is client-supplied and trivially spoofed otherwise.
    """
    if os.environ.get("TRUST_PROXY", "").lower() in ("1", "true", "yes"):
        fwd = request.headers.get("X-Forwarded-For", "")
        if fwd:
            return fwd.split(",")[0].strip()
    return request.remote_addr or "unknown"


def _auth_throttled(bucket):
    """Record an attempt; return seconds to wait if the caller is over the limit."""
    ip = _client_ip()
    key = f"{bucket}:{ip}"
    now = time.time()
    with _AUTH_LOCK:
        hits = [t for t in _AUTH_HITS.get(key, []) if now - t < AUTH_WINDOW_SECONDS]
        if len(hits) >= AUTH_MAX_ATTEMPTS:
            _AUTH_HITS[key] = hits
            return int(AUTH_WINDOW_SECONDS - (now - hits[0])) + 1
        hits.append(now)
        _AUTH_HITS[key] = hits
        # Opportunistic prune so the dict can't grow without bound.
        if len(_AUTH_HITS) > 2048:
            for k in [k for k, v in _AUTH_HITS.items() if not any(now - t < AUTH_WINDOW_SECONDS for t in v)]:
                _AUTH_HITS.pop(k, None)
    return 0


# Serving path prefix. The app is reachable two ways: at the root (LAN NodePort) and
# under /worthly on a shared port 80 behind Traefik. The frontend is built with
# base=/worthly/ so its asset and API URLs are absolute under that prefix, which works
# through Traefik's stripPrefix but would 404 at the root. Stripping it here makes both
# entry points serve the same bundle, and keeps the app working even if the middleware
# is missing or misordered.
URL_PREFIX = os.environ.get("URL_PREFIX", "/worthly")


class _PrefixMiddleware:
    """Strip URL_PREFIX from PATH_INFO before Flask routes the request.

    This has to be WSGI middleware, not @app.before_request: Flask matches the URL when
    the request context is pushed, which is BEFORE before_request handlers run. Rewriting
    PATH_INFO there is too late — the route has already been chosen, so /worthly/api/...
    fell through to the SPA catch-all and returned index.html instead of JSON.

    SCRIPT_NAME is extended rather than discarded so url_for() and redirects still emit
    URLs that include the prefix.
    """

    def __init__(self, wsgi_app, prefix):
        self.wsgi_app = wsgi_app
        self.prefix = (prefix or "").rstrip("/")

    def __call__(self, environ, start_response):
        if self.prefix:
            path = environ.get("PATH_INFO", "")
            if path == self.prefix or path.startswith(self.prefix + "/"):
                environ["SCRIPT_NAME"] = environ.get("SCRIPT_NAME", "") + self.prefix
                environ["PATH_INFO"] = path[len(self.prefix):] or "/"
        return self.wsgi_app(environ, start_response)


# Hostnames this instance answers to. Unset (the default) serves any Host, which is what
# local dev and the LAN NodePort rely on — both arrive by bare IP. Setting it is how you
# stop the public IP from serving the app once a DNS name is in front of it: gunicorn has
# no notion of a canonical hostname, so without this check it answers to the raw public IP
# and to every name anyone points at that IP, and the bare-IP URL stays live forever.
ALLOWED_HOSTS = [h.strip().lower() for h in os.environ.get("ALLOWED_HOSTS", "").split(",") if h.strip()]

# Loopback always passes, so container health checks and `curl localhost:5050` keep working
# even when ALLOWED_HOSTS lists only a public name.
_ALWAYS_ALLOWED_HOSTS = ("localhost", "127.0.0.1", "[::1]")


class _HostAllowlistMiddleware:
    """Reject requests whose Host header isn't in ALLOWED_HOSTS.

    Wrapped outermost so this runs before URL matching and before the prefix rewrite — a
    request for the wrong hostname should never reach a route or the SPA catch-all.

    Only HTTP_HOST is consulted, never X-Forwarded-Host: that header is client-supplied
    and would let a caller spoof its way past the allowlist. Traefik passes the original
    Host through untouched, so the real hostname is already in HTTP_HOST.
    """

    def __init__(self, wsgi_app, allowed):
        self.wsgi_app = wsgi_app
        self.allowed = tuple(allowed)

    def __call__(self, environ, start_response):
        if self.allowed:
            host = environ.get("HTTP_HOST", "").strip().lower()
            # Host usually carries a port (name:5050); the allowlist holds hostnames only.
            if host.startswith("[") and "]" in host:
                host = host[: host.index("]") + 1]  # [::1]:5050 -> [::1]
            else:
                host = host.partition(":")[0]
            if host not in self.allowed and host not in _ALWAYS_ALLOWED_HOSTS:
                # 404, not 403: a wrong-Host caller learns nothing about what runs here.
                body = b'{"error": "not found"}'
                start_response(
                    "404 NOT FOUND",
                    [("Content-Type", "application/json"), ("Content-Length", str(len(body)))],
                )
                return [body]
        return self.wsgi_app(environ, start_response)


app.wsgi_app = _PrefixMiddleware(app.wsgi_app, URL_PREFIX)
app.wsgi_app = _HostAllowlistMiddleware(app.wsgi_app, ALLOWED_HOSTS)


@app.after_request
def _security_headers(resp):
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "DENY")
    resp.headers.setdefault("Referrer-Policy", "no-referrer")
    resp.headers.setdefault("Cross-Origin-Opener-Policy", "same-origin")
    # No inline-script CSP: the built Vite bundle is external, but Recharts injects
    # inline styles, so style-src has to allow them.
    resp.headers.setdefault(
        "Content-Security-Policy",
        # fonts.googleapis.com serves the webfont stylesheet and fonts.gstatic.com the
        # font files; style-src 'self' alone silently killed the app's typeface.
        "default-src 'self'; img-src 'self' data: https:; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' data: https://fonts.gstatic.com; "
        "script-src 'self'; connect-src 'self'; frame-ancestors 'none'; base-uri 'self'",
    )
    if PUBLIC_MODE:
        resp.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return resp

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)
FRONTEND_DIST = os.path.join(BASE_DIR, "frontend", "dist")
DB_FILE = os.path.join(DATA_DIR, "prices.db")
CSV_FILE = os.path.join(DATA_DIR, "all_trades.csv")
EXCEL_FILE = os.path.join(DATA_DIR, "AllTradesReport.xlsx")
SNAPSHOT_FILE = os.path.join(DATA_DIR, "snapshots.json")

# yfinance suffixes for exchanges
EXCHANGE_SUFFIX = {
    "US": "",
    "NASDAQ": "",
    "NYSE": "",
    "ASX": ".AX",
    "LSE": ".L",
    "TSX": ".TO",
}

def db():
    """Create and return a database connection, initializing tables if they don't exist."""
    conn = sqlite3.connect(DB_FILE, timeout=30)
    # WAL lets readers and writers work concurrently instead of blocking on a
    # single file lock — needed now that sync fetches multiple symbols in parallel.
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS dashboard_layout (
            user_id INTEGER PRIMARY KEY,
            widget_order TEXT NOT NULL DEFAULT '',
            widget_visible TEXT NOT NULL DEFAULT '',
            stat_keys TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ibkr_credentials (
            user_id INTEGER PRIMARY KEY,
            flex_token TEXT NOT NULL,
            query_id TEXT NOT NULL,
            last_synced TEXT,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS prices (
            symbol TEXT NOT NULL,
            date TEXT NOT NULL,
            close REAL NOT NULL,
            PRIMARY KEY (symbol, date)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sync_log (
            symbol TEXT PRIMARY KEY,
            last_synced TEXT NOT NULL,
            cached_from TEXT,
            cached_to TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS snapshots (
            date TEXT NOT NULL,
            super REAL NOT NULL,
            cash REAL NOT NULL,
            user_id INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (user_id, date)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            exchange TEXT NOT NULL,
            ticker TEXT NOT NULL,
            name TEXT NOT NULL,
            action TEXT NOT NULL,
            units REAL NOT NULL,
            price REAL NOT NULL,
            currency TEXT NOT NULL,
            brokerage REAL NOT NULL DEFAULT 0,
            brokerage_currency TEXT NOT NULL DEFAULT 'AUD',
            exch_rate REAL NOT NULL DEFAULT 1.0,
            value REAL NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cash_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            institution TEXT,
            type TEXT,
            name TEXT,
            balance REAL NOT NULL DEFAULT 0,
            country TEXT NOT NULL DEFAULT 'AU'
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS super_holdings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            class TEXT,
            allocation_pct REAL NOT NULL,
            country TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS country_overrides (
            symbol TEXT NOT NULL,
            country TEXT NOT NULL,
            user_id INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (user_id, symbol)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS holding_meta (
            symbol TEXT PRIMARY KEY,
            sector TEXT,
            industry TEXT,
            long_name TEXT,
            website TEXT,
            logo_url TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS milestones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            category TEXT NOT NULL,
            value REAL,
            type TEXT DEFAULT 'achievement',
            target_value REAL,
            current_value REAL,
            is_achieved INTEGER DEFAULT 0,
            linked_metric TEXT,
            achieved_date TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS records (
            key TEXT NOT NULL,
            value REAL NOT NULL,
            date TEXT NOT NULL,
            user_id INTEGER NOT NULL DEFAULT 1,
            PRIMARY KEY (user_id, key)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS holding_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            symbols TEXT NOT NULL DEFAULT ''
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS dividends (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            symbol TEXT NOT NULL,
            ticker TEXT NOT NULL,
            exchange TEXT NOT NULL,
            per_share REAL NOT NULL,
            units REAL NOT NULL,
            currency TEXT NOT NULL,
            gross_amount REAL NOT NULL,
            gross_amount_aud REAL NOT NULL,
            franking_pct REAL NOT NULL DEFAULT 0,
            franking_credit_aud REAL NOT NULL DEFAULT 0,
            withholding_tax_pct REAL NOT NULL DEFAULT 0,
            net_amount_aud REAL NOT NULL,
            source TEXT NOT NULL DEFAULT 'manual',
            user_id INTEGER NOT NULL DEFAULT 1,
            UNIQUE(user_id, symbol, date)
        )
    """)
    # Migrate existing sync_log table — tracks sync health, not just successful ranges
    for col, defn in [
        ("last_error", "TEXT"),      # non-null when the most recent sync attempt failed
        ("last_attempt", "TEXT"),    # timestamp of the most recent attempt, success or not
    ]:
        try:
            conn.execute(f"ALTER TABLE sync_log ADD COLUMN {col} {defn}")
        except:
            pass
    # Migrate snapshots to track source (manual vs auto)
    try:
        conn.execute("ALTER TABLE snapshots ADD COLUMN source TEXT DEFAULT 'manual'")
    except:
        pass
    # Migrate existing milestones table
    for col, defn in [
        ("type", "TEXT DEFAULT 'achievement'"),
        ("target_value", "REAL"),
        ("current_value", "REAL"),
        ("is_achieved", "INTEGER DEFAULT 0"),
        ("linked_metric", "TEXT"),
        ("achieved_date", "TEXT"),
        ("linked_metrics", "TEXT"),   # comma-separated list, e.g. "cash,portfolio" — supersedes linked_metric
        ("currency", "TEXT DEFAULT 'AUD'"),  # currency the target/current value is expressed in: AUD or USD
    ]:
        try:
            conn.execute(f"ALTER TABLE milestones ADD COLUMN {col} {defn}")
        except:
            pass

    # Multi-user migration: every table holding a person's own financial data gets a
    # user_id column. Deliberately NOT applied to prices/sync_log/holding_meta — those
    # are the shared market-data cache (GOOG's price is the same for every user), and
    # keeping them global avoids N users redundantly hitting yfinance for the same symbol.
    # Existing single-user rows get user_id=1 so pre-migration data isn't orphaned —
    # whichever account is created first inherits the pre-existing data.
    for table in ["transactions", "cash_accounts", "super_holdings",
                  "milestones", "holding_groups"]:
        try:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN user_id INTEGER NOT NULL DEFAULT 1")
        except:
            pass

    # Allocation widget definitions (name, dimension, custom slices). These used to
    # live only in the browser's localStorage while their order and visibility were
    # saved server-side — so on a new browser the layout still referenced widgets
    # whose definitions no longer existed, and they silently vanished.
    try:
        conn.execute("ALTER TABLE dashboard_layout ADD COLUMN alloc_widgets TEXT")
    except:
        pass

    # Migrate transactions to track where a row came from (manual entry vs IBKR import)
    # and a stable per-broker-execution key so re-syncing IBKR updates existing rows
    # instead of duplicating them. external_id is NULL for every manual row — SQLite
    # treats each NULL as distinct in a unique index, so they never collide with each
    # other or with IBKR rows. Must run after the user_id migration above, since the
    # index is keyed on (user_id, external_id).
    for col, defn in [
        ("source", "TEXT NOT NULL DEFAULT 'manual'"),
        ("external_id", "TEXT"),
    ]:
        try:
            conn.execute(f"ALTER TABLE transactions ADD COLUMN {col} {defn}")
        except:
            pass
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_transactions_user_external_id "
        "ON transactions(user_id, external_id)"
    )

    # AMIT / attribution components of a trust distribution. Australian ETFs (VAS, NDQ,
    # IVV are unit trusts) pass through realised capital gains, tax-deferred amounts and
    # foreign income inside what looks like one "dividend". Treating the whole payment as
    # ordinary income both over-declares income and omits the capital-gains component
    # from the CGT report — on the reference portfolio Sharesight shows $4,289.86 of
    # discounted capital-gain distributions that were invisible here. These figures come
    # from the annual tax statement; no price feed publishes them, so they're entered.
    for col, defn in [
        ("cg_discounted_aud", "REAL NOT NULL DEFAULT 0"),   # discountable capital gain
        ("cg_other_aud", "REAL NOT NULL DEFAULT 0"),        # non-discountable capital gain
        ("tax_deferred_aud", "REAL NOT NULL DEFAULT 0"),    # reduces cost base, not income
        ("foreign_income_aud", "REAL NOT NULL DEFAULT 0"),  # attributed foreign income
        ("foreign_tax_paid_aud", "REAL NOT NULL DEFAULT 0"),# foreign tax offset
        ("pay_date", "TEXT"),                               # `date` is the ex-date
    ]:
        try:
            conn.execute(f"ALTER TABLE dividends ADD COLUMN {col} {defn}")
        except Exception:
            pass

    # Per-user tax settings. The 50% CGT discount is an INDIVIDUAL/trust rate; an SMSF
    # gets 33 1/3% and a company gets none, so it can't stay hardcoded. carry_forward
    # losses are prior-year net capital losses the ATO requires be applied before the
    # discount — there was previously no way to enter them at all.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS tax_settings (
            user_id INTEGER PRIMARY KEY,
            entity_type TEXT NOT NULL DEFAULT 'individual',
            allocation_method TEXT NOT NULL DEFAULT 'fifo',
            updated_at TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS capital_loss_carryforward (
            user_id INTEGER NOT NULL,
            fy_start TEXT NOT NULL,
            amount_aud REAL NOT NULL DEFAULT 0,
            note TEXT,
            PRIMARY KEY (user_id, fy_start)
        )
    """)

    # Locked sale allocations. Parcel selection is otherwise recomputed on every request,
    # so changing the method would retroactively rewrite an already-filed year AND shift
    # the leftover parcels, corrupting every later year's cost base. Once a year is
    # lodged, which parcels each disposal actually consumed becomes a stored fact.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sale_allocations (
            user_id INTEGER NOT NULL,
            sell_txn_id INTEGER NOT NULL,
            buy_txn_id INTEGER NOT NULL,
            units REAL NOT NULL,
            cost_aud REAL NOT NULL,
            method TEXT NOT NULL,
            locked_at TEXT NOT NULL,
            PRIMARY KEY (user_id, sell_txn_id, buy_txn_id)
        )
    """)

    # snapshots originally had `date` alone as PRIMARY KEY — two users both getting a
    # snapshot on the same date (e.g. the monthly auto-snapshot) would collide and
    # overwrite each other. Same fix as dividends/records/country_overrides above.
    existing_snap_sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='snapshots'"
    ).fetchone()
    if existing_snap_sql and "PRIMARY KEY (user_id, date)" not in existing_snap_sql[0]:
        conn.execute("ALTER TABLE snapshots RENAME TO snapshots_old")
        conn.execute("""
            CREATE TABLE snapshots (
                date TEXT NOT NULL, super REAL NOT NULL, cash REAL NOT NULL,
                user_id INTEGER NOT NULL DEFAULT 1, PRIMARY KEY (user_id, date)
            )
        """)
        has_user_id = conn.execute("PRAGMA table_info(snapshots_old)").fetchall()
        if any(col[1] == "user_id" for col in has_user_id):
            conn.execute("INSERT INTO snapshots (date, super, cash, user_id) SELECT date, super, cash, COALESCE(user_id, 1) FROM snapshots_old")
        else:
            conn.execute("INSERT INTO snapshots (date, super, cash, user_id) SELECT date, super, cash, 1 FROM snapshots_old")
        conn.execute("DROP TABLE snapshots_old")

    # country_overrides originally had `symbol` alone as PRIMARY KEY — two users
    # overriding the same symbol's country would collide. Same fix as dividends/records.
    existing_co_sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='country_overrides'"
    ).fetchone()
    if existing_co_sql and "PRIMARY KEY (user_id, symbol)" not in existing_co_sql[0]:
        conn.execute("ALTER TABLE country_overrides RENAME TO country_overrides_old")
        conn.execute("""
            CREATE TABLE country_overrides (
                symbol TEXT NOT NULL, country TEXT NOT NULL,
                user_id INTEGER NOT NULL DEFAULT 1, PRIMARY KEY (user_id, symbol)
            )
        """)
        has_user_id = conn.execute("PRAGMA table_info(country_overrides_old)").fetchall()
        if any(col[1] == "user_id" for col in has_user_id):
            conn.execute("INSERT INTO country_overrides (symbol, country, user_id) SELECT symbol, country, COALESCE(user_id, 1) FROM country_overrides_old")
        else:
            conn.execute("INSERT INTO country_overrides (symbol, country, user_id) SELECT symbol, country, 1 FROM country_overrides_old")
        conn.execute("DROP TABLE country_overrides_old")

    # records originally had `key` alone as PRIMARY KEY (e.g. 'portfolio_high') — two
    # different users would collide and overwrite each other's all-time-high record.
    # Same rebuild-required situation as dividends above.
    existing_records_sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='records'"
    ).fetchone()
    if existing_records_sql and "PRIMARY KEY (user_id, key)" not in existing_records_sql[0]:
        conn.execute("ALTER TABLE records RENAME TO records_old")
        conn.execute("""
            CREATE TABLE records (
                key TEXT NOT NULL, value REAL NOT NULL, date TEXT NOT NULL,
                user_id INTEGER NOT NULL DEFAULT 1, PRIMARY KEY (user_id, key)
            )
        """)
        has_user_id = conn.execute("PRAGMA table_info(records_old)").fetchall()
        if any(col[1] == "user_id" for col in has_user_id):
            conn.execute("INSERT INTO records (key, value, date, user_id) SELECT key, value, date, COALESCE(user_id, 1) FROM records_old")
        else:
            conn.execute("INSERT INTO records (key, value, date, user_id) SELECT key, value, date, 1 FROM records_old")
        conn.execute("DROP TABLE records_old")

    # dividends needs UNIQUE(user_id, symbol, date), not just UNIQUE(symbol, date) — two
    # different users holding the same stock would otherwise collide and overwrite each
    # other's dividend rows. SQLite can't ALTER a UNIQUE constraint in place, so any
    # database created before this fix needs the table rebuilt, not just a column added.
    existing_sql = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='dividends'"
    ).fetchone()
    if existing_sql and "UNIQUE(user_id, symbol, date)" not in existing_sql[0]:
        conn.execute("ALTER TABLE dividends RENAME TO dividends_old")
        conn.execute("""
            CREATE TABLE dividends (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL, symbol TEXT NOT NULL, ticker TEXT NOT NULL, exchange TEXT NOT NULL,
                per_share REAL NOT NULL, units REAL NOT NULL, currency TEXT NOT NULL,
                gross_amount REAL NOT NULL, gross_amount_aud REAL NOT NULL,
                franking_pct REAL NOT NULL DEFAULT 0, franking_credit_aud REAL NOT NULL DEFAULT 0,
                withholding_tax_pct REAL NOT NULL DEFAULT 0, net_amount_aud REAL NOT NULL,
                source TEXT NOT NULL DEFAULT 'manual', user_id INTEGER NOT NULL DEFAULT 1,
                UNIQUE(user_id, symbol, date)
            )
        """)
        old_cols = "id,date,symbol,ticker,exchange,per_share,units,currency,gross_amount," \
                   "gross_amount_aud,franking_pct,franking_credit_aud,withholding_tax_pct,net_amount_aud,source"
        has_user_id = conn.execute("PRAGMA table_info(dividends_old)").fetchall()
        if any(col[1] == "user_id" for col in has_user_id):
            conn.execute(f"INSERT INTO dividends ({old_cols}, user_id) SELECT {old_cols}, COALESCE(user_id, 1) FROM dividends_old")
        else:
            conn.execute(f"INSERT INTO dividends ({old_cols}, user_id) SELECT {old_cols}, 1 FROM dividends_old")
        conn.execute("DROP TABLE dividends_old")
    conn.commit()
    return conn

def current_user_id():
    """Int user id from the JWT — flask-jwt-extended requires string identities,
    so this is the one place that casts back to int for SQL params."""
    return int(get_jwt_identity())

@app.route("/api/register", methods=["POST"])
def register():
    wait = _auth_throttled("register")
    if wait:
        return jsonify({"ok": False, "error": f"Too many attempts. Try again in {wait}s."}), 429

    data = request.json or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    if not email:
        return jsonify({"ok": False, "error": "Username required"}), 400
    if not password:
        return jsonify({"ok": False, "error": "Password required"}), 400

    conn = db()
    # Closed unless explicitly enabled — except on a fresh instance with no users, so a
    # new deployment can still create its first account.
    existing_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if existing_users > 0 and not ALLOW_REGISTRATION:
        conn.close()
        return jsonify({"ok": False,
                        "error": "Registration is closed on this instance."}), 403
    if len(password) < MIN_PASSWORD_LEN:
        conn.close()
        return jsonify({"ok": False,
                        "error": f"Password must be at least {MIN_PASSWORD_LEN} characters."}), 400
    if conn.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone():
        conn.close()
        return jsonify({"ok": False, "error": "An account with this email already exists"}), 409
    cur = conn.execute(
        "INSERT INTO users (email, password_hash, created_at) VALUES (?, ?, ?)",
        (email, generate_password_hash(password), datetime.now().isoformat()),
    )
    user_id = cur.lastrowid
    conn.commit()
    conn.close()
    token = create_access_token(identity=str(user_id))
    return jsonify({"ok": True, "token": token, "user": {"id": user_id, "email": email}})

@app.route("/api/auth/config", methods=["GET"])
def auth_config():
    """Unauthenticated: lets the login screen hide Register when it's closed.

    Deliberately exposes nothing but the flags — no user count, no email, since this is
    reachable pre-login.
    """
    conn = db()
    bootstrap = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0
    conn.close()
    return jsonify({
        "registration_open": bool(ALLOW_REGISTRATION or bootstrap),
        "bootstrap": bootstrap,
        "min_password_len": MIN_PASSWORD_LEN,
    })


@app.route("/api/login", methods=["POST"])
def login():
    wait = _auth_throttled("login")
    if wait:
        return jsonify({"ok": False, "error": f"Too many attempts. Try again in {wait}s."}), 429
    data = request.json or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    conn = db()
    row = conn.execute("SELECT id, password_hash FROM users WHERE email = ?", (email,)).fetchone()
    conn.close()
    # Deliberately identical error for "no such user" and "wrong password" — distinguishing
    # them lets an attacker enumerate registered emails.
    if not row or not check_password_hash(row[1], password):
        return jsonify({"ok": False, "error": "Invalid email or password"}), 401
    token = create_access_token(identity=str(row[0]))
    return jsonify({"ok": True, "token": token, "user": {"id": row[0], "email": email}})

@app.route("/api/me", methods=["GET"])
@jwt_required()
def get_me():
    conn = db()
    row = conn.execute("SELECT id, email, created_at FROM users WHERE id = ?", (current_user_id(),)).fetchone()
    conn.close()
    if not row:
        return jsonify({"ok": False}), 404
    return jsonify({"id": row[0], "email": row[1], "created_at": row[2]})

@app.route("/api/change-password", methods=["POST"])
@jwt_required()
def change_password():
    data = request.json or {}
    current_password = data.get("current_password") or ""
    new_password = data.get("new_password") or ""
    if not new_password:
        return jsonify({"ok": False, "error": "New password required"}), 400

    uid = current_user_id()
    conn = db()
    row = conn.execute("SELECT password_hash FROM users WHERE id = ?", (uid,)).fetchone()
    if not row or not check_password_hash(row[0], current_password):
        conn.close()
        return jsonify({"ok": False, "error": "Current password is incorrect"}), 401

    conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), uid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/dashboard-layout", methods=["GET"])
@jwt_required()
def get_dashboard_layout():
    conn = db()
    row = conn.execute(
        "SELECT widget_order, widget_visible, stat_keys, alloc_widgets FROM dashboard_layout WHERE user_id = ?",
        (current_user_id(),),
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({"widget_order": None, "widget_visible": None,
                        "stat_keys": None, "alloc_widgets": None})
    return jsonify({
        "widget_order": json.loads(row[0]) if row[0] else None,
        "widget_visible": json.loads(row[1]) if row[1] else None,
        "stat_keys": json.loads(row[2]) if row[2] else None,
        "alloc_widgets": json.loads(row[3]) if row[3] else None,
    })

@app.route("/api/dashboard-layout", methods=["POST"])
@jwt_required()
def save_dashboard_layout():
    data = request.json or {}
    conn = db()
    conn.execute(
        "INSERT INTO dashboard_layout (user_id, widget_order, widget_visible, stat_keys, alloc_widgets, updated_at) "
        "VALUES (?, ?, ?, ?, ?, ?) ON CONFLICT(user_id) DO UPDATE SET "
        "widget_order=excluded.widget_order, widget_visible=excluded.widget_visible, "
        "stat_keys=excluded.stat_keys, alloc_widgets=excluded.alloc_widgets, "
        "updated_at=excluded.updated_at",
        (current_user_id(), json.dumps(data.get("widget_order")), json.dumps(data.get("widget_visible")),
         json.dumps(data.get("stat_keys")), json.dumps(data.get("alloc_widgets")),
         datetime.now().isoformat()),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

def yf_symbol(ticker, exchange):
    """Normalize the ticker symbol for yfinance based on exchange."""
    if "=X" in ticker:
        return ticker
    suffix = EXCHANGE_SUFFIX.get((exchange or "").upper(), "")
    return f"{ticker.upper()}{suffix}"

def _pick_monthly_snapshots(all_snapshots):
    """One snapshot per calendar month, anchored to the month's START.

    Returns (snapshots, is_current_month_flags), both date-ascending.

    Snapshots are taken on the 1st, but mid-month ones also get added (corrections, a
    manual top-up). This previously kept the LAST snapshot in each month, which meant a
    month with extras reported the wrong date entirely — July 2026 has snapshots on the
    1st, 7th, 16th, 27th and 30th, so the series showed 2026-07-30 and the step from
    2026-06-01 spanned 59 days while the next step to 2026-08-01 spanned 2. Those two
    then averaged into month-over-month stats as if each were a normal month.

    So: completed months use their FIRST snapshot (the 1st), giving clean ~30-day steps.
    The month still in progress uses its LATEST snapshot, which is the month-to-date
    position and the number worth seeing now.
    """
    current_ym = pd.Timestamp(date.today()).to_period("M")
    by_month: dict = {}
    for s in sorted(all_snapshots, key=lambda r: r[0]):
        ym = pd.Timestamp(s[0]).to_period("M")
        if ym == current_ym:
            by_month[ym] = s          # in-progress month: latest = month-to-date
        elif ym not in by_month:
            by_month[ym] = s          # completed month: first = the 1st
    keys = sorted(by_month)
    # No snapshot yet this month leaves the series ending last month, which the current
    # FY then closes on — it reported 0.0% over 0 months with an NW End weeks out of
    # date, and the longer since the last snapshot the wider the gap to the dashboard.
    # Carry the most recent cash and super forward to today instead: they're the latest
    # readings that exist, and the same pair /api/breakdown feeds the dashboard from.
    # Flagged current, so _compute_monthly_nw_series values the portfolio leg live and
    # the point becomes a true month-to-date rather than a stale one.
    if keys and keys[-1] != current_ym:
        _, latest_super, latest_cash, _ = max(all_snapshots, key=lambda r: r[0])
        by_month[current_ym] = (date.today().isoformat(), latest_super, latest_cash, "carried")
        keys = sorted(by_month)
    return [by_month[k] for k in keys], [k == current_ym for k in keys]


def _split_ratios_by_symbol(txns):
    """Per-symbol [(split_timestamp, ratio)] for retro-adjusting historical unit counts.

    Yahoo's historical closes are ALWAYS split-adjusted — auto_adjust=False removes only
    the dividend adjustment, not the split one, and there is no way to fetch genuine
    pre-split prices. So pairing a raw pre-split unit count with those prices undervalues
    the position by the split ratio: IVV showed 1 unit x $44.00 on 2021-12-30 when the
    holding was really worth the $659.87 actually paid, a 15x understatement.

    The fix is to scale UNITS instead. For any valuation date, multiply the raw unit count
    by every split ratio that happened AFTER that date, which puts units and prices on the
    same post-split basis. `units` on a split row holds the ADDED shares, so the ratio is
    (held_before + added) / held_before.
    """
    held, splits = {}, {}
    for t in sorted(txns, key=lambda x: (x["date"], x.get("id") or 0)):
        sym = yf_symbol(t["ticker"], t["exchange"])
        action = (t.get("action") or "").lower()
        u = held.get(sym, 0.0)
        if action == "buy":
            held[sym] = u + t["units"]
        elif action == "sell":
            held[sym] = max(0.0, u - t["units"])
        elif action == "split":
            if u > 1e-9:
                ratio = (u + t["units"]) / u
                if ratio > 0:
                    splits.setdefault(sym, []).append((pd.Timestamp(t["date"]), ratio))
            held[sym] = u + t["units"]
    return splits


def _split_adj_factor(sym_splits, as_of):
    """Multiplier putting a raw unit count at `as_of` onto today's post-split basis."""
    f = 1.0
    for d, ratio in sym_splits or ():
        if d > as_of:
            f *= ratio
    return f


def get_currency_from_exchange(exchange):
    """Determine instrument currency based on exchange."""
    if (exchange or "").upper() in ["NASDAQ", "NYSE", "US"]:
        return "USD"
    return "AUD"  # Default to AUD (ASX)

def get_historical_exchange_rate(conn, date_str):
    """Retrieve AUD/USD exchange rate for a given date. Try cache, then yfinance."""
    # Try database cache
    row = conn.execute(
        "SELECT close FROM prices WHERE symbol = 'AUDUSD=X' AND date = ?", (date_str,)
    ).fetchone()
    if row:
        return float(row[0])
    
    # Try fetching from yfinance for a short window around the date
    try:
        t_date = pd.Timestamp(date_str)
        t = yf.Ticker("AUDUSD=X")
        hist = t.history(start=t_date - timedelta(days=3), end=t_date + timedelta(days=4))
        if not hist.empty:
            # Find the closest date before or equal to target date
            hist.index = hist.index.tz_localize(None)
            available_dates = hist.index[hist.index <= t_date]
            if len(available_dates) > 0:
                closest_date = available_dates[-1]
            else:
                closest_date = hist.index[0]
            rate = float(hist.loc[closest_date, "Close"])
            
            # Cache it
            conn.execute(
                "INSERT OR REPLACE INTO prices (symbol, date, close) VALUES ('AUDUSD=X', ?, ?)",
                (closest_date.strftime("%Y-%m-%d"), rate)
            )
            conn.commit()
            return rate
    except Exception as e:
        print(f"[warning] Failed to fetch exchange rate for {date_str}: {e}")
        
    # Fallback if everything else fails: return latest cached rate, or 0.65
    fallback_row = conn.execute(
        "SELECT close FROM prices WHERE symbol = 'AUDUSD=X' ORDER BY date DESC LIMIT 1"
    ).fetchone()
    if fallback_row:
        return float(fallback_row[0])
    return 0.65

# ── IBKR Flex Web Service integration ────────────────────────────────────────
IBKR_SEND_URL = "https://ndcdyn.interactivebrokers.com/AccountManagement/FlexWebService/SendRequest"
IBKR_ERROR_HINTS = {
    "1003": "Invalid or expired Flex token.",
    "1004": "Query ID not found or not permissioned for this token.",
    "1005": "Flex query returned no data for the configured date range.",
    "1018": "Statement generation still in progress.",
    "1021": "Too many requests to IBKR — try again shortly.",
}

# Venue/route strings IBKR reports are NOT the same bucket set this app uses
# (ASX/NASDAQ/NYSE/US) — currency is the primary signal (this app is an AU retail
# tracker: AUD-denominated trades are assumed ASX). Venue only refines a USD trade
# into NASDAQ vs NYSE vs the generic "US" bucket when it literally matches.
IBKR_VENUE_MAP = {"NASDAQ": "NASDAQ", "NYSE": "NYSE"}


class IBKRFlexError(Exception):
    def __init__(self, code, message):
        self.code = code
        super().__init__(f"IBKR error {code}: {message}")


def fetch_ibkr_flex_report(token, query_id, send_timeout=15, get_timeout=30,
                            max_poll_attempts=8, poll_delay=5):
    """Fetch a Flex Query report via IBKR's 2-step send/poll Web Service. Retries on
    error 1018 (statement still generating) — the naive single 3s sleep isn't enough
    for larger accounts. Any other error code is terminal."""
    resp1 = requests.get(IBKR_SEND_URL, params={"t": token, "q": query_id, "v": 3}, timeout=send_timeout)
    resp1.raise_for_status()
    root1 = ET.fromstring(resp1.content)
    if root1.findtext("Status") != "Success":
        code = root1.findtext("ErrorCode") or "unknown"
        raise IBKRFlexError(code, root1.findtext("ErrorMessage") or IBKR_ERROR_HINTS.get(code, "Unknown error"))

    ref_code = root1.findtext("ReferenceCode")
    get_url = root1.findtext("Url")

    last_err = None
    for _ in range(max_poll_attempts):
        time.sleep(poll_delay)
        resp2 = requests.get(get_url, params={"t": token, "q": ref_code, "v": 3}, timeout=get_timeout)
        resp2.raise_for_status()
        text = resp2.text.strip()
        if text.startswith("<"):
            try:
                root2 = ET.fromstring(resp2.content)
            except ET.ParseError:
                last_err = IBKRFlexError("unknown", "Malformed response from IBKR")
                continue
            code = root2.findtext("ErrorCode") or "unknown"
            msg = root2.findtext("ErrorMessage") or IBKR_ERROR_HINTS.get(code, "Unknown error")
            if code == "1018":
                last_err = IBKRFlexError(code, msg)
                continue
            raise IBKRFlexError(code, msg)
        return text  # looks like real CSV data

    raise last_err or IBKRFlexError("timeout", "IBKR statement did not finish generating in time")


def _ibkr_date_to_iso(raw_date):
    """IBKR TradeDate is usually YYYYMMDD — normalize to this app's YYYY-MM-DD."""
    digits = (raw_date or "").replace("-", "").strip()
    if len(digits) == 8 and digits.isdigit():
        return f"{digits[:4]}-{digits[4:6]}-{digits[6:]}"
    return raw_date


def map_ibkr_exchange(currency, venue):
    """Map an IBKR currency + raw venue string onto this app's exchange bucket set."""
    venue_u = (venue or "").strip().upper()
    if venue_u in IBKR_VENUE_MAP:
        return IBKR_VENUE_MAP[venue_u]
    if currency == "AUD":
        return "ASX"
    return "US"  # generic bucket — already resolves to USD via get_currency_from_exchange


def parse_and_aggregate_ibkr_trades(raw_payload):
    """Parse a Flex trade-confirmation CSV and merge split/partial fills within the
    same minute into one consolidated trade per (date, minute, symbol, action) — the
    same aggregation the user's own reference script already does. Returns
    (trades, stats) where trades is a list of dicts ready to insert into
    `transactions`, and stats reports what was skipped and why."""
    rows = list(csv.DictReader(io.StringIO(raw_payload.strip())))
    stats = {"skipped_options": 0, "skipped_currency": {}}
    groups = {}

    for row in rows:
        asset_class = row.get("AssetClass") or row.get("AssetCategory") or ""
        if asset_class == "OPT":
            stats["skipped_options"] += 1
            continue
        if asset_class not in ("STK", "ETF"):
            continue

        currency = (row.get("CurrencyPrimary") or row.get("Currency") or "").strip().upper()
        try:
            fx_to_base = float(row.get("FXRateToBase") or 1.0)
        except ValueError:
            fx_to_base = 1.0
        if not currency:
            # Weak fallback if the Flex Query wasn't configured to include a currency
            # field at all — a base-currency (AUD) fill reports FXRateToBase ~= 1.0.
            currency = "AUD" if abs(fx_to_base - 1.0) < 1e-6 else "USD"
        if currency not in ("AUD", "USD"):
            stats["skipped_currency"][currency] = stats["skipped_currency"].get(currency, 0) + 1
            continue

        action_raw = (row.get("Buy/Sell") or row.get("BuySell") or "").strip().upper()
        if action_raw not in ("BUY", "SELL"):
            continue
        action = action_raw.lower()

        trade_date = _ibkr_date_to_iso(row.get("TradeDate", "").strip())
        full_time = row.get("TradeTime") or (row.get("DateTime", "").split(";")[-1]) or "000000"
        minute_bucket = full_time.replace(":", "").strip()[:4]
        symbol = row.get("Symbol", "").strip().upper()
        # ListingExchange (where the stock is actually listed) is a far more stable
        # signal than Exchange (the per-fill execution venue — IBKR's smart order
        # router can fill the same NASDAQ-listed stock through ARCA, EDGX, IEX, etc.
        # on different days), so prefer it when the Flex Query includes it.
        venue = row.get("ListingExchange") or row.get("Exchange") or ""

        try:
            quantity = abs(float(row.get("Quantity") or 0))
            price = float(row.get("TradePrice") or 0)
            commission = abs(float(row.get("IBCommission") or row.get("Commission") or 0))
        except ValueError:
            continue
        if quantity <= 0:
            continue

        key = (trade_date, minute_bucket, symbol, action)
        if key not in groups:
            groups[key] = {
                "currency": currency, "venue": venue, "fx_to_base": fx_to_base,
                "total_quantity": 0.0, "total_native": 0.0, "total_commission": 0.0,
            }
        g = groups[key]
        g["total_quantity"] += quantity
        g["total_native"] += quantity * price
        g["total_commission"] += commission

    trades = []
    for (trade_date, minute_bucket, symbol, action), g in groups.items():
        currency = g["currency"]
        avg_price = g["total_native"] / g["total_quantity"] if g["total_quantity"] else 0.0
        # exch_rate convention matches add_transaction(): price_native / exch_rate = price_aud.
        # Force exactly 1.0 for AUD fills regardless of FXRateToBase — IBKR sometimes
        # reports e.g. 0.9999 even for AUD trades, which would invent phantom FX gain/loss.
        if currency == "AUD":
            exch_rate = 1.0
        else:
            exch_rate = 1.0 / g["fx_to_base"] if g["fx_to_base"] else 1.0
        trades.append({
            "date": trade_date,
            "exchange": map_ibkr_exchange(currency, g["venue"]),
            "ticker": symbol,
            "name": f"{symbol} Stock",
            "action": action,
            "units": round(g["total_quantity"], 6),
            "price": round(avg_price, 6),
            "currency": currency,
            "brokerage": round(g["total_commission"], 2),
            "exch_rate": exch_rate,
            "external_id": f"ibkr:{trade_date}:{minute_bucket}:{symbol}:{action}",
        })

    return trades, stats


def find_ibkr_manual_duplicates(conn, user_id):
    """After an IBKR sync, flag existing MANUAL rows that look like the same trade as
    a newly-imported IBKR row, so the user can review/delete the manual one themselves.
    Never deletes or modifies anything."""
    ibkr_rows = conn.execute(
        "SELECT id, date, ticker, units, price, action FROM transactions "
        "WHERE user_id=? AND source='ibkr'", (user_id,)
    ).fetchall()
    manual_rows = conn.execute(
        "SELECT id, date, ticker, units, price, action FROM transactions "
        "WHERE user_id=? AND source='manual'", (user_id,)
    ).fetchall()

    warnings = []
    for ib_id, ib_date, ib_ticker, ib_units, ib_price, ib_action in ibkr_rows:
        for m_id, m_date, m_ticker, m_units, m_price, m_action in manual_rows:
            if ib_date != m_date:
                continue
            if ib_ticker.upper() != m_ticker.upper():
                continue
            if ib_action.lower() != m_action.lower():
                continue
            if abs(ib_units - m_units) > 0.01:
                continue
            tol = max(0.01, abs(m_price) * 0.01)
            if abs(ib_price - m_price) > tol:
                continue
            warnings.append({
                "ibkr_txn_id": ib_id,
                "manual_txn_id": m_id,
                "ticker": ib_ticker,
                "date": ib_date,
                "units": ib_units,
                "price": ib_price,
            })
    return warnings


def ingest_excel_to_json():
    """Ingest transactions from AllTradesReport.xlsx Combined sheet to JSON file."""
    if not os.path.exists(EXCEL_FILE):
        return []
    
    import openpyxl
    transactions = []
    conn = sqlite3.connect(DB_FILE)
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        if "Combined" not in wb.sheetnames:
            print("[error] 'Combined' sheet not found in AllTradesReport.xlsx")
            return []
            
        ws = wb["Combined"]
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx < 3:  # Skip metadata and headers (header is on row 3)
                continue
            if not row or row[0] is None:
                continue
                
            code = str(row[0]).strip().upper()
            if code in ['TOTAL', 'CODE', 'ALL TRADES']:
                continue
                
            market = str(row[1]).strip().upper() if row[1] else ''
            name = str(row[2]).strip() if row[2] else ''
            dt = str(row[3])[:10] if row[3] else None
            action_type = str(row[4]).strip().lower() if row[4] else ''
            qty = float(row[5]) if row[5] is not None else 0.0
            price = float(row[6]) if row[6] is not None else 0.0
            currency = str(row[7]).strip().upper() if row[7] else ''
            
            brokerage = float(row[9]) if row[9] is not None else 0.0
            brokerage_currency = str(row[10]).strip().upper() if row[10] else ''
            
            exch_rate = float(row[11]) if row[11] is not None else 1.0
            val = float(row[12]) if row[12] is not None else 0.0
            
            # Double-check multi-currency rate calculation
            if currency == "USD" and exch_rate == 1.0:
                exch_rate = get_historical_exchange_rate(conn, dt)
                
            if val == 0.0:
                sign = 1 if action_type == "buy" else -1
                cost_in_instrument = (sign * abs(qty) * price) + brokerage
                if currency == "USD":
                    val = cost_in_instrument / exch_rate
                else:
                    val = cost_in_instrument

            transactions.append({
                "date": dt,
                "exchange": market,
                "ticker": code,
                "name": name,
                "action": action_type,
                "units": abs(qty),
                "price": price,
                "currency": currency,
                "brokerage": brokerage,
                "brokerage_currency": brokerage_currency,
                "exch_rate": exch_rate,
                "value": val
            })
        
        # Sort chronologically
        transactions.sort(key=lambda x: x["date"])
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(transactions, f, indent=2)
        print(f"[backend] Successfully auto-ingested {len(transactions)} trades from Excel.")
    except Exception as e:
        print(f"[error] Failed to ingest Excel: {e}")
    finally:
        conn.close()
        
    return transactions

def ingest_csv_to_json():
    """Ingest transactions from CSV to JSON file."""
    if not os.path.exists(CSV_FILE):
        return []
    
    import csv
    transactions = []
    conn = sqlite3.connect(DB_FILE)
    
    try:
        with open(CSV_FILE, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for idx, row in enumerate(reader):
                code = row["Code"].strip().upper()
                market = row["Market Code"].strip().upper()
                name = row["Name"].strip()
                dt = row["Date"].strip()
                action_type = row["Type"].strip().lower()
                qty = float(row["Qty"])
                price = float(row["Price"])
                currency = row["Instrument Currency"].strip().upper()
                
                brokerage_str = row.get("Brokerage", "0").strip()
                brokerage = float(brokerage_str) if brokerage_str else 0.0
                brokerage_currency = row.get("Brokerage Currency", currency).strip().upper()
                
                exch_rate_str = row.get("Exch. Rate", "").strip()
                exch_rate = float(exch_rate_str) if exch_rate_str else 1.0
                
                val_str = row.get("Value", "").strip()
                val = float(val_str) if val_str else 0.0
                
                # Double-check multi-currency rate calculation
                if currency == "USD" and exch_rate == 1.0:
                    exch_rate = get_historical_exchange_rate(conn, dt)
                    
                if val == 0.0:
                    sign = 1 if action_type == "buy" else -1
                    cost_in_instrument = (sign * abs(qty) * price) + brokerage
                    if currency == "USD":
                        val = cost_in_instrument / exch_rate
                    else:
                        val = cost_in_instrument

                transactions.append({
                    "date": dt,
                    "exchange": market,
                    "ticker": code,
                    "name": name,
                    "action": action_type,
                    "units": abs(qty),
                    "price": price,
                    "currency": currency,
                    "brokerage": brokerage,
                    "brokerage_currency": brokerage_currency,
                    "exch_rate": exch_rate,
                    "value": val
                })
        
        # Sort chronologically
        transactions.sort(key=lambda x: x["date"])
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(transactions, f, indent=2)
        print(f"[backend] Successfully auto-ingested {len(transactions)} trades from CSV.")
    except Exception as e:
        print(f"[error] Failed to ingest CSV: {e}")
    finally:
        conn.close()
        
    return transactions

def load_transactions(user_id):
    """Load one user's transactions from DB."""
    conn = db()
    rows = conn.execute(
        "SELECT id, date, exchange, ticker, name, action, units, price, currency, "
        "brokerage, brokerage_currency, exch_rate, value, source FROM transactions "
        "WHERE user_id = ? ORDER BY date, id", (user_id,)
    ).fetchall()
    conn.close()
    cols = ["id", "date", "exchange", "ticker", "name", "action", "units", "price",
            "currency", "brokerage", "brokerage_currency", "exch_rate", "value", "source"]
    return [dict(zip(cols, row)) for row in rows]

def save_transactions(txns, user_id):
    """Save transactions to DB (full replace for this user only, sorted by date)."""
    txns_sorted = sorted(txns, key=lambda x: x.get("date", ""))
    cols = ("date", "exchange", "ticker", "name", "action", "units", "price",
            "currency", "brokerage", "brokerage_currency", "exch_rate", "value")
    conn = db()
    conn.execute("DELETE FROM transactions WHERE user_id = ?", (user_id,))
    conn.executemany(
        f"INSERT INTO transactions ({','.join(cols)}, user_id) VALUES ({','.join('?' * len(cols))}, ?)",
        [[t.get(c, 0 if c in ("units","price","brokerage","exch_rate","value") else "") for c in cols] + [user_id]
         for t in txns_sorted]
    )
    conn.commit()
    conn.close()

# ─── Cash Accounts, Super Holdings, Country Overrides ───────

def load_cash_accounts(user_id):
    conn = db()
    rows = conn.execute(
        "SELECT institution, type, name, balance, country FROM cash_accounts WHERE user_id = ? ORDER BY id", (user_id,)
    ).fetchall()
    conn.close()
    return [{"institution": r[0], "type": r[1], "name": r[2], "balance": r[3], "country": r[4]} for r in rows]

def save_cash_accounts(accounts, user_id):
    conn = db()
    conn.execute("DELETE FROM cash_accounts WHERE user_id = ?", (user_id,))
    conn.executemany(
        "INSERT INTO cash_accounts (institution, type, name, balance, country, user_id) VALUES (?,?,?,?,?,?)",
        [(a.get("institution",""), a.get("type",""), a.get("name",""), a.get("balance",0), a.get("country","AU"), user_id) for a in accounts]
    )
    conn.commit()
    conn.close()

def load_super_holdings(user_id):
    conn = db()
    rows = conn.execute(
        "SELECT name, class, allocation_pct, country FROM super_holdings WHERE user_id = ? ORDER BY id", (user_id,)
    ).fetchall()
    conn.close()
    return [{"name": r[0], "class": r[1], "allocation_pct": r[2], "country": r[3]} for r in rows]

def save_super_holdings(holdings, user_id):
    conn = db()
    conn.execute("DELETE FROM super_holdings WHERE user_id = ?", (user_id,))
    conn.executemany(
        "INSERT INTO super_holdings (name, class, allocation_pct, country, user_id) VALUES (?,?,?,?,?)",
        [(h.get("name",""), h.get("class",""), h.get("allocation_pct",0), h.get("country",""), user_id) for h in holdings]
    )
    conn.commit()
    conn.close()

def load_country_overrides(user_id):
    conn = db()
    rows = conn.execute("SELECT symbol, country FROM country_overrides WHERE user_id = ?", (user_id,)).fetchall()
    conn.close()
    return {r[0]: r[1] for r in rows}

def save_country_overrides(overrides, user_id):
    conn = db()
    conn.execute("DELETE FROM country_overrides WHERE user_id = ?", (user_id,))
    conn.executemany(
        "INSERT INTO country_overrides (symbol, country, user_id) VALUES (?,?,?)",
        [(k, v, user_id) for k, v in overrides.items()]
    )
    conn.commit()
    conn.close()

def load_holding_meta():
    conn = db()
    rows = conn.execute(
        "SELECT symbol, sector, industry, long_name, website, logo_url FROM holding_meta"
    ).fetchall()
    conn.close()
    return {r[0]: {"sector": r[1], "industry": r[2], "longName": r[3], "website": r[4], "logo_url": r[5]} for r in rows}

def save_holding_meta_one(symbol, entry):
    """Upsert a single symbol's metadata. Safe to call concurrently from multiple
    sync threads — unlike a delete-all-then-reinsert-all pattern, this can't lose
    another thread's in-flight write."""
    conn = db()
    conn.execute(
        "INSERT INTO holding_meta (symbol, sector, industry, long_name, website, logo_url) VALUES (?,?,?,?,?,?) "
        "ON CONFLICT(symbol) DO UPDATE SET sector=excluded.sector, industry=excluded.industry, "
        "long_name=excluded.long_name, website=excluded.website, logo_url=excluded.logo_url",
        (symbol, entry.get("sector", ""), entry.get("industry", ""), entry.get("longName", ""),
         entry.get("website", ""), entry.get("logo_url", "")),
    )
    conn.commit()
    conn.close()

_KNOWN_WEBSITES = {
    "VAS.AX": "vanguard.com.au",
    "VGS.AX": "vanguard.com.au",
    "VTS.AX": "vanguard.com.au",
    "VEU.AX": "vanguard.com.au",
    "VDHG.AX": "vanguard.com.au",
    "VGAD.AX": "vanguard.com.au",
    "VAF.AX": "vanguard.com.au",
    "IVV.AX": "blackrock.com",
    "IAA.AX": "blackrock.com",
    "IEM.AX": "blackrock.com",
    "IJR.AX": "blackrock.com",
    "NDQ.AX": "betashares.com.au",
    "A200.AX": "betashares.com.au",
    "HGBL.AX": "betashares.com.au",
    "DHHF.AX": "betashares.com.au",
    "STW.AX": "ssga.com",
    "SFY.AX": "ssga.com",
    "QOZ.AX": "betashares.com.au",
    "IOZ.AX": "blackrock.com",
    "MVW.AX": "vaneck.com.au",
    "QUAL.AX": "vaneck.com.au",
}

def fetch_holding_meta(ticker, exchange):
    """Fetch sector, industry, logo from yfinance for a ticker. Returns dict or None."""
    ysym = yf_symbol(ticker, exchange)
    conn = db()
    existing = conn.execute(
        "SELECT sector, industry, long_name, website, logo_url FROM holding_meta WHERE symbol = ?", (ysym,)
    ).fetchone()
    conn.close()
    if existing and existing[4]:  # has logo_url
        return {"sector": existing[0], "industry": existing[1], "longName": existing[2], "website": existing[3], "logo_url": existing[4]}

    try:
        t = yf.Ticker(ysym)
        info = t.info
        website = info.get("website", "")
        domain = website.replace("https://", "").replace("http://", "").split("/")[0] if website else ""

        # Fall back to known domain mapping when yfinance has no website
        if not domain and ysym in _KNOWN_WEBSITES:
            domain = _KNOWN_WEBSITES[ysym]
            website = f"https://{domain}"

        logo_url = f"https://www.google.com/s2/favicons?domain={domain}&sz=128" if domain else ""

        entry = {
            "sector": info.get("sector", ""),
            "industry": info.get("industry", ""),
            "longName": info.get("longName", ""),
            "website": website,
            "logo_url": logo_url,
        }
        save_holding_meta_one(ysym, entry)
        return entry
    except Exception as e:
        print(f"[meta] Failed to fetch metadata for {ysym}: {e}")
        return None

def get_holding_country(ticker, exchange, name, user_id):
    """Determine country for a holding: override > exchange heuristic > 'Unknown'."""
    overrides = load_country_overrides(user_id)
    if ticker in overrides:
        return overrides[ticker]
    sym = yf_symbol(ticker, exchange)
    if sym in overrides:
        return overrides[sym]
    # Heuristic: US exchanges → US, ASX → AU, else based on name
    ex_upper = (exchange or "").upper()
    if ex_upper in ("NASDAQ", "NYSE", "US"):
        return "US"
    if ex_upper == "ASX":
        return "AU"
    return "Unknown"

def get_total_cash(user_id):
    """Sum all cash account balances."""
    accounts = load_cash_accounts(user_id)
    return round(sum(a.get("balance", 0) for a in accounts), 2)

# ─── End Config Helpers ────────────────────────────────────

def seed_historical_snapshots():
    """Seed the snapshots table from snapshots.json on first run (when table is empty)."""
    conn = db()
    count = conn.execute("SELECT COUNT(*) FROM snapshots").fetchone()[0]
    conn.close()
    if count > 0:
        return

    if not os.path.exists(SNAPSHOT_FILE):
        return

    try:
        with open(SNAPSHOT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        conn = db()
        for entry in data:
            conn.execute(
                "INSERT OR REPLACE INTO snapshots (date, super, cash) VALUES (?, ?, ?)",
                (entry["date"], entry["super"], entry["cash"]),
            )
        conn.commit()
        conn.close()
        print(f"[backend] Loaded {len(data)} cash/super snapshots from snapshots.json.")
    except Exception as e:
        print(f"[backend] Failed to seed snapshots: {e}")

def _fetch_with_retry(fn, attempts=2, delay=1.5):
    """Run fn() with a couple of retries for transient network/rate-limit errors."""
    last_exc = None
    for i in range(attempts):
        try:
            return fn()
        except Exception as e:
            last_exc = e
            if i < attempts - 1:
                time.sleep(delay)
    raise last_exc

def sync_symbol(symbol, needed_start, needed_end, force=False):
    """Fetch and cache only missing historical daily close prices for symbol.
    Opens its own DB connection so it's safe to call from a worker thread.
    Returns (ok, message)."""
    conn = db()
    try:
        # 15-minute cooldown check
        row = conn.execute(
            "SELECT cached_from, cached_to, last_synced FROM sync_log WHERE symbol = ?", (symbol,)
        ).fetchone()

        now = pd.Timestamp.now()
        if row is not None and not force:
            cached_to = pd.Timestamp(row[1])
            last_synced = pd.Timestamp(row[2])
            if now - last_synced < timedelta(minutes=15) and needed_end < cached_to:
                return True, "Cached recently"

        # Yahoo doesn't publish a daily candle for a day that hasn't closed yet —
        # asking for "today" via the daily endpoint reliably throws a
        # "possibly delisted; no price data found" error, every symbol, every day,
        # until market close. Today's live price comes from the intraday fetch
        # below instead, so daily requests never go past the last fully elapsed day.
        daily_cap = pd.Timestamp(date.today()) - timedelta(days=1)

        ranges_to_fetch = []
        if row is None:
            ranges_to_fetch.append((needed_start, needed_end))
        else:
            cached_from = pd.Timestamp(row[0])
            cached_to = pd.Timestamp(row[1])
            if needed_start < cached_from:
                ranges_to_fetch.append((needed_start, cached_from - timedelta(days=1)))
            if needed_end > cached_to:
                ranges_to_fetch.append((cached_to + timedelta(days=1), needed_end))

        # Unconditionally (re)fetch the most recently closed trading day, even if
        # cached_to already reaches it — a sync that ran while the market was still
        # open only captures an in-progress intraday snapshot for that date, and
        # that snapshot's own date makes cached_to look "caught up" even though the
        # value isn't final. This guarantees it eventually gets overwritten by the
        # real close, instead of getting permanently stuck (which is what was
        # happening: the gap-fill range above collapses to an empty/inverted range
        # once cached_to already equals daily_cap, so that date was never revisited).
        ranges_to_fetch.append((daily_cap, daily_cap))

        errors = []
        for f_start, f_end in ranges_to_fetch:
            f_end = min(f_end, daily_cap)
            if f_start > f_end:
                continue
            try:
                # auto_adjust=False is load-bearing. yfinance defaults to True, which
                # rewrites historical closes DOWN to strip out dividends paid and splits
                # since — great for measuring one stock's total return, wrong for asking
                # "what was this holding worth on that date". It understated VAS at
                # 2021-11-18 as $77.99 against an actual traded $95.16 (-18%), and IVV
                # pre-split as $41.58 against $659.87 (-94%, the 15:1 split), so every
                # historical net worth was too low — worst in the past, converging to
                # correct at today's date, which is exactly the shape of the error.
                #
                # It also fixes splits properly: with raw prices, pre-split dates pair the
                # pre-split unit count with the pre-split price, and post-split dates pair
                # the adjusted count with the adjusted price. Both come out right.
                hist = _fetch_with_retry(
                    lambda: yf.Ticker(symbol).history(
                        start=f_start, end=f_end + timedelta(days=1), auto_adjust=False
                    )
                )
                if hist.empty:
                    continue
                hist.index = hist.index.tz_localize(None)
                rows = [
                    (symbol, d.strftime("%Y-%m-%d"), float(c))
                    for d, c in hist["Close"].items()
                ]
                conn.executemany(
                    "INSERT OR REPLACE INTO prices (symbol, date, close) VALUES (?, ?, ?)",
                    rows,
                )
            except Exception as e:
                msg = str(e)
                # yfinance raises this for any window with zero rows (market holiday,
                # exchange-specific non-trading day, etc.) — genuinely benign, not a
                # sync failure, so don't record it as an error.
                if "possibly delisted" in msg.lower() or "no price data found" in msg.lower():
                    print(f"[sync] No data for {symbol} {f_start.date()} -> {f_end.date()} (holiday/non-trading day, not an error)")
                    continue
                errors.append(msg)
                print(f"[sync] Failed to fetch {symbol} {f_start.date()} -> {f_end.date()}: {e}")

        # Fetch today's intraday price so we have live data even before market close.
        # yfinance daily history() excludes the incomplete current day; intraday 1m
        # data includes it as soon as the first trade prints. Always overwrite (not
        # just insert-if-missing) — the daily-history path above never touches today's
        # date (clamped to daily_cap = yesterday), so nothing else writes this row
        # until tomorrow, meaning it's always safe to refresh it as many times as this
        # runs today without risk of clobbering a finalized close.
        try:
            intraday = _fetch_with_retry(
                lambda: yf.Ticker(symbol).history(period="1d", interval="1m")
            )
            if not intraday.empty:
                intraday.index = intraday.index.tz_localize(None)
                latest = intraday.iloc[-1]
                latest_date = latest.name.strftime("%Y-%m-%d")
                latest_close = float(latest["Close"])
                conn.execute(
                    "INSERT OR REPLACE INTO prices (symbol, date, close) VALUES (?, ?, ?)",
                    (symbol, latest_date, latest_close),
                )
        except Exception as e:
            # Intraday is best-effort — a failure here alone shouldn't mark the whole
            # sync as failed as long as daily history above succeeded.
            print(f"[sync] Failed to fetch intraday for {symbol}: {e}")

        # last_attempt/last_error record every attempt, success or not, so the UI can
        # show accurate sync health even when a symbol has been failing for days.
        # last_synced/cached_from/cached_to only advance on a clean, error-free run —
        # using the ACTUAL date range present in the prices table, not the requested
        # range, so cached_to isn't advanced to today when nothing was actually stored.
        if errors:
            conn.execute(
                "INSERT INTO sync_log (symbol, last_synced, cached_from, cached_to, last_error, last_attempt) "
                "VALUES (?, COALESCE((SELECT last_synced FROM sync_log WHERE symbol = ?), ?), "
                "(SELECT cached_from FROM sync_log WHERE symbol = ?), (SELECT cached_to FROM sync_log WHERE symbol = ?), ?, ?) "
                "ON CONFLICT(symbol) DO UPDATE SET last_error = excluded.last_error, last_attempt = excluded.last_attempt",
                (symbol, symbol, now.isoformat(), symbol, symbol, "; ".join(errors), now.isoformat()),
            )
        else:
            actual_range = conn.execute(
                "SELECT MIN(date), MAX(date) FROM prices WHERE symbol = ?", (symbol,)
            ).fetchone()
            if actual_range and actual_range[0] is not None:
                conn.execute(
                    "INSERT INTO sync_log (symbol, last_synced, cached_from, cached_to, last_error, last_attempt) "
                    "VALUES (?, ?, ?, ?, NULL, ?) "
                    "ON CONFLICT(symbol) DO UPDATE SET last_synced = excluded.last_synced, "
                    "cached_from = excluded.cached_from, cached_to = excluded.cached_to, "
                    "last_error = NULL, last_attempt = excluded.last_attempt",
                    (symbol, now.isoformat(), actual_range[0], actual_range[1], now.isoformat()),
                )
        conn.commit()

        if not ranges_to_fetch:
            return True, "Up to date"
        if errors:
            return False, "; ".join(errors)
        return True, "Synced successfully"
    finally:
        conn.close()

@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_spa(path):
    if path.startswith("api/"):
        return jsonify({"error": "not found"}), 404
    full = os.path.join(FRONTEND_DIST, path)
    if path and os.path.exists(full):
        return send_from_directory(FRONTEND_DIST, path)
    return send_from_directory(FRONTEND_DIST, "index.html")

@app.route("/api/transactions", methods=["GET"])
@jwt_required()
def get_transactions():
    """Retrieve list of transactions enriched with current price and gain/loss."""
    txns = load_transactions(current_user_id())
    if not txns:
        return jsonify([])

    # Get latest prices for all symbols
    conn = db()
    latest_prices = {}
    rows = conn.execute("""
        SELECT symbol, close FROM prices
        WHERE (symbol, date) IN (
            SELECT symbol, MAX(date) FROM prices GROUP BY symbol
        )
    """).fetchall()
    for sym, close in rows:
        latest_prices[sym] = close
    conn.close()

    audusd = latest_prices.get("AUDUSD=X", 0.65)

    meta = load_holding_meta()
    enriched = []
    for t in txns:
        entry = dict(t)
        ysym = yf_symbol(t["ticker"], t["exchange"])
        m = meta.get(ysym, {})
        entry["logo_url"] = m.get("logo_url", "")
        entry["currency_label"] = t.get("currency", "AUD")
        ysym = yf_symbol(t["ticker"], t["exchange"])
        current_price = latest_prices.get(ysym, 0.0)

        if t["action"].lower() == "buy":
            if t.get("currency", "AUD") == "USD":
                current_price_aud = current_price / audusd
                # Split the total gain into a price-driven component (current price vs
                # buy price, FX held at the original rate) and an FX-driven component
                # (same price, FX moved from the original rate to today's). The two
                # sum back to gain_aud exactly.
                exch_rate_then = t.get("exch_rate") or 1.0
                value_at_original_fx = (current_price / exch_rate_then) * t["units"]
                price_gain_aud = value_at_original_fx - abs(t["value"])
                fx_gain_aud = current_price_aud * t["units"] - value_at_original_fx
            else:
                current_price_aud = current_price
                price_gain_aud = current_price_aud * t["units"] - abs(t["value"])
                fx_gain_aud = 0.0
            current_value = current_price_aud * t["units"]
            entry["current_price"] = round(current_price, 4)
            entry["current_value_aud"] = round(current_value, 2)
            entry["gain_aud"] = round(current_value - abs(t["value"]), 2)
            entry["gain_pct"] = round((current_value - abs(t["value"])) / abs(t["value"]) * 100, 2) if t["value"] != 0 else 0.0
            entry["price_gain_aud"] = round(price_gain_aud, 2)
            entry["fx_gain_aud"] = round(fx_gain_aud, 2)
        elif t["action"].lower() == "sell":
            # For sells, "gain" is the realized gain — the sell value minus the proportional cost
            # Simplified: show the sell proceeds as the reference
            entry["current_price"] = round(current_price, 4)
            entry["current_value_aud"] = round(t["value"], 2)
            entry["gain_aud"] = 0.0
            entry["gain_pct"] = 0.0
            entry["price_gain_aud"] = 0.0
            entry["fx_gain_aud"] = 0.0
        else:
            entry["current_price"] = 0.0
            entry["current_value_aud"] = 0.0
            entry["gain_aud"] = 0.0
            entry["gain_pct"] = 0.0
            entry["price_gain_aud"] = 0.0
            entry["fx_gain_aud"] = 0.0

        enriched.append(entry)

    return jsonify(enriched)

@app.route("/api/transactions", methods=["POST"])
@jwt_required()
def add_transaction():
    """Add a new transaction, auto-calculating values and exchange rates."""
    data = request.json
    try:
        date_str = data["date"]
        exchange = data["exchange"].upper().strip()
        ticker = data["ticker"].upper().strip()
        action = data["action"].lower().strip()
        units = float(data["units"])
        price = float(data["price"])
        brokerage = float(data.get("brokerage") or 0.0)
        
        currency = get_currency_from_exchange(exchange)

        conn = db()
        manual_fx = data.get("exch_rate")
        if manual_fx not in (None, "", 0):
            exch_rate = float(manual_fx)
        elif currency == "USD":
            exch_rate = get_historical_exchange_rate(conn, date_str)
        else:
            exch_rate = 1.0
        conn.close()
        
        # Calculate signed AUD value
        if action == "split":
            # Split: units are additional shares, zero cost, zero cash flow
            price = 0.0
            brokerage = 0.0
            aud_value = 0.0
        elif action == "sell":
            # Sell: negative cash flow
            aud_value = (-units * price + brokerage) / exch_rate
        else:
            # Buy: positive cash flow
            aud_value = (units * price + brokerage) / exch_rate
        
        # Determine name if possible, or use ticker
        name = data.get("name", "").strip() or f"{ticker} Stock"

        conn = db()
        conn.execute(
            "INSERT INTO transactions (date, exchange, ticker, name, action, units, price, currency, "
            "brokerage, brokerage_currency, exch_rate, value, user_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (date_str, exchange, ticker, name, action, units, price, currency,
             brokerage, currency, exch_rate, round(aud_value, 2), current_user_id()),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/transactions/<int:idx>", methods=["DELETE"])
@jwt_required()
def delete_transaction(idx):
    """Delete a transaction. idx is treated as a real transaction id when it exists as
    one (the frontend now has stable ids to work with); falls back to the old
    date-ordered positional lookup for any caller still using the legacy convention.
    Scoped to the current user either way — you cannot delete another account's row
    even by guessing its id."""
    uid = current_user_id()
    conn = db()
    exists = conn.execute("SELECT id FROM transactions WHERE id = ? AND user_id = ?", (idx, uid)).fetchone()
    if exists:
        conn.execute("DELETE FROM transactions WHERE id = ? AND user_id = ?", (idx, uid))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    row = conn.execute(
        "SELECT id FROM transactions WHERE user_id = ? ORDER BY date, id LIMIT 1 OFFSET ?", (uid, idx)
    ).fetchone()
    if row:
        conn.execute("DELETE FROM transactions WHERE id = ? AND user_id = ?", (row[0], uid))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    conn.close()
    return jsonify({"ok": False, "error": "Transaction not found"}), 404

@app.route("/api/transactions/<int:txn_id>", methods=["PUT"])
@jwt_required()
def update_transaction(txn_id):
    """Edit an existing transaction by its stable id, recalculating value/exch_rate
    the same way add_transaction does (same manual-FX-override support)."""
    data = request.json
    uid = current_user_id()
    try:
        conn = db()
        exists = conn.execute("SELECT id FROM transactions WHERE id = ? AND user_id = ?", (txn_id, uid)).fetchone()
        if not exists:
            conn.close()
            return jsonify({"ok": False, "error": "Transaction not found"}), 404

        date_str = data["date"]
        exchange = data["exchange"].upper().strip()
        ticker = data["ticker"].upper().strip()
        action = data["action"].lower().strip()
        units = float(data["units"])
        price = float(data["price"])
        brokerage = float(data.get("brokerage") or 0.0)
        currency = get_currency_from_exchange(exchange)

        manual_fx = data.get("exch_rate")
        if manual_fx not in (None, "", 0):
            exch_rate = float(manual_fx)
        elif currency == "USD":
            exch_rate = get_historical_exchange_rate(conn, date_str)
        else:
            exch_rate = 1.0

        if action == "split":
            price = 0.0
            brokerage = 0.0
            aud_value = 0.0
        elif action == "sell":
            aud_value = (-units * price + brokerage) / exch_rate
        else:
            aud_value = (units * price + brokerage) / exch_rate

        name = data.get("name", "").strip() or f"{ticker} Stock"

        conn.execute(
            "UPDATE transactions SET date=?, exchange=?, ticker=?, name=?, action=?, units=?, price=?, "
            "currency=?, brokerage=?, brokerage_currency=?, exch_rate=?, value=? WHERE id=? AND user_id=?",
            (date_str, exchange, ticker, name, action, units, price, currency,
             brokerage, currency, exch_rate, round(aud_value, 2), txn_id, uid),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

def _run_intraday_refresh():
    """Lightweight, frequent refresh of just today's intraday price — this is what
    makes the dashboard feel 'live' between the two full daily syncs, without the
    overhead of re-checking each symbol's entire daily-history backfill or metadata.
    Same DISTINCT-across-all-accounts symbol discovery as _run_sync."""
    conn = db()
    rows = conn.execute(
        "SELECT DISTINCT ticker, exchange FROM transactions"
    ).fetchall()
    conn.close()
    if not rows:
        return []

    symbols = {yf_symbol(t, e) for t, e in rows}
    symbols.add("AUDUSD=X")

    def _refresh_one(symbol):
        try:
            intraday = _fetch_with_retry(
                lambda: yf.Ticker(symbol).history(period="1d", interval="1m")
            )
            if intraday.empty:
                return {"symbol": symbol, "ok": True, "message": "No intraday data (market likely closed)"}
            intraday.index = intraday.index.tz_localize(None)
            latest = intraday.iloc[-1]
            latest_date = latest.name.strftime("%Y-%m-%d")
            latest_close = float(latest["Close"])
            conn = db()
            conn.execute(
                "INSERT OR REPLACE INTO prices (symbol, date, close) VALUES (?, ?, ?)",
                (symbol, latest_date, latest_close),
            )
            conn.commit()
            conn.close()
            return {"symbol": symbol, "ok": True, "message": f"Refreshed @ {latest_close}"}
        except Exception as e:
            return {"symbol": symbol, "ok": False, "message": str(e)}

    results = []
    with ThreadPoolExecutor(max_workers=min(8, len(symbols))) as pool:
        futures = {pool.submit(_refresh_one, sym): sym for sym in symbols}
        for fut in as_completed(futures):
            results.append(fut.result())
    return results

def _run_sync(force=False):
    """Core sync logic — fetch missing prices and metadata for all holdings.

    Each symbol's sync is I/O-bound (yfinance network calls), so they run
    concurrently in a small thread pool instead of one at a time — this is
    what actually makes "Sync All" fast instead of a serial 15-20 call chain.
    sync_symbol and fetch_holding_meta each open/close their own DB connection
    (WAL mode + busy_timeout handle the concurrent writes), so this is safe.
    """
    # Sync price history for the DISTINCT set of tickers held across every user account
    # combined — NOT per-user. yfinance gets hit once per symbol regardless of how many
    # accounts hold it, since the prices table is a shared global cache, not per-user data.
    conn = db()
    rows = conn.execute(
        "SELECT ticker, exchange, MIN(date) as min_date FROM transactions GROUP BY ticker, exchange"
    ).fetchall()
    conn.close()
    if not rows:
        return []

    df = pd.DataFrame(rows, columns=["ticker", "exchange", "date"])
    df["date"] = pd.to_datetime(df["date"])
    df["sym"] = df.apply(lambda r: yf_symbol(r["ticker"], r["exchange"]), axis=1)
    # Two different holders' transactions can map to the same yfinance symbol (e.g. both
    # bought VAS) — collapse to one job per symbol using the earliest date needed across
    # every account that holds it.
    df = df.groupby("sym", as_index=False).agg(date=("date", "min"), ticker=("ticker", "first"), exchange=("exchange", "first"))

    end = pd.Timestamp(date.today())
    # Roll back to Friday if min transaction date falls on a weekend
    min_date = df["date"].min()
    if min_date.weekday() == 5:   # Saturday
        min_date -= timedelta(days=1)
    elif min_date.weekday() == 6: # Sunday
        min_date -= timedelta(days=2)

    def _sync_one(symbol, sym_start):
        if sym_start.weekday() == 5:
            sym_start -= timedelta(days=1)
        elif sym_start.weekday() == 6:
            sym_start -= timedelta(days=2)
        ok, msg = sync_symbol(symbol, sym_start, end, force=force)
        return {"symbol": symbol, "ok": ok, "message": msg}

    jobs = [("AUDUSD=X", min_date)]
    holdings_by_sym = {}
    for sym, grp in df.groupby("sym"):
        jobs.append((sym, grp["date"].min()))
        holdings_by_sym[sym] = grp.iloc[0]

    results = []
    with ThreadPoolExecutor(max_workers=min(6, len(jobs))) as pool:
        futures = {pool.submit(_sync_one, sym, start): sym for sym, start in jobs}
        for fut in as_completed(futures):
            results.append(fut.result())

    # Sort back to a stable order (thread completion order is non-deterministic)
    order = {sym: i for i, (sym, _) in enumerate(jobs)}
    results.sort(key=lambda r: order.get(r["symbol"], 999))

    # Metadata fetches hit yfinance too, so parallelize them the same way.
    with ThreadPoolExecutor(max_workers=min(6, len(holdings_by_sym) or 1)) as pool:
        for sym, row in holdings_by_sym.items():
            pool.submit(fetch_holding_meta, row["ticker"], row["exchange"])

    return results


_sync_jobs: dict = {}  # job_id -> {status, results, started_at, finished_at, error}
_sync_lock = threading.Lock()

def _run_sync_job(job_id: str, force: bool):
    try:
        results = _run_sync(force=force)
        with _sync_lock:
            _sync_jobs[job_id] = {
                "status": "done",
                "results": results,
                "finished_at": datetime.now().isoformat(),
            }
    except Exception as e:
        with _sync_lock:
            _sync_jobs[job_id]["status"] = "error"
            _sync_jobs[job_id]["error"] = str(e)
            _sync_jobs[job_id]["finished_at"] = datetime.now().isoformat()


@app.route("/api/sync", methods=["POST"])
@jwt_required()
def sync_data():
    """Start an async sync — returns a job_id immediately. Poll /api/sync/progress/<job_id>."""
    force = request.args.get("force", "").lower() == "true"
    job_id = uuid.uuid4().hex[:10]
    with _sync_lock:
        _sync_jobs[job_id] = {"status": "running", "started_at": datetime.now().isoformat()}
    threading.Thread(target=_run_sync_job, args=(job_id, force), daemon=True).start()
    return jsonify({"job_id": job_id, "status": "running"})


@app.route("/api/sync/progress/<job_id>", methods=["GET"])
@jwt_required()
def sync_progress(job_id):
    with _sync_lock:
        job = _sync_jobs.get(job_id)
    if not job:
        return jsonify({"status": "not_found"}), 404
    return jsonify(job)


# ── IBKR Flex Web Service sync ───────────────────────────────────────────────
@app.route("/api/ibkr/credentials", methods=["POST"])
@jwt_required()
def save_ibkr_credentials():
    """Save (or replace) this user's Flex token + query ID. The token is only ever
    written here — no GET route on this table ever selects flex_token."""
    data = request.json or {}
    token = (data.get("flex_token") or "").strip()
    query_id = (data.get("query_id") or "").strip()
    if not token or not query_id:
        return jsonify({"ok": False, "error": "Both flex_token and query_id are required"}), 400
    conn = db()
    conn.execute(
        "INSERT INTO ibkr_credentials (user_id, flex_token, query_id, updated_at) VALUES (?,?,?,?) "
        "ON CONFLICT(user_id) DO UPDATE SET flex_token=excluded.flex_token, "
        "query_id=excluded.query_id, updated_at=excluded.updated_at",
        (current_user_id(), token, query_id, datetime.now().isoformat()),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/ibkr/credentials", methods=["GET"])
@jwt_required()
def get_ibkr_credentials():
    conn = db()
    row = conn.execute(
        "SELECT query_id, last_synced FROM ibkr_credentials WHERE user_id=?",
        (current_user_id(),),
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({"configured": False, "query_id": None, "last_synced": None})
    query_id, last_synced = row
    return jsonify({"configured": True, "query_id": query_id, "last_synced": last_synced})


@app.route("/api/ibkr/credentials", methods=["DELETE"])
@jwt_required()
def delete_ibkr_credentials():
    conn = db()
    conn.execute("DELETE FROM ibkr_credentials WHERE user_id=?", (current_user_id(),))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


_ibkr_sync_jobs: dict = {}  # job_id -> {status, results, duplicate_warnings, started_at, finished_at, error}
_ibkr_sync_lock = threading.Lock()

def _run_ibkr_sync_job(job_id, user_id, token, query_id):
    try:
        raw = fetch_ibkr_flex_report(token, query_id)
        trades, stats = parse_and_aggregate_ibkr_trades(raw)

        conn = db()

        # A single real-world ticker must always resolve to the same exchange bucket —
        # the per-ticker transaction list filters by (ticker, exchange) together, so a
        # split would silently hide rows. IBKR reports a different execution venue per
        # fill (NASDAQ vs ARCA vs blank), which fragments one position across exchange
        # values if trusted as-is. Prefer whatever exchange this user already holds the
        # ticker under (manual entry or an earlier sync); for a ticker with no prior
        # record, pin every fill in this batch to the first-resolved exchange for it.
        existing_exchange = {}
        for tk, ex in conn.execute(
            "SELECT ticker, exchange FROM transactions WHERE user_id=? ORDER BY id", (user_id,)
        ).fetchall():
            existing_exchange.setdefault(tk, ex)
        batch_exchange = {}
        for t in trades:
            t["exchange"] = existing_exchange.get(t["ticker"]) or batch_exchange.setdefault(t["ticker"], t["exchange"])

        for t in trades:
            if t["action"] == "sell":
                aud_value = (-t["units"] * t["price"] + t["brokerage"]) / t["exch_rate"]
            else:
                aud_value = (t["units"] * t["price"] + t["brokerage"]) / t["exch_rate"]
            conn.execute(
                "INSERT INTO transactions (date, exchange, ticker, name, action, units, price, "
                "currency, brokerage, brokerage_currency, exch_rate, value, user_id, source, external_id) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
                "ON CONFLICT(user_id, external_id) DO UPDATE SET date=excluded.date, exchange=excluded.exchange, "
                "ticker=excluded.ticker, name=excluded.name, action=excluded.action, units=excluded.units, "
                "price=excluded.price, currency=excluded.currency, brokerage=excluded.brokerage, "
                "brokerage_currency=excluded.brokerage_currency, exch_rate=excluded.exch_rate, value=excluded.value "
                "WHERE transactions.source='ibkr'",
                (t["date"], t["exchange"], t["ticker"], t["name"], t["action"], t["units"], t["price"],
                 t["currency"], t["brokerage"], t["currency"], t["exch_rate"], round(aud_value, 2),
                 user_id, "ibkr", t["external_id"]),
            )
        warnings = find_ibkr_manual_duplicates(conn, user_id)
        conn.execute(
            "UPDATE ibkr_credentials SET last_synced=? WHERE user_id=?",
            (datetime.now().isoformat(), user_id),
        )
        conn.commit()
        conn.close()

        # A ticker this import touched might be brand new (no price history yet) — left
        # alone it'd sit at a $0 value until the next manual price sync. Piggyback a
        # "missing only" price pass right away so newly-imported holdings show a real
        # value immediately. Cheap for tickers already synced recently: sync_symbol has
        # its own per-symbol cooldown, so this only actually hits yfinance for the new
        # ones. Filtered down to just this import's tickers so the reported summary
        # stays relevant instead of dumping the whole tracked-symbol universe.
        price_sync_results = []
        if trades:
            imported_symbols = {yf_symbol(t["ticker"], t["exchange"]) for t in trades}
            price_sync_results = [r for r in _run_sync(force=False) if r["symbol"] in imported_symbols]

        with _ibkr_sync_lock:
            _ibkr_sync_jobs[job_id] = {
                "status": "done",
                "results": {
                    "trades_processed": len(trades),
                    "skipped_options": stats["skipped_options"],
                    "skipped_currency": stats["skipped_currency"],
                },
                "duplicate_warnings": warnings,
                "price_sync_results": price_sync_results,
                "finished_at": datetime.now().isoformat(),
            }
    except Exception as e:
        with _ibkr_sync_lock:
            _ibkr_sync_jobs[job_id] = {
                "status": "error",
                "error": str(e),
                "finished_at": datetime.now().isoformat(),
            }


@app.route("/api/ibkr/sync", methods=["POST"])
@jwt_required()
def sync_ibkr():
    """Start an async IBKR trade import — returns a job_id immediately. Poll
    /api/ibkr/sync/progress/<job_id>. Uses its own job dict (not /api/sync's) since
    the result shape is different and job ids shouldn't collide across sync kinds."""
    uid = current_user_id()
    conn = db()
    row = conn.execute("SELECT flex_token, query_id FROM ibkr_credentials WHERE user_id=?", (uid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"ok": False, "error": "IBKR credentials not configured"}), 400
    token, query_id = row
    job_id = uuid.uuid4().hex[:10]
    with _ibkr_sync_lock:
        _ibkr_sync_jobs[job_id] = {"status": "running", "started_at": datetime.now().isoformat()}
    threading.Thread(target=_run_ibkr_sync_job, args=(job_id, uid, token, query_id), daemon=True).start()
    return jsonify({"job_id": job_id, "status": "running"})


@app.route("/api/ibkr/sync/progress/<job_id>", methods=["GET"])
@jwt_required()
def ibkr_sync_progress(job_id):
    with _ibkr_sync_lock:
        job = _ibkr_sync_jobs.get(job_id)
    if not job:
        return jsonify({"status": "not_found"}), 404
    return jsonify(job)


@app.route("/api/sync-status", methods=["GET"])
@jwt_required()
def get_sync_status():
    """Return full sync status: prices + metadata + health for all cached symbols.
    Deliberately global (not scoped to the caller) — the sync log tracks the shared
    price cache, same reasoning as /api/sync above."""
    conn = db()
    rows = conn.execute("""
        SELECT
            s.symbol,
            s.last_synced,
            s.cached_from,
            s.cached_to,
            s.last_error,
            s.last_attempt,
            (SELECT COUNT(*) FROM prices p WHERE p.symbol = s.symbol) as record_count,
            (SELECT MIN(p.date) FROM prices p WHERE p.symbol = s.symbol) as actual_from,
            (SELECT MAX(p.date) FROM prices p WHERE p.symbol = s.symbol) as actual_to
        FROM sync_log s
        ORDER BY s.symbol
    """).fetchall()
    conn.close()

    # Merge with holding metadata
    meta = load_holding_meta()
    result = []
    for r in rows:
            sym = r[0]
            m = meta.get(sym, {})
            result.append({
                "symbol": sym,
                "last_synced": r[1],
                "cached_from": r[2],
                "cached_to": r[3],
                "last_error": r[4],
                "last_attempt": r[5],
                "record_count": r[6],
                "actual_from": r[7],
                "actual_to": r[8],
                "sector": m.get("sector", ""),
                "industry": m.get("industry", ""),
                "website": m.get("website", ""),
                "logo_url": m.get("logo_url", ""),
                "has_meta": bool(m.get("sector") or m.get("website")),
            })
    return jsonify(result)

AU_FRANKING_TAX_RATE = 0.30  # Australian corporate tax rate used to gross up franked dividends
US_TREATY_WITHHOLDING_PCT = 15.0  # Withholding on US-source dividends under the AU-US tax treaty

def _units_held_on(symbol_txns, as_of):
    """Cumulative units held for a symbol's transactions as of (and including) a date."""
    units = 0.0
    for t in symbol_txns:
        if t["date"] > as_of:
            continue
        if t["action"].lower() == "buy":
            units += t["units"]
        elif t["action"].lower() == "sell":
            units -= t["units"]
        elif t["action"].lower() == "split":
            units += t["units"]
    return units

def _compute_dividend_row(conn, symbol, ticker, exchange, currency, ex_date, per_share, units, source="manual", franking_pct=0.0):
    """Compute the full AUD-converted, franking/withholding-aware dividend record for one payment."""
    gross_amount = per_share * units
    exch_rate = get_historical_exchange_rate(conn, ex_date) if currency == "USD" else 1.0
    gross_amount_aud = gross_amount / exch_rate if currency == "USD" else gross_amount

    withholding_tax_pct = US_TREATY_WITHHOLDING_PCT if currency == "USD" else 0.0
    # Franking credits are a tax offset, not a cash reduction — franked AU dividends
    # are paid in full; withholding is what actually reduces the cash you receive.
    franking_credit_aud = gross_amount_aud * (franking_pct / 100.0) * (AU_FRANKING_TAX_RATE / (1 - AU_FRANKING_TAX_RATE)) if currency != "USD" else 0.0
    net_amount_aud = gross_amount_aud * (1 - withholding_tax_pct / 100.0)

    return {
        "date": ex_date, "symbol": symbol, "ticker": ticker, "exchange": exchange,
        "per_share": per_share, "units": round(units, 4), "currency": currency,
        "gross_amount": round(gross_amount, 2), "gross_amount_aud": round(gross_amount_aud, 2),
        "franking_pct": franking_pct, "franking_credit_aud": round(franking_credit_aud, 2),
        "withholding_tax_pct": withholding_tax_pct, "net_amount_aud": round(net_amount_aud, 2),
        "source": source,
    }

@app.route("/api/dividends", methods=["GET"])
@jwt_required()
def get_dividends():
    conn = db()
    rows = conn.execute("""
        SELECT id, date, symbol, ticker, exchange, per_share, units, currency,
               gross_amount, gross_amount_aud, franking_pct, franking_credit_aud,
               withholding_tax_pct, net_amount_aud, source
        FROM dividends WHERE user_id = ? ORDER BY date DESC
    """, (current_user_id(),)).fetchall()
    conn.close()
    cols = ["id", "date", "symbol", "ticker", "exchange", "per_share", "units", "currency",
            "gross_amount", "gross_amount_aud", "franking_pct", "franking_credit_aud",
            "withholding_tax_pct", "net_amount_aud", "source"]
    return jsonify([dict(zip(cols, r)) for r in rows])

@app.route("/api/dividends", methods=["POST"])
@jwt_required()
def add_dividend():
    """Manually add a dividend payment yfinance didn't catch."""
    data = request.json
    try:
        conn = db()
        row = _compute_dividend_row(
            conn, yf_symbol(data["ticker"], data["exchange"]), data["ticker"], data["exchange"],
            data.get("currency") or get_currency_from_exchange(data["exchange"]),
            data["date"], float(data["per_share"]), float(data["units"]),
            source="manual", franking_pct=float(data.get("franking_pct", 0)),
        )
        conn.execute(
            "INSERT OR REPLACE INTO dividends (date, symbol, ticker, exchange, per_share, units, currency, "
            "gross_amount, gross_amount_aud, franking_pct, franking_credit_aud, withholding_tax_pct, net_amount_aud, source, user_id) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (row["date"], row["symbol"], row["ticker"], row["exchange"], row["per_share"], row["units"],
             row["currency"], row["gross_amount"], row["gross_amount_aud"], row["franking_pct"],
             row["franking_credit_aud"], row["withholding_tax_pct"], row["net_amount_aud"], row["source"],
             current_user_id()),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/dividends/<int:div_id>", methods=["PUT"])
@jwt_required()
def update_dividend(div_id):
    """Mainly used to fill in franking_pct on an auto-fetched row — nobody publishes
    franking data programmatically, so this always has to be a manual edit."""
    data = request.json
    uid = current_user_id()
    try:
        conn = db()
        existing = conn.execute(
            "SELECT symbol, ticker, exchange, per_share, units, currency, date FROM dividends WHERE id = ? AND user_id = ?",
            (div_id, uid),
        ).fetchone()
        if not existing:
            conn.close()
            return jsonify({"ok": False, "error": "Not found"}), 404
        symbol, ticker, exchange, per_share, units, currency, ex_date = existing
        franking_pct = float(data.get("franking_pct", 0))
        row = _compute_dividend_row(conn, symbol, ticker, exchange, currency, ex_date, per_share, units,
                                     source=data.get("source", "manual"), franking_pct=franking_pct)
        conn.execute(
            "UPDATE dividends SET franking_pct=?, franking_credit_aud=?, net_amount_aud=? WHERE id=? AND user_id=?",
            (row["franking_pct"], row["franking_credit_aud"], row["net_amount_aud"], div_id, uid),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/dividends/<int:div_id>", methods=["DELETE"])
@jwt_required()
def delete_dividend(div_id):
    conn = db()
    conn.execute("DELETE FROM dividends WHERE id = ? AND user_id = ?", (div_id, current_user_id()))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/dividends/sync", methods=["POST"])
@jwt_required()
def sync_dividends():
    """Pull dividend history from yfinance for every symbol ever held, sized by the
    units actually held on each ex-dividend date. Franking % is never set by this —
    there's no feed for it — existing rows keep whatever franking_pct was already
    entered; only genuinely new payments are inserted (at franking_pct=0, since a
    reasonable default can't be assumed)."""
    uid = current_user_id()
    txns = load_transactions(uid)
    if not txns:
        return jsonify({"results": [], "message": "No transactions to sync dividends for"})

    by_symbol = {}
    for t in txns:
        sym = yf_symbol(t["ticker"], t["exchange"])
        by_symbol.setdefault(sym, {"ticker": t["ticker"], "exchange": t["exchange"],
                                    "currency": t.get("currency") or get_currency_from_exchange(t["exchange"]),
                                    "txns": []})["txns"].append(t)

    conn = db()
    results = []
    for sym, info in by_symbol.items():
        try:
            divs = _fetch_with_retry(lambda: yf.Ticker(sym).dividends)
            if divs is None or divs.empty:
                results.append({"symbol": sym, "ok": True, "message": "No dividend history"})
                continue
            divs.index = divs.index.tz_localize(None)
            inserted = 0
            for ts, per_share in divs.items():
                ex_date = ts.strftime("%Y-%m-%d")
                units = _units_held_on(info["txns"], ex_date)
                if units <= 1e-5:
                    continue  # wasn't held on this ex-date
                existing = conn.execute(
                    "SELECT franking_pct FROM dividends WHERE symbol = ? AND date = ? AND user_id = ?", (sym, ex_date, uid)
                ).fetchone()
                franking_pct = existing[0] if existing else 0.0
                row = _compute_dividend_row(conn, sym, info["ticker"], info["exchange"], info["currency"],
                                             ex_date, float(per_share), units,
                                             source="yfinance", franking_pct=franking_pct)
                conn.execute(
                    "INSERT INTO dividends (date, symbol, ticker, exchange, per_share, units, currency, "
                    "gross_amount, gross_amount_aud, franking_pct, franking_credit_aud, withholding_tax_pct, net_amount_aud, source, user_id) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
                    # franking_credit_aud must be re-set too: it's derived from the gross
                    # amount, so leaving it out left a stale credit attached to a freshly
                    # recomputed dividend whenever units or FX changed.
                    "ON CONFLICT(user_id, symbol, date) DO UPDATE SET per_share=excluded.per_share, units=excluded.units, "
                    "gross_amount=excluded.gross_amount, gross_amount_aud=excluded.gross_amount_aud, "
                    "franking_credit_aud=excluded.franking_credit_aud, "
                    "net_amount_aud=excluded.net_amount_aud",
                    (row["date"], row["symbol"], row["ticker"], row["exchange"], row["per_share"], row["units"],
                     row["currency"], row["gross_amount"], row["gross_amount_aud"], row["franking_pct"],
                     row["franking_credit_aud"], row["withholding_tax_pct"], row["net_amount_aud"], row["source"], uid),
                )
                inserted += 1
            conn.commit()
            results.append({"symbol": sym, "ok": True, "message": f"{inserted} payment(s) synced"})
        except Exception as e:
            results.append({"symbol": sym, "ok": False, "message": str(e)})
    conn.close()
    return jsonify({"results": results})


def _compute_active_holdings(user_id):
    """Core holdings computation shared by /api/portfolio and anything else that
    needs per-holding value/cost/return (e.g. holding groups) — returns a plain
    list of dicts, not a Flask Response, so it's safe to call from other routes."""
    txns = load_transactions(user_id)
    if not txns:
        return []

    conn = db()
    latest_prices = {}
    prev_prices = {}
    rows = conn.execute("""
        SELECT symbol, close, date
        FROM prices
        WHERE (symbol, date) IN (
            SELECT symbol, MAX(date)
            FROM prices
            GROUP BY symbol
        )
    """).fetchall()
    for sym, close, dt in rows:
        latest_prices[sym] = close

    # Get previous close for daily change calculation
    prev_rows = conn.execute("""
        SELECT p.symbol, p.close
        FROM prices p
        INNER JOIN (
            SELECT symbol, MAX(date) as max_date
            FROM prices
            GROUP BY symbol
        ) latest ON p.symbol = latest.symbol
        WHERE p.date < latest.max_date
        AND (p.symbol, p.date) IN (
            SELECT symbol, MAX(date)
            FROM prices
            WHERE (symbol, date) NOT IN (
                SELECT symbol, MAX(date) FROM prices GROUP BY symbol
            )
            GROUP BY symbol
        )
    """).fetchall()
    for sym, close in prev_rows:
        prev_prices[sym] = close

    # Last-synced timestamps per symbol for staleness display in Holdings tab
    sync_rows = conn.execute("SELECT symbol, last_synced FROM sync_log").fetchall()
    last_synced_map = {r[0]: r[1] for r in sync_rows}
    conn.close()

    audusd = latest_prices.get("AUDUSD=X", 0.65)

    # Sort chronologically to compute cost bases properly
    txns_sorted = sorted(txns, key=lambda x: x["date"])

    holdings = {}
    for t in txns_sorted:
        sym = yf_symbol(t["ticker"], t["exchange"])
        if sym not in holdings:
            holdings[sym] = {
                "ticker": t["ticker"],
                "exchange": t["exchange"],
                "name": t.get("name") or f"{t['ticker']} Stock",
                "currency": t.get("currency") or "AUD",
                "units": 0.0,
                "cost_aud": 0.0,
                "cost_local": 0.0,
                # Every dollar ever put into the ticker, never reduced by a sell.
                # cost_aud shrinks as parcels go out the door, which makes it the
                # right denominator for unrealised-only return and the wrong one for
                # lifetime return — see total_return_pct below.
                "gross_cost_aud": 0.0,
                # Cost in the trade's own currency, brokerage included — so it is
                # exactly cost_aud × the rate of the day, and cost_local_gross/cost_aud
                # gives the cost-weighted average FX the position was built at. The
                # existing cost_local excludes brokerage (it feeds avg_price display),
                # which would put a spurious FX component on AUD holdings.
                "cost_local_gross": 0.0,
                "buys_count": 0,
                "sells_count": 0,
                "realised_aud": 0.0,
                # Realised gain split into what the price did and what the rate did.
                "realised_price_aud": 0.0,
                "realised_fx_aud": 0.0,
            }

        h = holdings[sym]
        qty = t["units"]

        if t["action"].lower() == "buy":
            h["units"] += qty
            h["cost_aud"] += t["value"]
            h["cost_local"] += qty * t["price"]
            h["gross_cost_aud"] += t["value"]
            h["cost_local_gross"] += t["value"] * (t.get("exch_rate") or 1.0)
            h["buys_count"] += 1
        elif t["action"].lower() == "split":
            # Stock split: adjust units without changing cost basis
            # split "units" field stores the additional shares from the split
            h["units"] += qty
        elif t["action"].lower() == "sell":
            if h["units"] > 0:
                avg_cost_before = h["cost_aud"] / h["units"]
                avg_cost_local_before = h["cost_local"] / h["units"]
                avg_local_gross_before = h["cost_local_gross"] / h["units"]
                # Realised gain on this sale, average-cost basis: proceeds less the
                # cost the units carried. `value` is AUD-and-signed (negative for a
                # sell, brokerage already folded in — see add_transaction), so the
                # proceeds are -value. Accumulated here rather than derived later
                # because this is the only place the pre-sale average cost exists.
                h["realised_aud"] += (-t["value"]) - (qty * avg_cost_before)

                # Split that realised gain the same way the unrealised one is split
                # below: value the proceeds at the rate the position was BUILT at, and
                # whatever is left over is the rate having moved since.
                #   price = proceeds at the old rate − what those units cost
                #   fx    = proceeds at today's rate − proceeds at the old rate
                # The two always sum back to the realised gain.
                proceeds_aud = -t["value"]
                qty_cost_aud = qty * avg_cost_before
                fx_avg = (h["cost_local_gross"] / h["cost_aud"]) if h["cost_aud"] > 0 else 1.0
                if fx_avg > 0:
                    proceeds_local = proceeds_aud * (t.get("exch_rate") or 1.0)
                    proceeds_at_buy_fx = proceeds_local / fx_avg
                else:
                    proceeds_at_buy_fx = proceeds_aud
                h["realised_price_aud"] += proceeds_at_buy_fx - qty_cost_aud
                h["realised_fx_aud"] += proceeds_aud - proceeds_at_buy_fx

                h["units"] -= qty
                h["cost_aud"] -= qty * avg_cost_before
                h["cost_local"] -= qty * avg_cost_local_before
                h["cost_local_gross"] -= qty * avg_local_gross_before
            else:
                h["units"] = 0
                h["cost_aud"] = 0
                h["cost_local"] = 0
                h["cost_local_gross"] = 0
            h["sells_count"] += 1

    # Lifetime income per holding, reported in dollars only. Deliberately NOT
    # turned into a percentage against cost_aud: this is income earned across every
    # unit ever held, while cost_aud covers only the units still held. Dividing one
    # by the other overstates badly on trimmed positions (VAS earned ~$29k of income
    # on a position that peaked near 2,600 units but was down to 430 — the ratio
    # reads >100% and means nothing). total_return_pct below does express it as a
    # percentage, but over gross_cost_aud, which is the matching denominator: every
    # unit ever bought, against income from every unit ever held.
    # Franking stays out of both because it's a tax credit, not cash received.
    div_conn = db()
    income_by_symbol, franking_by_symbol = {}, {}
    for sym_, net_, fr_ in div_conn.execute(
        "SELECT symbol, COALESCE(SUM(net_amount_aud), 0), COALESCE(SUM(franking_credit_aud), 0) "
        "FROM dividends WHERE user_id = ? GROUP BY symbol", (user_id,)
    ).fetchall():
        income_by_symbol[sym_] = net_
        franking_by_symbol[sym_] = fr_
    div_conn.close()

    active_holdings = []
    total_portfolio_value = 0.0

    for sym, h in holdings.items():
        if h["units"] <= 1e-5:
            continue

        current_price = latest_prices.get(sym, 0.0)
        prev_price = prev_prices.get(sym, current_price)  # fallback to current if no prev
        if h["currency"] == "USD":
            current_price_aud = current_price / audusd
            prev_price_aud = prev_price / audusd
        else:
            current_price_aud = current_price
            prev_price_aud = prev_price

        value_aud = h["units"] * current_price_aud
        total_portfolio_value += value_aud

        # Daily change
        daily_change = (current_price_aud - prev_price_aud) * h["units"]
        daily_change_pct = ((current_price_aud - prev_price_aud) / prev_price_aud * 100) if prev_price_aud > 0 else 0.0

        avg_price_aud = h["cost_aud"] / h["units"] if h["units"] > 0 else 0.0
        avg_price_local = h["cost_local"] / h["units"] if h["units"] > 0 else 0.0

        return_aud = value_aud - h["cost_aud"]
        return_pct = (return_aud / h["cost_aud"] * 100) if h["cost_aud"] > 0 else 0.0

        # ONE definition of lifetime total return, used everywhere it's reported:
        # unrealised on units still held + realised on parcels already sold + income
        # + franking. Summed over active and closed positions this reconciles exactly
        # to total_return_all in /api/stats — it previously didn't, and the Holdings
        # tab's Total column came up $50,226 short of the dashboard headline because
        # it carried neither the realised leg nor franking.
        # return_pct stays capital-only (and honestly labelled "Capital %") so the
        # unrealised-on-units-held view is still available beside it.
        income_aud = round(income_by_symbol.get(sym, 0.0), 2)
        franking_aud = round(franking_by_symbol.get(sym, 0.0), 2)
        total_return_aud = return_aud + h["realised_aud"] + income_aud + franking_aud

        # The same figure as a percentage, over gross cost — every dollar ever put in.
        # return_pct answers "how are the units I still hold doing"; this answers
        # "was this position worth owning", and they diverge sharply on anything
        # trimmed. AMZN reads +12.5% on units held but +6.5% on the whole position,
        # because two earlier parcels went out at a thinner gain — and it's the
        # second number that lines up with Sharesight.
        # Not annualised, deliberately: Sharesight annualises anything held over a
        # year, which would leave these tiles showing a %/yr that reconciles with no
        # dollar figure anywhere else in the app.
        total_return_pct = (
            total_return_aud / h["gross_cost_aud"] * 100 if h["gross_cost_aud"] > 0 else 0.0
        )

        # Split the unrealised gain into price and rate, the same way the realised leg
        # was split during the replay. fx_avg is the cost-weighted rate the position
        # was built at; value the units at that rate and the remainder is FX.
        fx_avg_now = (h["cost_local_gross"] / h["cost_aud"]) if h["cost_aud"] > 0 else 1.0
        if fx_avg_now > 0:
            value_at_buy_fx = h["units"] * current_price / fx_avg_now
        else:
            value_at_buy_fx = value_aud
        unreal_price = value_at_buy_fx - h["cost_aud"]
        unreal_fx = value_aud - value_at_buy_fx

        # ── What the table reports: the position you actually hold ──────────────
        # Parcels already sold are NOT in these. A row in a list of holdings should
        # describe those holdings; AMZN reading 6.53% because two parcels left at a
        # thinner gain months ago is a portfolio-level fact wearing a holding's row.
        # Capital here is therefore the local price move on the units still held
        # (AMZN: US$227.18 → US$274.44 = 20.81%, the figure a US broker quotes), and
        # Currency is what the rate did to those same units.
        #
        # Income is pro-rated by the share of gross cost still invested — the whole
        # dividend history divided by the units left is the VAS trap (its $28,999 over
        # 430 remaining units reads 126.85%). Note the pro-rated dollars over cost_aud
        # give exactly the same percentage as the full dollars over gross_cost_aud,
        # since cost per unit is constant under average-cost. Only the $ figure moves.
        held_share = (h["cost_aud"] / h["gross_cost_aud"]) if h["gross_cost_aud"] > 0 else 0.0
        income_held_aud = income_aud * held_share
        franking_held_aud = franking_aud * held_share
        holding_return_aud = unreal_price + unreal_fx + income_held_aud + franking_held_aud
        pct_of_cost = (lambda v: v / h["cost_aud"] * 100) if h["cost_aud"] > 0 else (lambda v: 0.0)

        # ── Lifetime, over everything ever invested ─────────────────────────────
        # Kept for the dashboard headline and anything reconciling to it: this is the
        # only basis on which realised gain and the full dividend history are counted.
        capital_gain_aud = unreal_price + h["realised_price_aud"]
        currency_gain_aud = unreal_fx + h["realised_fx_aud"]
        pct_of_gross = (lambda v: v / h["gross_cost_aud"] * 100) if h["gross_cost_aud"] > 0 else (lambda v: 0.0)

        meta = fetch_holding_meta(h["ticker"], h["exchange"])

        active_holdings.append({
            "symbol": sym,
            "ticker": h["ticker"],
            "exchange": h["exchange"],
            "name": h["name"],
            "sector": meta.get("sector", "") if meta else "",
            "industry": meta.get("industry", "") if meta else "",
            "logo_url": meta.get("logo_url", "") if meta else "",
            "currency": h["currency"],
            "units": round(h["units"], 4),
            "cost_aud": round(h["cost_aud"], 2),
            "gross_cost_aud": round(h["gross_cost_aud"], 2),
            "avg_price": round(avg_price_local, 4),
            "avg_price_aud": round(avg_price_aud, 4),
            "current_price": round(current_price, 4),
            "current_price_aud": round(current_price_aud, 4),
            "value_aud": round(value_aud, 2),
            "return_aud": round(return_aud, 2),
            "return_pct": round(return_pct, 2),
            "income_aud": income_aud,
            "franking_aud": franking_aud,
            "realised_aud": round(h["realised_aud"], 2),
            "total_return_aud": round(total_return_aud, 2),
            "total_return_pct": round(total_return_pct, 2),
            # ── Holdings basis: the units you still own. What the tables show. ──
            # These four add to holding_return_aud, all over cost_aud.
            "held_capital_aud": round(unreal_price, 2),
            "held_capital_pct": round(pct_of_cost(unreal_price), 2),
            "held_currency_aud": round(unreal_fx, 2),
            "held_currency_pct": round(pct_of_cost(unreal_fx), 2),
            "held_income_aud": round(income_held_aud, 2),
            "held_income_pct": round(pct_of_cost(income_held_aud), 2),
            "held_franking_aud": round(franking_held_aud, 2),
            "held_franking_pct": round(pct_of_cost(franking_held_aud), 2),
            "holding_return_aud": round(holding_return_aud, 2),
            "holding_return_pct": round(pct_of_cost(holding_return_aud), 2),
            # ── Lifetime basis: everything ever bought. Feeds the headline. ──
            "capital_gain_aud": round(capital_gain_aud, 2),
            "capital_gain_pct": round(pct_of_gross(capital_gain_aud), 2),
            "currency_gain_aud": round(currency_gain_aud, 2),
            "currency_gain_pct": round(pct_of_gross(currency_gain_aud), 2),
            "income_pct": round(pct_of_gross(income_aud), 2),
            "franking_pct_of_cost": round(pct_of_gross(franking_aud), 2),
            "daily_change": round(daily_change, 2),
            "daily_change_pct": round(daily_change_pct, 2),
            "buys_count": h["buys_count"],
            "sells_count": h["sells_count"],
            "last_synced": last_synced_map.get(sym),
        })

    # Portfolio weightings
    for h in active_holdings:
        h["weight"] = round((h["value_aud"] / total_portfolio_value * 100), 2) if total_portfolio_value > 0 else 0.0

    # Sort holdings by value descending
    active_holdings.sort(key=lambda x: x["value_aud"], reverse=True)
    return active_holdings

@app.route("/api/portfolio", methods=["GET"])
@jwt_required()
def get_portfolio():
    """Return detailed analytics of current active holdings.

    Kicks a background price refresh when the cache is stale — the frontend polls this
    every 60s, so simply having the app open keeps prices near-live during trading
    without the scheduler having to poll yfinance around the clock.
    """
    _maybe_refresh_prices_async()
    return jsonify(_compute_active_holdings(current_user_id()))


@app.route("/api/prices/refresh", methods=["POST"])
@jwt_required()
def refresh_prices_now():
    """Force an immediate intraday refresh, ignoring the staleness throttle.

    Synchronous (~1s for 11 symbols) so the caller can re-read and see new numbers.
    Backs a manual 'refresh' control, and works outside market hours too — the last
    close is still the right answer then, it just won't have changed.
    """
    try:
        results = _run_intraday_refresh()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502
    with _LIVE_REFRESH_LOCK:
        _LIVE_REFRESH["at"] = time.time()
    ok = sum(1 for r in results if r.get("ok"))
    return jsonify({
        "ok": True,
        "symbols_refreshed": ok,
        "symbols_total": len(results),
        "market_active": _is_extended_hours(),
        "results": results,
    })

RANGE_DAYS = {"1W": 7, "1M": 30, "3M": 90, "6M": 180, "1Y": 365}
# YTD isn't a fixed number of days, so it's resolved against the calendar instead.
DEFAULT_BENCHMARK = "IVV.AX"


@app.route("/api/performance", methods=["GET"])
@jwt_required()
def get_performance():
    """Cumulative time-weighted return vs a benchmark, both rebased to 0% at the
    start of the window so the two lines answer "what would $1 have done over this
    period" rather than carrying history from before it.

    Params: range = 1W|1M|3M|6M|YTD|1Y|Max, benchmark = a yfinance symbol.
    """
    rng = request.args.get("range", "Max")
    symbol = (request.args.get("benchmark") or DEFAULT_BENCHMARK).strip().upper()
    uid = current_user_id()
    conn = db()
    try:
        twr = _build_twr_series(conn, uid)
        if twr is None or twr.empty:
            return jsonify({"dates": [], "portfolio": [], "benchmark": [],
                            "benchmark_symbol": symbol, "benchmark_available": False,
                            "portfolio_return": 0.0, "benchmark_return": 0.0})

        if rng == "YTD":
            cutoff = pd.Timestamp(date(date.today().year, 1, 1))
        elif rng in RANGE_DAYS:
            cutoff = pd.Timestamp(date.today() - timedelta(days=RANGE_DAYS[rng]))
        else:
            cutoff = twr.index[0]
        window = twr[twr.index >= cutoff]
        if len(window) < 2:
            window = twr

        bench = _benchmark_series(conn, symbol, twr.index)
    finally:
        conn.close()

    # A benchmark the user just typed has no prices yet. Kick a background fetch so it
    # fills in and the next poll draws the line, rather than making them find the Sync
    # tab to make a picker work. Fire-and-forget: this request still answers now.
    if bench is None and symbol:
        def _warm(sym, start):
            try:
                # sync_symbol compares these against cached Timestamps, so they must be
                # Timestamps too — passing ISO strings raises on the '<' comparison.
                sync_symbol(sym, start, pd.Timestamp.now().normalize())
            except Exception as e:
                print(f"[performance] benchmark warm-up failed for {sym}: {e}")
        threading.Thread(target=_warm, args=(symbol, twr.index[0]), daemon=True).start()

    # Rebase: a cumulative series restarted mid-life has to be un-compounded from its
    # own value at the window's start, not merely shifted down by it.
    base = window.iloc[0]
    port = ((1 + window / 100.0) / (1 + base / 100.0) - 1.0) * 100.0

    bench_out, bench_ret = [], None
    if bench is not None:
        bw = bench[bench.index >= window.index[0]]
        if len(bw) >= 2 and bw.iloc[0] > 0:
            bseries = (bw / bw.iloc[0] - 1.0) * 100.0
            bench_out = [round(float(v), 2) for v in bseries]
            bench_ret = round(float(bseries.iloc[-1]), 2)

    return jsonify({
        "dates": [d.strftime("%Y-%m-%d") for d in port.index],
        "portfolio": [round(float(v), 2) for v in port],
        "benchmark": bench_out,
        "benchmark_symbol": symbol,
        "benchmark_available": bool(bench_out),
        "portfolio_return": round(float(port.iloc[-1]), 2),
        "benchmark_return": bench_ret,
        "range": rng,
    })



@app.route("/api/portfolio/range-performance", methods=["GET"])
@jwt_required()
def get_range_performance():
    """Per-holding performance scoped to a time window, so the Holding Performance
    treemap can follow the same 1M/3M/6M/1Y/All selector as the net worth chart.

    'All' is holding_return_pct — the same number as the Portfolio table's Return
    column: price, currency and a pro-rated share of dividends on the units you
    still hold. Parcels already sold are excluded, because the treemap sizes each
    tile by CURRENT value and a percentage covering capital you no longer have
    invested puts two different timeframes in one rectangle. It also has to agree
    with its own range selector — 1M/3M/6M/1Y all report a move on current
    holdings, so 'All' must too.
    Lifetime total return, which does count sold parcels, lives in the dashboard
    headline and the closed-positions table.
    A bounded window can't use cost basis at all (you may not have held the position
    for the whole window), so it reports the price move over that window instead,
    converted to AUD at each end so FX is included the same way it is everywhere
    else in the app.
    """
    rng = request.args.get("range", "All")
    holdings = _compute_active_holdings(current_user_id())

    if rng not in RANGE_DAYS:
        return jsonify([
            {"ticker": h["ticker"], "value_aud": h["value_aud"], "return_pct": h["holding_return_pct"]}
            for h in holdings
        ])

    cutoff = (date.today() - timedelta(days=RANGE_DAYS[rng])).isoformat()
    conn = db()

    # When did each position actually open? A window that reaches back further than
    # you've held something would otherwise report the market's move over a period
    # you had no exposure to — e.g. NVDA bought four weeks ago showing a full year
    # of price action. For those, your return since purchase is the honest number.
    first_buy = dict(conn.execute(
        "SELECT ticker, MIN(date) FROM transactions WHERE user_id = ? AND action = 'buy' "
        "GROUP BY ticker", (current_user_id(),)
    ).fetchall())

    def close_on_or_before(symbol, day):
        row = conn.execute(
            "SELECT close FROM prices WHERE symbol = ? AND date <= ? ORDER BY date DESC LIMIT 1",
            (symbol, day),
        ).fetchone()
        return float(row[0]) if row else None

    def latest_close(symbol):
        row = conn.execute(
            "SELECT close FROM prices WHERE symbol = ? ORDER BY date DESC LIMIT 1", (symbol,)
        ).fetchone()
        return float(row[0]) if row else None

    fx_then = close_on_or_before("AUDUSD=X", cutoff) or 0.65
    fx_now = latest_close("AUDUSD=X") or 0.65

    out = []
    for h in holdings:
        sym = h["symbol"]
        opened = first_buy.get(h["ticker"])
        held_whole_window = opened is not None and opened <= cutoff

        start, end = close_on_or_before(sym, cutoff), latest_close(sym)
        if held_whole_window and start and end and start > 0:
            if h.get("currency") == "USD":
                start_aud, end_aud = start / fx_then, end / fx_now
            else:
                start_aud, end_aud = start, end
            pct = (end_aud - start_aud) / start_aud * 100
        else:
            # Position younger than the window (or no price that far back): report
            # its return since purchase rather than a market move you didn't own.
            pct = h["return_pct"]
        out.append({"ticker": h["ticker"], "value_aud": h["value_aud"], "return_pct": round(pct, 2)})

    conn.close()
    return jsonify(out)

# Extended-hours quotes are a live call per symbol (~0.5s each), so they're cached briefly.
# Without this, every dashboard poll would fan out 7 yfinance requests.
_EXT_HOURS_CACHE = {"at": 0.0, "data": None}
_EXT_HOURS_TTL = 60
# Last payload that actually had quotes. Served (flagged stale) when the price feed is
# unreachable, so a transient DNS or rate-limit blip doesn't blank the card and read as
# "the market has no data" — which is a different and much more alarming thing.
_EXT_HOURS_LAST_GOOD = {"at": 0.0, "data": None}


@app.route("/api/portfolio/extended-hours", methods=["GET"])
@jwt_required()
def extended_hours():
    """Pre-market or after-hours movement of the portfolio, in AUD.

    Which session is reported depends on the time: Yahoo's marketState drives it, and the
    card flips between pre-market and after-hours on its own. Only US-listed holdings are
    covered — yfinance exposes no extended session for the ASX, so ASX holdings are
    excluded from the total rather than silently counted as flat.
    """
    uid = current_user_id()
    now = time.time()
    if _EXT_HOURS_CACHE["data"] and now - _EXT_HOURS_CACHE["at"] < _EXT_HOURS_TTL:
        return jsonify(_EXT_HOURS_CACHE["data"])

    holdings = [h for h in _compute_active_holdings(uid) if h["currency"] == "USD"]
    total_value = sum(h["value_aud"] for h in _compute_active_holdings(uid))
    if not holdings:
        out = {"session": "none", "label": "Extended hours", "total_aud": 0.0, "pct": 0.0,
               "covered": 0, "total_holdings": 0, "movers": [],
               "note": "No US-listed holdings — the ASX has no extended session."}
        _EXT_HOURS_CACHE.update(at=now, data=out)
        return jsonify(out)

    fx_conn = db()
    fx_row = fx_conn.execute(
        "SELECT close FROM prices WHERE symbol = 'AUDUSD=X' ORDER BY date DESC LIMIT 1"
    ).fetchone()
    fx_conn.close()
    audusd = float(fx_row[0]) if fx_row else 0.65

    def quote(h):
        try:
            info = yf.Ticker(h["symbol"]).info
            state = info.get("marketState") or ""
            reg = info.get("regularMarketPrice")
            pre = info.get("preMarketPrice")
            post = info.get("postMarketPrice")
            prev = info.get("regularMarketPreviousClose")
            # Prefer whichever session is actually live; pre-market wins when both exist.
            # During the regular session neither exists, so fall back to today's move
            # against the previous close — otherwise the card is blank 6.5 hours a day.
            if pre:
                ext, which, base = pre, "pre", reg
            elif post:
                ext, which, base = post, "post", reg
            elif reg and prev:
                ext, which, base = reg, "regular", prev
            else:
                return {"failed": True}
            if not base or not ext:
                return {"failed": True}
            return {"ticker": h["ticker"], "state": state, "which": which,
                    "reg": float(base), "ext": float(ext), "units": h["units"]}
        except Exception:
            return {"failed": True}

    with ThreadPoolExecutor(max_workers=min(8, len(holdings))) as pool:
        results = list(pool.map(quote, holdings))
    quotes = [q for q in results if q and not q.get("failed")]
    failures = sum(1 for q in results if not q or q.get("failed"))

    if not quotes:
        # Distinguish "the feed is down" from "there is genuinely nothing to report", and
        # keep showing the last good figure rather than a bare dash.
        last = _EXT_HOURS_LAST_GOOD.get("data")
        if last:
            out = dict(last)
            out["stale"] = True
            out["note"] = (f"Price feed unreachable — showing the last reading from "
                           f"{datetime.fromtimestamp(_EXT_HOURS_LAST_GOOD['at']):%H:%M}.")
        else:
            out = {"session": "unavailable", "label": "Pre / After Market",
                   "total_aud": 0.0, "pct": 0.0, "covered": 0,
                   "total_holdings": len(holdings), "movers": [], "stale": True,
                   "note": f"Couldn't reach the price feed ({failures} of {len(holdings)} "
                           f"symbols failed). This is a connection problem, not a market state."}
        _EXT_HOURS_CACHE.update(at=now, data=out)
        return jsonify(out)

    if any(q["which"] == "pre" for q in quotes):
        which = "pre"
    elif any(q["which"] == "post" for q in quotes):
        which = "post"
    else:
        which = "regular"
    states = {q["state"] for q in quotes}
    total_aud = 0.0
    base_aud = 0.0
    movers = []
    for q in quotes:
        delta_native = q["ext"] - q["reg"]
        delta_aud = delta_native * q["units"] / audusd
        total_aud += delta_aud
        base_aud += q["reg"] * q["units"] / audusd
        movers.append({
            "ticker": q["ticker"],
            "delta_aud": round(delta_aud, 2),
            "pct": round((delta_native / q["reg"] * 100) if q["reg"] else 0.0, 2),
            "price": round(q["ext"], 2),
        })
    movers.sort(key=lambda m: -abs(m["delta_aud"]))

    label = {"pre": "Pre-market", "post": "After hours", "regular": "Market open"}[which]
    # PREPRE means pre-market hasn't opened yet, so a post price is last night's session.
    if which == "post" and "PREPRE" in states:
        label = "After hours (last session)"

    out = {
        "session": which,
        "label": label,
        "as_of": datetime.now().strftime("%H:%M"),
        "stale": False,
        "failures": failures,
        "market_state": sorted(states)[0] if states else "",
        "total_aud": round(total_aud, 2),
        # Percent is against the US sleeve that actually has quotes, not whole net worth.
        "pct": round((total_aud / base_aud * 100) if base_aud else 0.0, 2),
        "us_value_aud": round(base_aud, 2),
        "portfolio_value_aud": round(total_value, 2),
        "covered": len(quotes),
        "total_holdings": len(holdings),
        "movers": movers[:5],
        "note": None,
    }
    _EXT_HOURS_CACHE.update(at=now, data=out)
    _EXT_HOURS_LAST_GOOD.update(at=now, data=out)
    return jsonify(out)


@app.route("/api/portfolio/sparklines", methods=["GET"])
@jwt_required()
def get_portfolio_sparklines():
    """Recent closing prices per held ticker, for the trend column in the holdings
    table. One round trip for the whole portfolio rather than a request per row.
    Keyed by ticker (not yfinance symbol) so the frontend can look up by the same
    key it already renders. Prices are the shared cache, so no user scoping here
    beyond choosing which tickers to return."""
    try:
        days = max(2, min(90, int(request.args.get("days", 30))))
    except (TypeError, ValueError):
        days = 30

    holdings = _compute_active_holdings(current_user_id())
    conn = db()
    series = {}
    for h in holdings:
        rows = conn.execute(
            "SELECT close FROM prices WHERE symbol = ? ORDER BY date DESC LIMIT ?",
            (h["symbol"], days),
        ).fetchall()
        if len(rows) >= 2:
            series[h["ticker"]] = [round(r[0], 4) for r in reversed(rows)]
    conn.close()
    return jsonify(series)

@app.route("/api/holding-groups", methods=["GET"])
@jwt_required()
def get_holding_groups():
    """Return every group with computed aggregates — value, capital gain (unrealised),
    income (net dividends received), currency, capital % and lifetime total return —
    plus a grand total row summing across every grouped holding. Mirrors the
    aggregates in a Sharesight custom-group report."""
    uid = current_user_id()
    conn = db()
    rows = conn.execute("SELECT id, name, symbols FROM holding_groups WHERE user_id = ? ORDER BY id", (uid,)).fetchall()
    conn.close()

    holdings_by_symbol = {h["symbol"]: h for h in _compute_active_holdings(uid)}
    # A group can name a symbol that's since been sold out entirely. Those have no
    # active holding, but they still carry realised gain, income and franking that
    # belong in the group's total — without this they'd silently contribute nothing
    # but their dividends.
    closed_by_symbol = {
        yf_symbol(p["ticker"], p["exchange"]): p for p in _compute_closed_positions(uid)
    }

    div_conn = db()
    div_rows = div_conn.execute("SELECT symbol, net_amount_aud FROM dividends WHERE user_id = ?", (uid,)).fetchall()
    div_conn.close()
    income_by_symbol = {}
    for sym, net in div_rows:
        income_by_symbol[sym] = income_by_symbol.get(sym, 0.0) + net

    def _aggregate(symbols):
        value = capital_gain = cost_basis = income = 0.0
        realised = franking = gross_cost = 0.0
        capital_only = currency_gain = 0.0
        held_capital = held_currency = held_income = held_franking = 0.0
        open_count = 0
        currencies = set()
        for sym in symbols:
            h = holdings_by_symbol.get(sym)
            if h:
                open_count += 1
                value += h["value_aud"]
                capital_gain += h["return_aud"]
                cost_basis += h["cost_aud"]
                realised += h["realised_aud"]
                franking += h["franking_aud"]
                gross_cost += h["gross_cost_aud"]
                capital_only += h["capital_gain_aud"]
                currency_gain += h["currency_gain_aud"]
                held_capital += h["held_capital_aud"]
                held_currency += h["held_currency_aud"]
                held_income += h["held_income_aud"]
                held_franking += h["held_franking_aud"]
                currencies.add(h["currency"])
            else:
                cp = closed_by_symbol.get(sym)
                if cp:
                    realised += cp["realised_aud"]
                    franking += cp["franking_aud"]
                    gross_cost += cp["invested"]
                    capital_only += cp["capital_gain_aud"]
                    currency_gain += cp["currency_gain_aud"]
                    # Nothing added to the held_* legs: a fully-exited position is not
                    # a holding, so it contributes nothing to a holdings-basis figure.
                    # Its history stays in the lifetime legs and the "1 closed" label.
                    currencies.add(cp["currency"])
            income += income_by_symbol.get(sym, 0.0)
        # Capital-only, matching the per-holding return_pct, and labelled "Capital %"
        # in the UI. Income is deliberately kept out of THIS ratio: it spans every
        # unit ever held, while cost_basis covers only units still held, so folding
        # it in here inflates the figure on any group holding a trimmed position.
        return_pct = (capital_gain / cost_basis * 100) if cost_basis > 0 else 0.0
        # The lifetime figure, same four legs as everywhere else in the app, over
        # gross cost — which is the denominator that makes income safe to include.
        total_return_aud = capital_gain + realised + income + franking
        total_return_pct = (total_return_aud / gross_cost * 100) if gross_cost > 0 else 0.0
        currency = currencies.pop() if len(currencies) == 1 else ("Mixed" if len(currencies) > 1 else "AUD")
        return {
            "value": round(value, 2), "capital_gain": round(capital_gain, 2),
            "income": round(income, 2), "currency": currency, "return_pct": round(return_pct, 2),
            # Sharesight-style decomposition, additive to total_return over gross cost.
            # capital_only excludes FX; capital_gain above keeps its original meaning
            # (unrealised, FX included) so nothing already reading it changes.
            "capital_only_aud": round(capital_only, 2),
            "capital_only_pct": round(capital_only / gross_cost * 100, 2) if gross_cost > 0 else 0.0,
            "currency_gain_aud": round(currency_gain, 2),
            "currency_gain_pct": round(currency_gain / gross_cost * 100, 2) if gross_cost > 0 else 0.0,
            "income_pct": round(income / gross_cost * 100, 2) if gross_cost > 0 else 0.0,
            "franking_aud": round(franking, 2),
            # Holdings basis — units still owned, over the cost of those units. What
            # the tables show, so groups and the Portfolio list agree row for row.
            "held_capital_aud": round(held_capital, 2),
            "held_capital_pct": round(held_capital / cost_basis * 100, 2) if cost_basis > 0 else 0.0,
            "held_currency_aud": round(held_currency, 2),
            "held_currency_pct": round(held_currency / cost_basis * 100, 2) if cost_basis > 0 else 0.0,
            "held_income_aud": round(held_income, 2),
            "held_income_pct": round(held_income / cost_basis * 100, 2) if cost_basis > 0 else 0.0,
            "holding_return_aud": round(held_capital + held_currency + held_income + held_franking, 2),
            "holding_return_pct": round(
                (held_capital + held_currency + held_income + held_franking) / cost_basis * 100, 2
            ) if cost_basis > 0 else 0.0,
            "cost_basis": round(cost_basis, 2),
            "realised": round(realised, 2), "franking": round(franking, 2),
            "gross_cost": round(gross_cost, 2),
            # Symbols still held. A group keeps a symbol after it's sold out — the
            # position's realised gain and income stay part of the group's history —
            # so "3 holdings" can mean three names of which one is closed.
            "open_count": open_count,
            "total_return_aud": round(total_return_aud, 2),
            "total_return_pct": round(total_return_pct, 2),
        }

    groups = []
    all_grouped_symbols = []
    for gid, name, symbols_str in rows:
        symbols = [s.strip() for s in symbols_str.split(",") if s.strip()]
        all_grouped_symbols.extend(symbols)
        agg = _aggregate(symbols)
        groups.append({"id": gid, "name": name, "symbols": symbols, **agg})

    grand_total = _aggregate(all_grouped_symbols)
    return jsonify({"groups": groups, "grand_total": grand_total})

@app.route("/api/holding-groups", methods=["POST"])
@jwt_required()
def add_holding_group():
    data = request.json
    try:
        symbols = data.get("symbols", [])
        conn = db()
        conn.execute(
            "INSERT INTO holding_groups (name, symbols, user_id) VALUES (?, ?, ?)",
            (data["name"], ",".join(symbols), current_user_id()),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/holding-groups/<int:group_id>", methods=["PUT"])
@jwt_required()
def update_holding_group(group_id):
    data = request.json
    try:
        symbols = data.get("symbols", [])
        conn = db()
        conn.execute(
            "UPDATE holding_groups SET name = ?, symbols = ? WHERE id = ? AND user_id = ?",
            (data["name"], ",".join(symbols), group_id, current_user_id()),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/holding-groups/<int:group_id>", methods=["DELETE"])
@jwt_required()
def delete_holding_group(group_id):
    conn = db()
    conn.execute("DELETE FROM holding_groups WHERE id = ? AND user_id = ?", (group_id, current_user_id()))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

def _build_daily_market_return_series(conn, user_id):
    """Daily 'market-only' return series — portfolio_value(t) minus cumulative net
    invested capital(t), same construction as the Return line in /api/networth.
    Diffing THIS (not raw portfolio value) is what Daily ATH needs: a buy moves both
    portfolio_value and invested capital up by the same amount on the same day, so it
    nets to zero here — only genuine price movement on existing holdings shows up.
    Diffing raw portfolio_value instead would misread "I bought $80k of APP today" as
    an $80k gain, which is exactly the bug this replaced.
    Returned as a pandas Series indexed by date."""
    txns = load_transactions(user_id)
    if not txns:
        return None
    df = pd.DataFrame(txns)
    df["date"] = pd.to_datetime(df["date"])
    df["sym"] = df.apply(lambda r: yf_symbol(r["ticker"], r["exchange"]), axis=1)
    start = df["date"].min()
    end = pd.Timestamp(date.today())
    all_dates = pd.date_range(start, end, freq="D")

    price_data = {}
    symbols_to_read = list(df["sym"].unique()) + ["AUDUSD=X"]
    for sym in symbols_to_read:
        rows = conn.execute("SELECT date, close FROM prices WHERE symbol = ? ORDER BY date", (sym,)).fetchall()
        if rows:
            s = pd.Series({pd.Timestamp(d): c for d, c in rows}).reindex(all_dates).ffill().bfill()
            s = s.fillna(0.0 if sym != "AUDUSD=X" else 1.0)
        else:
            s = pd.Series(0.0 if sym != "AUDUSD=X" else 1.0, index=all_dates)
        price_data[sym] = s

    fx_rates = price_data["AUDUSD=X"]
    units_changes = pd.DataFrame(0.0, index=all_dates, columns=df["sym"].unique())
    sym_currency = df.groupby("sym")["currency"].first().to_dict()
    cash_flow_changes = pd.Series(0.0, index=all_dates)
    for _, row in df.iterrows():
        sign = -1 if row["action"].lower() == "sell" else 1
        units_changes.loc[row["date"], row["sym"]] += sign * row["units"]
        cash_flow_changes.loc[row["date"]] += row["value"]
    units_df = units_changes.cumsum()
    cash_flow = cash_flow_changes.cumsum()

    portfolio_value = pd.Series(0.0, index=all_dates)
    split_ratios = _split_ratios_by_symbol(txns)
    for sym in df["sym"].unique():
        prices = price_data[sym]
        if sym_currency[sym] == "USD":
            prices = prices / fx_rates
        # Stored closes are split-adjusted; scale raw units by any split that happens
        # after each date so units and prices share one basis.
        factor = pd.Series(1.0, index=all_dates)
        for sd, ratio in split_ratios.get(sym, ()):
            factor.loc[all_dates < sd] *= ratio
        portfolio_value += units_df[sym] * factor * prices
    return portfolio_value - cash_flow

def _build_daily_value_and_flows(conn, user_id):
    """(portfolio_value, net_flow) as daily pandas Series, both AUD.

    Same replay as _build_daily_market_return_series, but returning the two pieces
    rather than their difference — a time-weighted return needs the flow on each day
    kept separate from the value it produced.
    """
    txns = load_transactions(user_id)
    if not txns:
        return None, None
    df = pd.DataFrame(txns)
    df["date"] = pd.to_datetime(df["date"])
    df["sym"] = df.apply(lambda r: yf_symbol(r["ticker"], r["exchange"]), axis=1)
    all_dates = pd.date_range(df["date"].min(), pd.Timestamp(date.today()), freq="D")

    price_data = {}
    for sym in list(df["sym"].unique()) + ["AUDUSD=X"]:
        rows = conn.execute(
            "SELECT date, close FROM prices WHERE symbol = ? ORDER BY date", (sym,)
        ).fetchall()
        if rows:
            s = pd.Series({pd.Timestamp(d): c for d, c in rows}).reindex(all_dates).ffill().bfill()
            s = s.fillna(0.0 if sym != "AUDUSD=X" else 1.0)
        else:
            s = pd.Series(0.0 if sym != "AUDUSD=X" else 1.0, index=all_dates)
        price_data[sym] = s

    fx_rates = price_data["AUDUSD=X"]
    units_changes = pd.DataFrame(0.0, index=all_dates, columns=df["sym"].unique())
    sym_currency = df.groupby("sym")["currency"].first().to_dict()
    flows = pd.Series(0.0, index=all_dates)
    for _, row in df.iterrows():
        sign = -1 if row["action"].lower() == "sell" else 1
        units_changes.loc[row["date"], row["sym"]] += sign * row["units"]
        flows.loc[row["date"]] += row["value"]
    units_df = units_changes.cumsum()

    value = pd.Series(0.0, index=all_dates)
    split_ratios = _split_ratios_by_symbol(txns)
    for sym in df["sym"].unique():
        prices = price_data[sym]
        if sym_currency[sym] == "USD":
            prices = prices / fx_rates
        factor = pd.Series(1.0, index=all_dates)
        for sd, ratio in split_ratios.get(sym, ()):
            factor.loc[all_dates < sd] *= ratio
        value += units_df[sym] * factor * prices
    return value, flows


def _build_twr_series(conn, user_id):
    """Cumulative TIME-WEIGHTED return, as a daily pandas Series of percentages.

    Each day's return is (value_end − flow) / value_start − 1, chained. Subtracting
    the day's flow before dividing is the whole point: money you paid in is not
    performance, and without that step every deposit reads as an instant gain.

    This is deliberately NOT the money-weighted figure the stat cards show. MWR
    answers "how did MY money do" and is the honest measure of a personal result,
    but it moves with the TIMING of contributions, so it cannot be compared against
    an index — which has no contributions. Only TWR can sit on the same axis as a
    benchmark line, which is the only reason this exists.
    """
    value, flows = _build_daily_value_and_flows(conn, user_id)
    if value is None or len(value) < 2:
        return None
    prev = value.shift(1)
    daily = (value - flows) / prev - 1.0
    # Before the first buy prev is 0 (or the position is empty) — no capital at work,
    # so no return to record. inf/NaN here would poison the whole cumulative product.
    daily = daily.replace([float("inf"), float("-inf")], 0.0).fillna(0.0)
    daily[prev <= 0] = 0.0
    return ((1.0 + daily).cumprod() - 1.0) * 100.0


def _benchmark_series(conn, symbol, index):
    """Cumulative % return for a benchmark symbol over `index`, or None if unpriced.

    AUD-converted when the symbol is USD-quoted, so it sits on the same axis as a
    portfolio measured in AUD — an Australian investor holding SPY really does earn
    the index return PLUS whatever the currency did, and hiding that would make the
    comparison flattering rather than true. ASX symbols are already AUD.
    """
    rows = conn.execute(
        "SELECT date, close FROM prices WHERE symbol = ? ORDER BY date", (symbol,)
    ).fetchall()
    if len(rows) < 2:
        return None
    s = pd.Series({pd.Timestamp(d): c for d, c in rows}).reindex(index).ffill().bfill()
    if s.isna().all():
        return None
    if not symbol.endswith((".AX", ".L", ".TO")):
        fx = conn.execute(
            "SELECT date, close FROM prices WHERE symbol = 'AUDUSD=X' ORDER BY date"
        ).fetchall()
        if fx:
            fxs = pd.Series({pd.Timestamp(d): c for d, c in fx}).reindex(index).ffill().bfill()
            s = s / fxs
    return s


def _calc_realised_gain(txns: list) -> float:
    """Lifetime realised gain in AUD, average-cost basis, across every symbol.

    Deliberately a standalone replay rather than a sum over _compute_active_holdings'
    output: a position sold down to zero is filtered out of that list, so summing it
    there would lose the gain the moment a holding is fully exited. `value` is AUD and
    signed — positive on a buy, negative on a sell, brokerage already folded in (see
    add_transaction) — so sale proceeds are -value.
    """
    units, cost, realised = {}, {}, 0.0
    for t in sorted(txns, key=lambda x: (x["date"], x.get("id") or 0)):
        sym = f"{t['ticker']}:{t.get('exchange', '')}"
        action = (t.get("action") or "").lower()
        u = units.get(sym, 0.0)
        if action == "buy":
            units[sym] = u + t["units"]
            cost[sym] = cost.get(sym, 0.0) + t["value"]
        elif action == "split":
            units[sym] = u + t["units"]
        elif action == "sell":
            if u > 0:
                avg = cost.get(sym, 0.0) / u
                realised += (-t["value"]) - (t["units"] * avg)
                units[sym] = u - t["units"]
                cost[sym] = cost.get(sym, 0.0) - t["units"] * avg
            else:
                units[sym], cost[sym] = 0.0, 0.0
    return round(realised, 2)


def _compute_closed_positions(user_id):
    """Positions sold down to zero — invisible in the holdings list, but not gone.

    _compute_active_holdings skips anything with no units left, which is right for a
    "what do I own" view but means a fully exited position disappears completely. Selling
    out of VAS removed a 60-trade, five-year holding with $50,056.86 of realised gain and
    $28,999.22 of dividends from every screen, even though the money still counts in the
    portfolio totals.

    Average-cost basis, matching _calc_realised_gain, so the per-position realised figures
    here sum to the realised_gain reported in /api/stats.
    """
    txns = load_transactions(user_id)
    if not txns:
        return []

    conn = db()
    income = dict(conn.execute(
        "SELECT ticker, COALESCE(SUM(net_amount_aud), 0) FROM dividends WHERE user_id = ? GROUP BY ticker",
        (user_id,)
    ).fetchall())
    franking = dict(conn.execute(
        "SELECT ticker, COALESCE(SUM(franking_credit_aud), 0) FROM dividends WHERE user_id = ? GROUP BY ticker",
        (user_id,)
    ).fetchall())
    conn.close()

    pos = {}
    for t in sorted(txns, key=lambda x: (x["date"], x.get("id") or 0)):
        key = t["ticker"]
        p = pos.setdefault(key, {
            "ticker": key, "exchange": t.get("exchange", ""),
            "name": t.get("name") or key, "currency": t.get("currency") or "AUD",
            "units": 0.0, "cost": 0.0, "invested": 0.0, "proceeds": 0.0,
            "realised": 0.0, "buys": 0, "sells": 0,
            # Same price/rate split as _compute_active_holdings, so a group holding a
            # closed position still foots capital + currency + income + franking.
            "cost_local": 0.0, "realised_price": 0.0, "realised_fx": 0.0,
            "first_date": t["date"], "last_date": t["date"],
        })
        p["last_date"] = t["date"]
        action = (t.get("action") or "").lower()
        if action == "buy":
            p["units"] += t["units"]
            p["cost"] += t["value"]
            p["invested"] += t["value"]
            p["cost_local"] += t["value"] * (t.get("exch_rate") or 1.0)
            p["buys"] += 1
        elif action == "split":
            p["units"] += t["units"]
        elif action == "sell":
            if p["units"] > 0:
                avg = p["cost"] / p["units"]
                avg_local = p["cost_local"] / p["units"]
                proceeds = -t["value"]
                p["realised"] += proceeds - (t["units"] * avg)
                fx_avg = (p["cost_local"] / p["cost"]) if p["cost"] > 0 else 1.0
                at_buy_fx = (proceeds * (t.get("exch_rate") or 1.0) / fx_avg) if fx_avg > 0 else proceeds
                p["realised_price"] += at_buy_fx - (t["units"] * avg)
                p["realised_fx"] += proceeds - at_buy_fx
                p["proceeds"] += proceeds
                p["cost"] -= t["units"] * avg
                p["cost_local"] -= t["units"] * avg_local
                p["units"] -= t["units"]
            p["sells"] += 1

    out = []
    for p in pos.values():
        # Closed means nothing left AND it was actually sold — not merely never bought.
        if p["units"] > 1e-5 or p["sells"] == 0:
            continue
        inc = round(float(income.get(p["ticker"], 0.0)), 2)
        frk = round(float(franking.get(p["ticker"], 0.0)), 2)
        realised = round(p["realised"], 2)
        out.append({
            "ticker": p["ticker"],
            "exchange": p["exchange"],
            "name": p["name"],
            "currency": p["currency"],
            "invested": round(p["invested"], 2),
            "proceeds": round(p["proceeds"], 2),
            "realised_aud": realised,
            "capital_gain_aud": round(p["realised_price"], 2),
            "currency_gain_aud": round(p["realised_fx"], 2),
            "income_aud": inc,
            "franking_aud": frk,
            "total_return_aud": round(realised + inc + frk, 2),
            # Must match total_return_aud, not just the realised leg. Quoting realised-only
            # next to a Total column that includes income read as a bad result: VAS showed
            # +20.14% beside $79,056.08, silently dropping $28,999.22 of dividends — the
            # honest figure is 31.81%. Franking is now inside both, matching active
            # holdings and the dashboard headline: one definition of total return across
            # the app, so per-position figures sum to it. It stays broken out as
            # franking_aud so the tax-credit portion is still visible.
            "return_pct": round((realised + inc + frk) / p["invested"] * 100, 2) if p["invested"] > 0 else None,
            "buys_count": p["buys"],
            "sells_count": p["sells"],
            "first_date": p["first_date"],
            "closed_date": p["last_date"],
            "held_days": (date.fromisoformat(p["last_date"][:10]) - date.fromisoformat(p["first_date"][:10])).days,
        })
    out.sort(key=lambda x: x["closed_date"], reverse=True)
    return out


@app.route("/api/portfolio/closed", methods=["GET"])
@jwt_required()
def get_closed_positions():
    """Fully-exited positions, newest close first."""
    rows = _compute_closed_positions(current_user_id())
    return jsonify({
        "positions": rows,
        "total_realised": round(sum(r["realised_aud"] for r in rows), 2),
        "total_income": round(sum(r["income_aud"] for r in rows), 2),
        "total_return": round(sum(r["total_return_aud"] for r in rows), 2),
    })


def _xirr(flows, tol=1e-7, max_iter=300):
    """Annualised money-weighted rate from dated (date, amount) flows, or None.

    Bisection rather than Newton: no derivative, no divergence, and a guaranteed
    answer whenever the NPV actually brackets zero. Returns None when it doesn't
    (all-positive or all-negative flows), so callers must handle the null instead
    of being handed a fabricated rate.
    """
    if len(flows) < 2:
        return None
    t0 = min(d for d, _ in flows)

    def npv(r):
        return sum(a / (1.0 + r) ** ((d - t0).days / 365.0) for d, a in flows)

    lo, hi = -0.9999, 10.0
    try:
        n_lo, n_hi = npv(lo), npv(hi)
    except (OverflowError, ZeroDivisionError):
        return None
    if n_lo * n_hi > 0:
        return None
    for _ in range(max_iter):
        mid = (lo + hi) / 2
        n_mid = npv(mid)
        if abs(n_mid) < tol:
            return mid
        if n_lo * n_mid <= 0:
            hi = mid
        else:
            lo, n_lo = mid, n_mid
    return (lo + hi) / 2


def _calc_mwr(txns: list, div_rows: list, total_value: float):
    """Money-weighted (dollar-weighted) return p.a. over the real dated cash flows.

    This replaces the old CAGR, which computed (total_value / cost_of_units_still_held)
    ** (1/years_since_first_buy) and was wrong on three counts: it ignored every dollar
    returned by a sale, ignored all dividend income, and treated capital deployed last
    month as though it had been invested since the first ever buy. On the reference
    portfolio it reported +4.07% p.a. where the money-weighted rate is +15.96%.

    Returns a dict. `annualised` is False for holding periods under a year, where the
    figure is a plain cumulative return instead — raising a short period to 1/years
    explodes (a few percent over two days annualises into the thousands).
    """
    buys_sells = [t for t in txns if (t.get("action") or "").lower() in ("buy", "sell")]
    if not buys_sells or total_value <= 0:
        return {"pct": None, "years": 0.0, "annualised": False, "pct_ex_income": None}

    # value is AUD and signed: +buy, -sell. A cash flow is its negation — money
    # leaving the pocket is negative, money coming back is positive.
    cap = [(date.fromisoformat(t["date"][:10]), -t["value"]) for t in buys_sells]
    income = [(date.fromisoformat(r["date"][:10]), r["net_amount_aud"]) for r in div_rows]
    today = date.today()
    first = min(d for d, _ in cap)
    years = (today - first).days / 365.25

    terminal = [(today, total_value)]
    r_all = _xirr(cap + income + terminal)
    r_ex = _xirr(cap + terminal)

    if years < 1:
        # Too short to annualise; report cumulative money in vs money out instead.
        out = sum(-a for _, a in cap if a < 0)
        back = sum(a for _, a in cap if a > 0) + sum(a for _, a in income) + total_value
        cum = ((back / out - 1) * 100) if out > 0 else None
        return {"pct": round(cum, 2) if cum is not None else None,
                "years": round(years, 2), "annualised": False, "pct_ex_income": None}

    return {
        "pct": round(r_all * 100, 2) if r_all is not None else None,
        "years": round(years, 2),
        "annualised": True,
        "pct_ex_income": round(r_ex * 100, 2) if r_ex is not None else None,
    }


def _calc_cagr(txns: list, total_value: float, total_cost: float):
    """Return (pct, years, annualised) from first buy date to today.

    Only annualises once a full year has elapsed. Below that, `(v/c) ** (1/years)`
    blows up as years approaches zero — a few percent over a couple of days
    annualises into the thousands — so the plain cumulative return is reported
    instead and `annualised` is False so the caller can label it honestly.
    """
    if not txns or total_cost <= 0 or total_value <= 0:
        return 0.0, 0.0, False
    buy_dates = [t["date"] for t in txns if t.get("action", "").lower() == "buy"]
    if not buy_dates:
        return 0.0, 0.0, False
    first = date.fromisoformat(min(buy_dates))
    years = (date.today() - first).days / 365.25
    cumulative = (total_value / total_cost - 1) * 100
    if years < 1:
        return round(cumulative, 2), round(years, 2), False
    return round(((total_value / total_cost) ** (1 / years) - 1) * 100, 2), round(years, 2), True


def _calc_dividend_income(user_id: int) -> float:
    """Total dividend income received (AUD) across all holdings."""
    conn = db()
    row = conn.execute(
        "SELECT COALESCE(SUM(net_amount_aud), 0) FROM dividends WHERE user_id = ?", (user_id,)
    ).fetchone()
    conn.close()
    return round(float(row[0]), 2) if row else 0.0


@app.route("/api/stats", methods=["GET"])
@jwt_required()
def get_stats():
    """Return top level aggregated portfolio statistics."""
    uid = current_user_id()
    txns = load_transactions(uid)
    if not txns:
        return jsonify({
            "total_value": 0.0,
            "total_principal": 0.0,
            "total_return": 0.0,
            "total_return_pct": 0.0,
            "best_performer": "-",
            "best_performer_pct": 0.0,
            "worst_performer": "-",
            "worst_performer_pct": 0.0,
            "usd_allocation_pct": 0.0,
            "audusd_rate": 0.65,
            "all_time_high": 0.0,
            "all_time_high_date": None,
            "daily_ath": 0.0,
            "daily_ath_date": None,
            "day_pl": 0.0,
            "day_pl_pct": 0.0,
        })

    # Fetch latest exchange rate and holdings
    conn = db()
    audusd_row = conn.execute(
        "SELECT close FROM prices WHERE symbol = 'AUDUSD=X' ORDER BY date DESC LIMIT 1"
    ).fetchone()
    conn.close()
    audusd = float(audusd_row[0]) if audusd_row else 0.65

    # Retrieve holdings breakdown
    holdings = _compute_active_holdings(uid)
    
    if not holdings:
        return jsonify({
            "total_value": 0.0,
            "total_principal": 0.0,
            "total_return": 0.0,
            "total_return_pct": 0.0,
            "best_performer": "-",
            "best_performer_pct": 0.0,
            "worst_performer": "-",
            "worst_performer_pct": 0.0,
            "usd_allocation_pct": 0.0,
            "audusd_rate": audusd,
            "all_time_high": 0.0,
            "all_time_high_date": None,
            "daily_ath": 0.0,
            "daily_ath_date": None,
            "day_pl": 0.0,
            "day_pl_pct": 0.0,
        })

    total_value = sum(h["value_aud"] for h in holdings)
    total_cost = sum(h["cost_aud"] for h in holdings)
    total_return = total_value - total_cost
    total_return_pct = (total_return / total_cost * 100) if total_cost > 0 else 0.0

    # The holdings-basis return, matching the Portfolio table's Total row and the
    # treemap exactly: price and currency on the units still held, plus the share of
    # dividends and franking belonging to the capital still invested. total_return
    # above is price+currency only, which is why the card reading it sat 1.71pp below
    # the table on the same screen.
    holding_return = sum(h["holding_return_aud"] for h in holdings)
    holding_return_pct = (holding_return / total_cost * 100) if total_cost > 0 else 0.0

    # Today's P&L — sum of each holding's own daily_change (already price-only, not
    # affected by units bought/sold today, same fix as Daily ATH needed)
    day_pl = sum(h["daily_change"] for h in holdings)
    prev_total_value = total_value - day_pl
    day_pl_pct = (day_pl / prev_total_value * 100) if prev_total_value > 0 else 0.0
    
    usd_value = sum(h["value_aud"] for h in holdings if h["currency"] == "USD")
    usd_allocation_pct = (usd_value / total_value * 100) if total_value > 0 else 0.0

    # Find best and worst performer by % return
    best_h = max(holdings, key=lambda x: x["return_pct"])
    worst_h = min(holdings, key=lambda x: x["return_pct"])

    # Track all-time portfolio high
    today = date.today().isoformat()
    conn2 = db()
    row = conn2.execute("SELECT value, date FROM records WHERE key = 'portfolio_high' AND user_id = ?", (uid,)).fetchone()
    if row is None or total_value > row[0]:
        conn2.execute(
            "INSERT OR REPLACE INTO records (key, value, date, user_id) VALUES ('portfolio_high', ?, ?, ?)",
            (round(total_value, 2), today, uid)
        )
        conn2.commit()
        ath_value, ath_date = round(total_value, 2), today
    else:
        ath_value, ath_date = round(row[0], 2), row[1]

    # Daily ATH — the single best day-over-day dollar increase in portfolio value ever
    # recorded (market-driven only — deliberately portfolio value, not net worth, so a
    # manually-entered cash/super update never gets misread as a market "gain").
    daily_ath_value, daily_ath_date = 0.0, None
    try:
        pv_series = _build_daily_market_return_series(conn2, uid)
        if pv_series is not None and len(pv_series) > 1:
            daily_diffs = pv_series.diff().dropna()
            if not daily_diffs.empty:
                best_idx = daily_diffs.idxmax()
                daily_ath_value = round(float(daily_diffs.loc[best_idx]), 2)
                daily_ath_date = best_idx.strftime("%Y-%m-%d")
    except Exception as e:
        print(f"[stats] daily_ath calc failed (non-fatal): {e}")
    conn2.close()

    cagr_pct, cagr_years, cagr_annualised = _calc_cagr(txns, total_value, total_cost)

    # Realised gain, income and the true total return. total_return above is
    # unrealised only (market value less the cost of units still held), which
    # understated lifetime profit by everything banked on a sale and every dividend.
    inc_conn = db()
    div_rows = [
        {"date": d, "net_amount_aud": n or 0.0, "franking_credit_aud": f or 0.0}
        for d, n, f in inc_conn.execute(
            "SELECT date, net_amount_aud, franking_credit_aud FROM dividends WHERE user_id = ?", (uid,)
        ).fetchall()
    ]
    inc_conn.close()
    realised_gain = _calc_realised_gain(txns)
    income_total = round(sum(r["net_amount_aud"] for r in div_rows), 2)
    franking_total = round(sum(r["franking_credit_aud"] or 0 for r in div_rows), 2)
    # Australian FY window, bounded at BOTH ends — a lower bound alone counted any
    # future-dated dividend into the current year.
    _t = date.today()
    fy_start = date(_t.year - (1 if _t.month < 7 else 0), 7, 1)
    fy_end = date(fy_start.year + 1, 6, 30)
    income_fy = round(sum(r["net_amount_aud"] for r in div_rows
                          if fy_start.isoformat() <= r["date"][:10] <= fy_end.isoformat()), 2)
    fy_start = fy_start.isoformat()
    # Franking counts toward the total. It isn't cash the portfolio received — it's an
    # ATO offset — but for an Australian resident it's refundable, so leaving it out
    # understates what a fully-franked position was actually worth holding. Sharesight
    # keeps it outside the return by default; this app deliberately doesn't.
    total_return_all = round(total_return + realised_gain + income_total + franking_total, 2)
    mwr = _calc_mwr(txns, div_rows, total_value)

    return jsonify({
        "total_value": round(total_value, 2),
        "total_principal": round(total_cost, 2),
        "total_return": round(total_return, 2),
        "total_return_pct": round(total_return_pct, 2),
        "holding_return": round(holding_return, 2),
        "holding_return_pct": round(holding_return_pct, 2),
        "realised_gain": realised_gain,
        "income_total": income_total,
        "franking_total": franking_total,
        "income_fy": income_fy,
        "total_return_all": total_return_all,
        "mwr_pct": mwr["pct"],
        "mwr_years": mwr["years"],
        "mwr_annualised": mwr["annualised"],
        "mwr_pct_ex_income": mwr["pct_ex_income"],
        "best_performer": f"{best_h['ticker']} ({best_h['return_pct']:+.1f}%)",
        "best_performer_pct": round(best_h["return_pct"], 2),
        "worst_performer": f"{worst_h['ticker']} ({worst_h['return_pct']:+.1f}%)",
        "worst_performer_pct": round(worst_h["return_pct"], 2),
        "usd_allocation_pct": round(usd_allocation_pct, 2),
        "audusd_rate": round(audusd, 4),
        "all_time_high": ath_value,
        "all_time_high_date": ath_date,
        "daily_ath": daily_ath_value,
        "daily_ath_date": daily_ath_date,
        "day_pl": round(day_pl, 2),
        "day_pl_pct": round(day_pl_pct, 2),
        "cost_basis": round(total_cost, 2),
        "cagr": cagr_pct,
        "cagr_years": cagr_years,
        "cagr_annualised": cagr_annualised,
        "dividend_income": _calc_dividend_income(uid),
    })

def _get_latest_portfolio_value(user_id):
    """Helper: return total portfolio value from current holdings."""
    txns = load_transactions(user_id)
    if not txns:
        return 0.0, 0.0, 0.0  # value, active_stocks, passive_stocks
    conn = db()
    latest_prices = {}
    rows = conn.execute("""
        SELECT symbol, close FROM prices
        WHERE (symbol, date) IN (
            SELECT symbol, MAX(date) FROM prices GROUP BY symbol
        )
    """).fetchall()
    for sym, close in rows:
        latest_prices[sym] = close
    conn.close()
    audusd = latest_prices.get("AUDUSD=X", 0.65)

    txns_sorted = sorted(txns, key=lambda x: x["date"])
    holdings = {}
    for t in txns_sorted:
        sym = yf_symbol(t["ticker"], t["exchange"])
        if sym not in holdings:
            holdings[sym] = {
                "ticker": t["ticker"], "exchange": t["exchange"],
                "name": t.get("name") or f"{t['ticker']} Stock",
                "currency": t.get("currency") or "AUD",
                "units": 0.0
            }
        h = holdings[sym]
        qty = t["units"]
        if t["action"].lower() == "buy":
            h["units"] += qty
        elif t["action"].lower() == "split":
            h["units"] += qty
        elif t["action"].lower() == "sell":
            h["units"] = max(0, h["units"] - qty)

    total_value = 0.0
    active_value = 0.0
    passive_value = 0.0
    for sym, h in holdings.items():
        if h["units"] <= 1e-5:
            continue
        price = latest_prices.get(sym, 0.0)
        if h["currency"] == "USD":
            price = price / audusd
        val = h["units"] * price
        total_value += val
        # Classification: ETFs contain "Etf" or "Index" in name
        name_lower = h["name"].lower()
        if "etf" in name_lower or "index" in name_lower:
            passive_value += val
        else:
            active_value += val
    return total_value, active_value, passive_value

# CGT discount by entity type. 50% is the individual/trust rate; a complying super fund
# gets one third; companies get no discount at all. Previously hardcoded to 0.5.
CGT_DISCOUNT_RATES = {"individual": 0.5, "trust": 0.5, "smsf": 1.0 / 3.0, "company": 0.0}


def _get_tax_settings(user_id):
    conn = db()
    row = conn.execute(
        "SELECT entity_type, allocation_method FROM tax_settings WHERE user_id = ?", (user_id,)
    ).fetchone()
    conn.close()
    if not row:
        return {"entity_type": "individual", "allocation_method": "fifo"}
    entity = row[0] if row[0] in CGT_DISCOUNT_RATES else "individual"
    method = row[1] if row[1] in ("fifo", "lifo", "hifo") else "fifo"
    return {"entity_type": entity, "allocation_method": method}


def _get_carryforward(user_id, fy_start):
    """Prior-year net capital loss the user has recorded against this FY."""
    if not fy_start:
        return 0.0
    conn = db()
    row = conn.execute(
        "SELECT amount_aud FROM capital_loss_carryforward WHERE user_id = ? AND fy_start = ?",
        (user_id, fy_start[:10]),
    ).fetchone()
    conn.close()
    return max(0.0, float(row[0])) if row else 0.0


def _distribution_capital_gains(user_id, from_date, to_date):
    """Capital-gain components attributed by trust distributions in the period.

    Australian ETFs distribute realised capital gains inside what the app records as a
    dividend. They are assessable as CAPITAL GAINS (the discounted portion attracting the
    CGT discount), not as ordinary income, and were previously omitted from CGT entirely.
    Also returns tax_deferred, which reduces the cost base rather than being assessable.
    """
    conn = db()
    sql = ("SELECT COALESCE(SUM(cg_discounted_aud),0), COALESCE(SUM(cg_other_aud),0), "
           "COALESCE(SUM(tax_deferred_aud),0) FROM dividends WHERE user_id = ?")
    args = [user_id]
    if from_date:
        sql += " AND date >= ?"; args.append(from_date)
    if to_date:
        sql += " AND date <= ?"; args.append(to_date)
    d, o, td = conn.execute(sql, args).fetchone()
    conn.close()
    return {"discounted": float(d or 0), "other": float(o or 0), "tax_deferred": float(td or 0)}


def _load_sale_allocations(user_id):
    """Locked allocations as {sell_txn_id: [(buy_txn_id, units), ...]} in stored order."""
    conn = db()
    rows = conn.execute(
        "SELECT sell_txn_id, buy_txn_id, units, method FROM sale_allocations "
        "WHERE user_id = ? ORDER BY sell_txn_id, rowid", (user_id,)
    ).fetchall()
    conn.close()
    locked, method_by_sell = {}, {}
    for sell_id, buy_id, units, method in rows:
        locked.setdefault(sell_id, []).append((buy_id, float(units)))
        method_by_sell[sell_id] = method
    return locked, method_by_sell


def _order_parcels_for_disposal(parcels, method):
    """Return parcels ordered by which should be treated as sold first, per method.
    fifo = oldest first (default, what Sharesight uses unless configured otherwise)
    lifo = newest first
    hifo = highest cost-per-unit first (minimizes reported gain — a legitimate,
           commonly-offered specific-identification strategy, not a shortcut)
    """
    live = [p for p in parcels if p["units"] > 1e-9]
    if method == "lifo":
        return sorted(live, key=lambda p: p["date"], reverse=True)
    if method == "hifo":
        return sorted(live, key=lambda p: (p["cost_aud"] / p["units"]) if p["units"] > 0 else 0, reverse=True)
    return sorted(live, key=lambda p: p["date"])  # fifo default

@app.route("/api/cgt", methods=["GET"])
@jwt_required()
def get_cgt():
    """Calculate Australian CGT for sells within a date range, using real per-parcel
    lot tracking (not blended average cost) so the 12-month discount test applies to
    the specific units actually disposed of — a single sale can legitimately be part
    discount-eligible and part not, if it draws from parcels of different ages."""
    uid = current_user_id()
    from_date = request.args.get("from", "")
    to_date = request.args.get("to", "")
    method = request.args.get("method", "").lower()

    settings = _get_tax_settings(uid)
    if method not in ("fifo", "lifo", "hifo"):
        method = settings["allocation_method"]
    discount_rate = CGT_DISCOUNT_RATES.get(settings["entity_type"], 0.5)

    # Prior-year net capital losses. The ATO requires these be applied to gains BEFORE
    # the discount, and there was previously no way to supply them, so every year was
    # computed in isolation and net gain was overstated for anyone carrying a loss.
    prior_losses = request.args.get("prior_losses")
    if prior_losses is None:
        prior_losses = _get_carryforward(uid, from_date)
    else:
        try:
            prior_losses = max(0.0, float(prior_losses))
        except (TypeError, ValueError):
            prior_losses = 0.0

    def empty(extra_note=None):
        return jsonify({
            "gains": [], "total_gain": 0, "gross_gains": 0, "gross_losses": 0,
            "discounted_gains": 0, "non_discounted_gains": 0,
            "distribution_gains": 0, "distribution_gains_discounted": 0,
            "losses_applied": 0, "prior_losses_available": round(prior_losses, 2),
            "prior_losses_applied": 0, "losses_carried_forward": round(prior_losses, 2),
            "cgt_discount": 0, "net_gain": 0, "net_capital_loss": 0,
            "from": from_date, "to": to_date, "method": method,
            "entity_type": settings["entity_type"], "discount_rate": discount_rate,
            "warnings": [extra_note] if extra_note else [],
        })

    txns = load_transactions(uid)
    if not txns:
        return empty()

    sells = [t for t in txns if t["action"].lower() == "sell"]
    if from_date:
        sells = [t for t in sells if t["date"] >= from_date]
    if to_date:
        sells = [t for t in sells if t["date"] <= to_date]
    dist = _distribution_capital_gains(uid, from_date, to_date)
    if not sells and not (dist["discounted"] or dist["other"]):
        return empty()

    locked, lock_method = _load_sale_allocations(uid)

    txns_sorted = sorted(txns, key=lambda x: x["date"])
    parcels_by_sym = {}  # sym -> list of {date, units, cost_aud} — one entry per buy lot, consumed over time

    gains = []
    warnings = []
    for t in txns_sorted:
        sym = yf_symbol(t["ticker"], t["exchange"])
        parcels = parcels_by_sym.setdefault(sym, [])
        action = t["action"].lower()

        if action == "buy":
            # buy_id gives the parcel a stable identity so a locked allocation can name it.
            parcels.append({"date": t["date"], "units": t["units"], "cost_aud": t["value"],
                            "buy_id": t.get("id")})

        elif action == "split":
            # Scale every existing parcel's units up proportionally, cost basis unchanged
            total_units = sum(p["units"] for p in parcels)
            if total_units > 1e-9:
                ratio = (total_units + t["units"]) / total_units
                for p in parcels:
                    p["units"] *= ratio

        elif action == "sell":
            units_to_sell = t["units"]
            proceeds_total = abs(t["value"])
            proceeds_per_unit = proceeds_total / units_to_sell if units_to_sell > 0 else 0

            in_range = (not from_date or t["date"] >= from_date) and (not to_date or t["date"] <= to_date)

            # A locked disposal replays the parcels it actually consumed when the year was
            # lodged, in that order, ignoring the currently-selected method. Without this,
            # switching method would rewrite a filed year and shift every later cost base.
            lock = locked.get(t.get("id"))
            if lock:
                by_id = {p.get("buy_id"): p for p in parcels}
                ordered = [by_id[b] for b, _u in lock if b in by_id]
                locked_units = {b: u for b, u in lock}
                effective_method = f"{lock_method.get(t.get('id'), method)} (locked)"
            else:
                ordered = _order_parcels_for_disposal(parcels, method)
                locked_units = None
                effective_method = method
            remaining = units_to_sell
            for p in ordered:
                if remaining <= 1e-9:
                    break
                # A locked allocation dictates the exact units drawn from this parcel.
                take = min(p["units"], remaining)
                if locked_units is not None:
                    take = min(take, locked_units.get(p.get("buy_id"), 0.0))
                    if take <= 1e-9:
                        continue
                per_unit_cost = p["cost_aud"] / p["units"] if p["units"] > 0 else 0
                slice_cost = take * per_unit_cost
                slice_proceeds = take * proceeds_per_unit
                slice_gain = slice_proceeds - slice_cost

                buy_dt = pd.Timestamp(p["date"])
                sell_dt = pd.Timestamp(t["date"])
                # MORE than 12 months, not "at least 365 days". The ATO test excludes the
                # acquisition day: bought 1 Jul 2020, the earliest qualifying disposal is
                # 2 Jul 2021 — 366 days. `>= 365` granted the 50% discount a day early.
                held_12m = (sell_dt - buy_dt).days > 365

                if in_range:
                    gains.append({
                        "date": t["date"],
                        "acquired_date": p["date"],
                        "ticker": t["ticker"],
                        "name": t.get("name", ""),
                        "units": round(take, 4),
                        "proceeds": round(slice_proceeds, 2),
                        "cost": round(slice_cost, 2),
                        "gain": round(slice_gain, 2),
                        "held_12m": held_12m,
                        "discount_eligible": held_12m and slice_gain > 0,
                        "buy_id": p.get("buy_id"),
                        "sell_id": t.get("id"),
                        "locked": bool(lock),
                        "method": effective_method,
                    })

                p["units"] -= take
                p["cost_aud"] -= slice_cost
                remaining -= take

            # Units disposed of with no parcel to draw from. Previously the loop simply
            # exited and the proceeds vanished from the report with no indication, so a
            # missing or mis-dated buy silently understated the gain.
            if remaining > 1e-6 and in_range:
                warnings.append(
                    f"{t['ticker']}: sold {remaining:.4f} units on {t['date']} with no matching "
                    f"purchase parcel — proceeds of "
                    f"{round(remaining * proceeds_per_unit, 2)} AUD are NOT in this report. "
                    f"Check for a missing buy transaction."
                )

    # Calculate CGT summary
    total_gain = sum(g["gain"] for g in gains)
    total_losses = sum(g["gain"] for g in gains if g["gain"] < 0)
    total_discountable = sum(g["gain"] for g in gains if g["discount_eligible"])
    # Trust distributions add capital gains that have no disposal behind them.
    total_discountable += dist["discounted"]

    # Capital losses must be applied BEFORE the 50% discount, and the ATO lets the
    # taxpayer choose which gains to apply them against. Non-discounted gains first is
    # always at least as good: a dollar of loss cancels a full taxable dollar there,
    # versus only fifty cents against a discounted gain. This previously ran the other
    # way round despite the comment, overstating the taxable gain.
    non_discounted = sum(g["gain"] for g in gains if g["gain"] > 0 and not g["discount_eligible"])
    non_discounted += dist["other"]

    # Current-year losses first, then prior-year carried-forward losses. Within each, hit
    # non-discounted gains before discounted ones: a dollar of loss cancels a full taxable
    # dollar there, versus only fifty cents against a discounted gain. Verified against
    # Sharesight, whose net capital gain matches this ordering to the cent.
    current_losses = abs(total_losses)
    pool = current_losses + prior_losses

    non_discounted_after = max(0.0, non_discounted - pool)
    pool = max(0.0, pool - non_discounted)

    discounted_after = max(0.0, total_discountable - pool)
    pool = max(0.0, pool - total_discountable)

    absorbed = (current_losses + prior_losses) - pool
    current_applied = min(current_losses, absorbed)
    prior_applied = max(0.0, absorbed - current_applied)

    cgt_discount = round(discounted_after * discount_rate, 2)
    net_gain = round(discounted_after * (1 - discount_rate) + non_discounted_after, 2)

    # A loss-making year has no taxable gain AND a loss to carry forward. Reporting only
    # a floored-at-zero net_gain hid the carry-forward amount the user must record.
    net_capital_loss = round(pool, 2)

    if len(warnings) == 0 and prior_losses > 0 and prior_applied == 0:
        warnings.append(
            f"{prior_losses:,.2f} of prior-year losses were available but no gains "
            f"remained to absorb them; the full amount still carries forward."
        )

    return jsonify({
        "gains": gains,
        # total_gain is NET of losses — gross_gains/gross_losses are the reconcilable pair.
        "total_gain": round(total_gain, 2),
        "gross_gains": round(sum(g["gain"] for g in gains if g["gain"] > 0), 2),
        "gross_losses": round(current_losses, 2),
        "discounted_gains": round(total_discountable, 2),
        "non_discounted_gains": round(non_discounted, 2),
        "distribution_gains_discounted": round(dist["discounted"], 2),
        "distribution_gains_other": round(dist["other"], 2),
        "tax_deferred_distributions": round(dist["tax_deferred"], 2),
        "losses_applied": round(current_applied, 2),
        "prior_losses_available": round(prior_losses, 2),
        "prior_losses_applied": round(prior_applied, 2),
        "losses_carried_forward": net_capital_loss,
        "net_capital_loss": net_capital_loss,
        "cgt_discount": cgt_discount,
        "net_gain": net_gain,
        "from": from_date,
        "to": to_date,
        "method": method,
        "entity_type": settings["entity_type"],
        "discount_rate": discount_rate,
        "warnings": warnings,
    })
@app.route("/api/tax/income", methods=["GET"])
@jwt_required()
def get_taxable_income():
    """Australian taxable income report for a date range — the counterpart to /api/cgt.

    Assessable dividend income is the GROSS amount plus the franking credit, not the net
    cash received. The app previously only ever surfaced net_amount_aud (gross less foreign
    withholding), which under-declares foreign income and loses the foreign tax offset the
    user is entitled to claim.

    Trust distributions are split into their attribution components: the capital-gains
    parts belong in the CGT report (and are excluded here to avoid double counting), while
    tax-deferred amounts are not assessable at all — they reduce the cost base.
    """
    uid = current_user_id()
    from_date = request.args.get("from", "")
    to_date = request.args.get("to", "")

    conn = db()
    sql = ("SELECT date, ticker, symbol, exchange, currency, gross_amount, gross_amount_aud, "
           "franking_pct, franking_credit_aud, withholding_tax_pct, net_amount_aud, "
           "cg_discounted_aud, cg_other_aud, tax_deferred_aud, foreign_income_aud, "
           "foreign_tax_paid_aud FROM dividends WHERE user_id = ?")
    args = [uid]
    if from_date:
        sql += " AND date >= ?"; args.append(from_date)
    if to_date:
        sql += " AND date <= ?"; args.append(to_date)
    rows = conn.execute(sql + " ORDER BY date", args).fetchall()
    conn.close()

    cols = ["date", "ticker", "symbol", "exchange", "currency", "gross_amount",
            "gross_amount_aud", "franking_pct", "franking_credit_aud",
            "withholding_tax_pct", "net_amount_aud", "cg_discounted_aud", "cg_other_aud",
            "tax_deferred_aud", "foreign_income_aud", "foreign_tax_paid_aud"]
    items, agg = [], {
        "gross_income": 0.0, "franking_credits": 0.0, "withholding_tax": 0.0,
        "net_cash": 0.0, "franked_income": 0.0, "unfranked_income": 0.0,
        "foreign_income": 0.0, "foreign_tax_offsets": 0.0,
        "capital_gain_distributions": 0.0, "tax_deferred": 0.0,
    }

    for r in rows:
        d = dict(zip(cols, r))
        for k in cols[5:]:
            d[k] = float(d[k] or 0)

        # Components that are NOT ordinary income: capital gains belong to the CGT
        # report, tax-deferred amounts reduce cost base and are never assessable.
        cg = d["cg_discounted_aud"] + d["cg_other_aud"]
        income_aud = max(0.0, d["gross_amount_aud"] - cg - d["tax_deferred_aud"])

        is_foreign = (d["currency"] or "AUD").upper() != "AUD"
        franked = income_aud * (d["franking_pct"] / 100.0)
        withheld = d["gross_amount_aud"] * (d["withholding_tax_pct"] / 100.0)
        foreign_tax = d["foreign_tax_paid_aud"] or (withheld if is_foreign else 0.0)

        agg["gross_income"] += income_aud
        agg["franking_credits"] += d["franking_credit_aud"]
        agg["withholding_tax"] += withheld
        agg["net_cash"] += d["net_amount_aud"]
        agg["franked_income"] += franked
        agg["unfranked_income"] += income_aud - franked
        agg["foreign_income"] += d["foreign_income_aud"] or (income_aud if is_foreign else 0.0)
        agg["foreign_tax_offsets"] += foreign_tax
        agg["capital_gain_distributions"] += cg
        agg["tax_deferred"] += d["tax_deferred_aud"]

        items.append({
            "date": d["date"], "ticker": d["ticker"], "exchange": d["exchange"],
            "currency": d["currency"],
            "income_aud": round(income_aud, 2),
            "franking_credit_aud": round(d["franking_credit_aud"], 2),
            "withholding_tax_aud": round(withheld, 2),
            "net_cash_aud": round(d["net_amount_aud"], 2),
            "capital_gain_aud": round(cg, 2),
            "tax_deferred_aud": round(d["tax_deferred_aud"], 2),
            "foreign": is_foreign,
        })

    out = {k: round(v, 2) for k, v in agg.items()}
    # Grossed-up: what actually goes in the return. Franking credits are assessable income
    # and then claimed back as a credit — omitting them under-declares income.
    out["assessable_income"] = round(agg["gross_income"] + agg["franking_credits"], 2)
    out["items"] = items
    out["from"] = from_date
    out["to"] = to_date
    out["components_entered"] = bool(agg["capital_gain_distributions"] or agg["tax_deferred"])
    return jsonify(out)


@app.route("/api/tax/settings", methods=["GET", "POST"])
@jwt_required()
def tax_settings():
    uid = current_user_id()
    if request.method == "GET":
        s = _get_tax_settings(uid)
        s["discount_rate"] = CGT_DISCOUNT_RATES.get(s["entity_type"], 0.5)
        s["entity_options"] = list(CGT_DISCOUNT_RATES.keys())
        return jsonify(s)
    data = request.json or {}
    entity = data.get("entity_type", "individual")
    method = data.get("allocation_method", "fifo")
    if entity not in CGT_DISCOUNT_RATES:
        return jsonify({"error": f"entity_type must be one of {list(CGT_DISCOUNT_RATES)}"}), 400
    if method not in ("fifo", "lifo", "hifo"):
        return jsonify({"error": "allocation_method must be fifo, lifo or hifo"}), 400
    conn = db()
    conn.execute(
        "INSERT INTO tax_settings (user_id, entity_type, allocation_method, updated_at) "
        "VALUES (?,?,?,?) ON CONFLICT(user_id) DO UPDATE SET entity_type=excluded.entity_type, "
        "allocation_method=excluded.allocation_method, updated_at=excluded.updated_at",
        (uid, entity, method, datetime.now().isoformat()),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "entity_type": entity, "allocation_method": method,
                    "discount_rate": CGT_DISCOUNT_RATES[entity]})


@app.route("/api/tax/lock", methods=["GET", "POST", "DELETE"])
@jwt_required()
def sale_allocation_lock():
    """Freeze which parcels each disposal consumed, up to and including a date.

    Lodge FY2025 under LIFO, then switch to FIFO next year, and without this the FY2025
    figures silently change — and so do the leftover parcels, which corrupts FY2026's cost
    base too. Locking records the actual allocation so a later method change only affects
    disposals after the locked date.
    """
    uid = current_user_id()
    conn = db()

    if request.method == "GET":
        row = conn.execute(
            "SELECT COUNT(DISTINCT sell_txn_id), MAX(locked_at) FROM sale_allocations WHERE user_id = ?",
            (uid,)).fetchone()
        rows = conn.execute(
            "SELECT sa.sell_txn_id, t.date, t.ticker, sa.method, COUNT(*), SUM(sa.units) "
            "FROM sale_allocations sa JOIN transactions t ON t.id = sa.sell_txn_id "
            "WHERE sa.user_id = ? GROUP BY sa.sell_txn_id ORDER BY t.date", (uid,)).fetchall()
        conn.close()
        return jsonify({
            "locked_disposals": row[0] or 0,
            "last_locked_at": row[1],
            "locked_to": max((r[1] for r in rows), default=None),
            "disposals": [{"sell_id": r[0], "date": r[1], "ticker": r[2], "method": r[3],
                           "parcels": r[4], "units": round(r[5], 4)} for r in rows],
        })

    if request.method == "DELETE":
        # Unlocking re-exposes filed years to method changes, so it is explicit and loud.
        to_date = request.args.get("to", "")
        if to_date:
            conn.execute(
                "DELETE FROM sale_allocations WHERE user_id = ? AND sell_txn_id IN "
                "(SELECT id FROM transactions WHERE user_id = ? AND date <= ?)", (uid, uid, to_date))
        else:
            conn.execute("DELETE FROM sale_allocations WHERE user_id = ?", (uid,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "unlocked_to": to_date or "all"})

    conn.close()
    data = request.json or {}
    to_date = (data.get("to") or request.args.get("to") or "")[:10]
    if not to_date:
        return jsonify({"error": "to (YYYY-MM-DD) required — the last date to lock"}), 400

    # Recompute allocations under the CURRENT method, then persist them for every disposal
    # up to the cut-off. Anything already locked is left as-is — re-locking must never
    # silently rewrite a year that was filed under a different method.
    settings = _get_tax_settings(uid)
    method = data.get("method") or settings["allocation_method"]
    if method not in ("fifo", "lifo", "hifo"):
        return jsonify({"error": "method must be fifo, lifo or hifo"}), 400

    already, _ = _load_sale_allocations(uid)
    txns = load_transactions(uid)
    parcels_by_sym, to_store = {}, []

    for t in sorted(txns, key=lambda x: (x["date"], x.get("id") or 0)):
        sym = yf_symbol(t["ticker"], t["exchange"])
        parcels = parcels_by_sym.setdefault(sym, [])
        action = (t["action"] or "").lower()
        if action == "buy":
            parcels.append({"date": t["date"], "units": t["units"], "cost_aud": t["value"],
                            "buy_id": t.get("id")})
        elif action == "split":
            tot = sum(p["units"] for p in parcels)
            if tot > 1e-9:
                ratio = (tot + t["units"]) / tot
                for p in parcels:
                    p["units"] *= ratio
        elif action == "sell":
            sell_id = t.get("id")
            prior = already.get(sell_id)
            if prior:
                # Replay the existing lock so downstream parcels stay consistent.
                by_id = {p.get("buy_id"): p for p in parcels}
                for buy_id, units in prior:
                    p = by_id.get(buy_id)
                    if not p or p["units"] <= 1e-9:
                        continue
                    take = min(p["units"], units)
                    p["cost_aud"] -= take * (p["cost_aud"] / p["units"])
                    p["units"] -= take
                continue
            remaining = t["units"]
            for p in _order_parcels_for_disposal(parcels, method):
                if remaining <= 1e-9:
                    break
                take = min(p["units"], remaining)
                per_unit = p["cost_aud"] / p["units"] if p["units"] > 0 else 0
                if t["date"] <= to_date:
                    to_store.append((uid, sell_id, p.get("buy_id"), take, take * per_unit, method))
                p["units"] -= take
                p["cost_aud"] -= take * per_unit
                remaining -= take

    conn = db()
    now = datetime.now().isoformat()
    for uid_, sell_id, buy_id, units, cost, m in to_store:
        conn.execute(
            "INSERT INTO sale_allocations (user_id, sell_txn_id, buy_txn_id, units, cost_aud, method, locked_at) "
            "VALUES (?,?,?,?,?,?,?) ON CONFLICT(user_id, sell_txn_id, buy_txn_id) DO NOTHING",
            (uid_, sell_id, buy_id, units, cost, m, now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "locked_to": to_date, "method": method,
                    "allocations_written": len(to_store),
                    "disposals_already_locked": len(already)})


@app.route("/api/tax/carryforward", methods=["GET", "POST"])
@jwt_required()
def carryforward_losses():
    """Prior-year net capital losses the user is bringing into a financial year."""
    uid = current_user_id()
    if request.method == "GET":
        conn = db()
        rows = conn.execute(
            "SELECT fy_start, amount_aud, note FROM capital_loss_carryforward "
            "WHERE user_id = ? ORDER BY fy_start", (uid,)
        ).fetchall()
        conn.close()
        return jsonify([{"fy_start": r[0], "amount_aud": r[1], "note": r[2]} for r in rows])
    data = request.json or {}
    fy_start = (data.get("fy_start") or "")[:10]
    if not fy_start:
        return jsonify({"error": "fy_start required (YYYY-MM-DD, the first day of the FY)"}), 400
    try:
        amount = max(0.0, float(data.get("amount_aud", 0)))
    except (TypeError, ValueError):
        return jsonify({"error": "amount_aud must be a number"}), 400
    conn = db()
    conn.execute(
        "INSERT INTO capital_loss_carryforward (user_id, fy_start, amount_aud, note) "
        "VALUES (?,?,?,?) ON CONFLICT(user_id, fy_start) DO UPDATE SET "
        "amount_aud=excluded.amount_aud, note=excluded.note",
        (uid, fy_start, amount, data.get("note", "")),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "fy_start": fy_start, "amount_aud": amount})


@app.route("/api/snapshots", methods=["GET"])
@jwt_required()
def get_snapshots():
    """Return all cash + super snapshots sorted by date."""
    conn = db()
    rows = conn.execute("SELECT date, super, cash FROM snapshots WHERE user_id = ? ORDER BY date", (current_user_id(),)).fetchall()
    conn.close()
    return jsonify([{"date": r[0], "super": r[1], "cash": r[2]} for r in rows])

@app.route("/api/snapshots", methods=["POST"])
@jwt_required()
def add_snapshot():
    """Add or update a cash + super snapshot. Persists to DB and snapshots.json."""
    data = request.json
    try:
        date_str = data["date"]
        super_val = float(data["super"])
        cash_val = float(data["cash"])

        # A future-dated snapshot silently breaks every "current value" lookup in the
        # app (they correctly filter WHERE date <= today, so a future row just gets
        # ignored until the calendar catches up) — this is exactly what caused cash to
        # appear stuck on an old value with no visible error. Reject it outright rather
        # than let it happen silently again.
        if date_str > date.today().isoformat():
            return jsonify({
                "ok": False,
                "error": f"Snapshot date {date_str} is in the future — it would be silently "
                         f"ignored by every 'current value' lookup until that date arrives. "
                         f"Use today's date ({date.today().isoformat()}) or an actual past date."
            }), 400

        conn = db()
        conn.execute(
            "INSERT OR REPLACE INTO snapshots (date, super, cash, user_id) VALUES (?, ?, ?, ?)",
            (date_str, super_val, cash_val, current_user_id()),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/snapshots/<snap_date>", methods=["DELETE"])
@jwt_required()
def delete_snapshot(snap_date):
    conn = db()
    conn.execute("DELETE FROM snapshots WHERE date = ? AND user_id = ?", (snap_date, current_user_id()))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/milestones", methods=["GET"])
@jwt_required()
def get_milestones():
    """Return all milestones with live current values for linked metrics.

    Goals can track multiple metrics at once (e.g. cash + portfolio), which are
    summed live. Targets can be set in AUD or USD — a USD target is converted to
    its AUD equivalent using the latest cached AUDUSD=X rate on every request, so
    the goal's progress moves with the exchange rate rather than freezing at the
    rate that was in effect when the milestone was created.
    """
    uid = current_user_id()
    # Get live metrics once
    live = {}
    audusd = 0.65
    try:
        conn2 = db()
        fx_row = conn2.execute(
            "SELECT close FROM prices WHERE symbol = 'AUDUSD=X' ORDER BY date DESC LIMIT 1"
        ).fetchone()
        if fx_row and fx_row[0]:
            audusd = fx_row[0]

        portfolio_val, active_val, passive_val = _get_latest_portfolio_value(uid)
        cash = get_total_cash(uid)
        today = date.today().isoformat()
        row = conn2.execute("SELECT super FROM snapshots WHERE date <= ? AND user_id = ? ORDER BY date DESC LIMIT 1", (today, uid)).fetchone()
        conn2.close()
        super_val = row[0] if row else 0.0
        holdings = _compute_active_holdings(uid)
        stats_data = {}
        if holdings:
            total_cost = sum(h.get("cost_aud", 0) for h in holdings)
            total_val = sum(h.get("value_aud", 0) for h in holdings)
            stats_data["return_pct"] = round((total_val - total_cost) / total_cost * 100, 2) if total_cost else 0
            stats_data["return_aud"] = round(total_val - total_cost, 2)
        live = {
            "portfolio": round(portfolio_val, 2),
            "networth": round(portfolio_val + cash + super_val, 2),
            "cash": round(cash, 2),
            "super": round(super_val, 2),
            "return_pct": stats_data.get("return_pct", 0),
            "return_aud": stats_data.get("return_aud", 0),
        }
    except:
        pass

    conn = db()
    rows = conn.execute(
        "SELECT id, date, title, description, category, value, type, target_value, current_value, "
        "is_achieved, linked_metric, achieved_date, linked_metrics, currency FROM milestones WHERE user_id = ? ORDER BY date DESC",
        (uid,)
    ).fetchall()

    results = []
    for r in rows:
        mtype = r[6] or "achievement"
        linked_legacy = r[10]
        current_val = r[8]
        is_achieved = bool(r[9])
        achieved_date = r[11]
        linked_metrics_raw = r[12]
        currency = r[13] or "AUD"
        target_value = r[7]

        # New multi-metric field wins; fall back to the legacy single-metric field
        metrics = (
            [m.strip() for m in linked_metrics_raw.split(",") if m.strip()]
            if linked_metrics_raw else ([linked_legacy] if linked_legacy else [])
        )

        # A target set in USD is re-converted to its AUD equivalent at the *current*
        # rate every time this endpoint runs, so it fluctuates with the market
        # rather than being fixed at entry time.
        target_value_aud = target_value
        if target_value is not None and currency == "USD" and audusd:
            target_value_aud = target_value / audusd

        if mtype == "goal" and metrics and all(m in live for m in metrics):
            current_val = round(sum(live[m] for m in metrics), 2)
            if target_value_aud is not None and current_val >= target_value_aud and not is_achieved:
                is_achieved = True
                achieved_date = date.today().isoformat()
                conn.execute("UPDATE milestones SET is_achieved=1, achieved_date=?, current_value=? WHERE id=? AND user_id=?",
                             (achieved_date, current_val, r[0], uid))
            elif not is_achieved:
                conn.execute("UPDATE milestones SET current_value=? WHERE id=? AND user_id=?", (current_val, r[0], uid))

        results.append({
            "id": r[0],
            "date": r[1],
            "title": r[2],
            "description": r[3],
            "category": r[4],
            "value": r[5],
            "type": mtype,
            "target_value": target_value,
            "target_value_aud": round(target_value_aud, 2) if target_value_aud is not None else None,
            "current_value": current_val,
            "is_achieved": is_achieved,
            "linked_metric": linked_legacy,
            "linked_metrics": metrics,
            "currency": currency,
            "achieved_date": achieved_date,
        })
    conn.commit()
    conn.close()
    return jsonify(results)

@app.route("/api/milestones", methods=["POST"])
@jwt_required()
def add_milestone():
    data = request.json
    try:
        conn = db()
        mtype = data.get("type", "achievement")

        # Accept either the new multi-metric list or the legacy single metric
        metrics = data.get("linked_metrics") or ([data["linked_metric"]] if data.get("linked_metric") else [])
        linked_metrics_str = ",".join(metrics) if metrics else None
        linked = metrics[0] if metrics else None  # keep legacy column populated for backward compat

        currency = data.get("currency") or "AUD"
        target_value = data.get("target_value")
        target_value_aud = target_value
        if target_value is not None and currency == "USD":
            fx_row = conn.execute(
                "SELECT close FROM prices WHERE symbol = 'AUDUSD=X' ORDER BY date DESC LIMIT 1"
            ).fetchone()
            audusd = fx_row[0] if fx_row and fx_row[0] else 0.65
            target_value_aud = target_value / audusd

        is_achieved = False
        achieved_date = None
        current_val = data.get("current_value")
        if mtype == "goal" and current_val is not None and target_value_aud is not None:
            if current_val >= target_value_aud:
                is_achieved = True
                achieved_date = date.today().isoformat()
        conn.execute(
            "INSERT INTO milestones (date, title, description, category, value, type, target_value, current_value, "
            "is_achieved, linked_metric, achieved_date, linked_metrics, currency, user_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (data["date"], data["title"], data.get("description", ""), data["category"],
             data.get("value"), mtype, target_value, current_val,
             int(is_achieved), linked, achieved_date, linked_metrics_str, currency, current_user_id())
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/milestones/<int:milestone_id>", methods=["PUT"])
@jwt_required()
def update_milestone(milestone_id):
    data = request.json
    try:
        conn = db()
        mtype = data.get("type", "achievement")

        metrics = data.get("linked_metrics") or ([data["linked_metric"]] if data.get("linked_metric") else [])
        linked_metrics_str = ",".join(metrics) if metrics else None
        linked = metrics[0] if metrics else None

        currency = data.get("currency") or "AUD"
        target_value = data.get("target_value")
        target_value_aud = target_value
        if target_value is not None and currency == "USD":
            fx_row = conn.execute(
                "SELECT close FROM prices WHERE symbol = 'AUDUSD=X' ORDER BY date DESC LIMIT 1"
            ).fetchone()
            audusd = fx_row[0] if fx_row and fx_row[0] else 0.65
            target_value_aud = target_value / audusd

        is_achieved = data.get("is_achieved", False)
        achieved_date = data.get("achieved_date")
        current_val = data.get("current_value")
        if mtype == "goal" and current_val is not None and target_value_aud is not None:
            if current_val >= target_value_aud and not is_achieved:
                is_achieved = True
                achieved_date = date.today().isoformat()
        conn.execute(
            "UPDATE milestones SET date=?,title=?,description=?,category=?,value=?,type=?,target_value=?,current_value=?,"
            "is_achieved=?,linked_metric=?,achieved_date=?,linked_metrics=?,currency=? WHERE id=? AND user_id=?",
            (data["date"], data["title"], data.get("description", ""), data["category"],
             data.get("value"), mtype, target_value, current_val,
             int(is_achieved), linked, achieved_date, linked_metrics_str, currency, milestone_id, current_user_id())
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400


@app.route("/api/milestones/<int:milestone_id>", methods=["DELETE"])
@jwt_required()
def delete_milestone(milestone_id):
    """Delete a milestone by ID."""
    try:
        conn = db()
        conn.execute("DELETE FROM milestones WHERE id = ? AND user_id = ?", (milestone_id, current_user_id()))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/breakdown", methods=["GET"])
@jwt_required()
def get_breakdown():
    """Return current asset breakdown: cash (from accounts), super (from snapshot), stocks."""
    uid = current_user_id()
    conn = db()
    today = date.today().isoformat()
    row = conn.execute(
        "SELECT super FROM snapshots WHERE date <= ? AND user_id = ? ORDER BY date DESC LIMIT 1", (today, uid)
    ).fetchone()
    conn.close()
    super_val = row[0] if row else 0.0
    cash = get_total_cash(uid)
    portfolio_val, active_val, passive_val = _get_latest_portfolio_value(uid)
    return jsonify({
        "cash": cash,
        "super": round(super_val, 2),
        "stocks_active": round(active_val, 2),
        "stocks_passive": round(passive_val, 2),
        "portfolio": round(portfolio_val, 2),
        "total": round(cash + super_val + portfolio_val, 2),
    })

@app.route("/api/allocation", methods=["GET"])
@jwt_required()
def get_allocation():
    """Return dynamic country allocation using overrides, super holdings, and cash accounts."""
    uid = current_user_id()
    conn = db()
    today = date.today().isoformat()
    row = conn.execute(
        "SELECT super FROM snapshots WHERE date <= ? AND user_id = ? ORDER BY date DESC LIMIT 1", (today, uid)
    ).fetchone()
    conn.close()
    super_total = row[0] if row else 0.0

    txns = load_transactions(uid)
    conn = db()
    latest_prices = {}
    rows = conn.execute("""
        SELECT symbol, close FROM prices
        WHERE (symbol, date) IN (
            SELECT symbol, MAX(date) FROM prices GROUP BY symbol
        )
    """).fetchall()
    for sym, close in rows:
        latest_prices[sym] = close
    conn.close()
    audusd = latest_prices.get("AUDUSD=X", 0.65)

    txns_sorted = sorted(txns, key=lambda x: x["date"])
    holdings = {}
    for t in txns_sorted:
        sym = yf_symbol(t["ticker"], t["exchange"])
        if sym not in holdings:
            holdings[sym] = {
                "ticker": t["ticker"], "exchange": t["exchange"],
                "name": t.get("name", ""), "currency": t.get("currency") or "AUD", "units": 0.0
            }
        h = holdings[sym]
        qty = t["units"]
        if t["action"].lower() == "buy":
            h["units"] += qty
        elif t["action"].lower() == "split":
            h["units"] += qty
        elif t["action"].lower() == "sell":
            h["units"] = max(0, h["units"] - qty)

    # Aggregate by dynamic country labels
    countries = {}
    for sym, h in holdings.items():
        if h["units"] <= 1e-5:
            continue
        price = latest_prices.get(sym, 0.0)
        if h["currency"] == "USD":
            price = price / audusd
        val = h["units"] * price
        country = get_holding_country(h["ticker"], h["exchange"], h["name"], uid)
        countries[country] = countries.get(country, 0) + val

    # Super: distribute across countries from super_holdings.json
    super_holdings = load_super_holdings(uid)
    if super_holdings:
        for sh in super_holdings:
            c = sh.get("country", "Unknown")
            pct = sh.get("allocation_pct", 0) / 100.0
            countries[c] = countries.get(c, 0) + (super_total * pct)
    else:
        countries["AU"] = countries.get("AU", 0) + super_total

    # Cash accounts: distribute by country
    cash_accounts = load_cash_accounts(uid)
    for ca in cash_accounts:
        c = ca.get("country", "AU")
        countries[c] = countries.get(c, 0) + ca.get("balance", 0)

    total = sum(countries.values())
    result = {"countries": {}, "total": round(total, 2)}
    for country, value in sorted(countries.items(), key=lambda x: x[1], reverse=True):
        result["countries"][country] = {
            "value": round(value, 2),
            "pct": round(value / total * 100, 2) if total > 0 else 0.0
        }
    return jsonify(result)

# ─── Config CRUD Endpoints ─────────────────────────────────

@app.route("/api/cash-accounts", methods=["GET"])
@jwt_required()
def get_cash_accounts():
    return jsonify(load_cash_accounts(current_user_id()))

@app.route("/api/cash-accounts", methods=["POST"])
@jwt_required()
def save_cash_accounts_route():
    data = request.json
    try:
        save_cash_accounts(data, current_user_id())
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/super-holdings", methods=["GET"])
@jwt_required()
def get_super_holdings_route():
    return jsonify(load_super_holdings(current_user_id()))

@app.route("/api/super-holdings", methods=["POST"])
@jwt_required()
def save_super_holdings_route():
    data = request.json
    try:
        save_super_holdings(data, current_user_id())
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/country-overrides", methods=["GET"])
@jwt_required()
def get_country_overrides_route():
    return jsonify(load_country_overrides(current_user_id()))

@app.route("/api/country-overrides", methods=["POST"])
@jwt_required()
def save_country_overrides_route():
    data = request.json
    try:
        save_country_overrides(data, current_user_id())
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

# ─── End Config CRUD ───────────────────────────────────────

@app.route("/api/networth", methods=["GET"])
@jwt_required()
def get_networth():
    """Return combined net worth timeline: portfolio + cash + super."""
    uid = current_user_id()
    txns = load_transactions(uid)
    if not txns:
        return jsonify({"dates": [], "portfolio": [], "cash": [], "super": [], "net_worth": []})

    # Build daily portfolio value (same logic as /api/performance)
    df = pd.DataFrame(txns)
    df["date"] = pd.to_datetime(df["date"])
    df["sym"] = df.apply(lambda r: yf_symbol(r["ticker"], r["exchange"]), axis=1)

    start = df["date"].min()
    end = pd.Timestamp(date.today())
    all_dates = pd.date_range(start, end, freq="D")

    conn = db()
    price_data = {}
    symbols_to_read = list(df["sym"].unique()) + ["AUDUSD=X"]
    for sym in symbols_to_read:
        rows = conn.execute(
            "SELECT date, close FROM prices WHERE symbol = ? ORDER BY date", (sym,)
        ).fetchall()
        if rows:
            s = pd.Series({pd.Timestamp(d): c for d, c in rows}).reindex(all_dates).ffill().bfill()
            s = s.fillna(0.0 if sym != "AUDUSD=X" else 1.0)
        else:
            s = pd.Series(0.0 if sym != "AUDUSD=X" else 1.0, index=all_dates)
        price_data[sym] = s

    fx_rates = price_data["AUDUSD=X"]
    units_changes = pd.DataFrame(0.0, index=all_dates, columns=df["sym"].unique())
    sym_currency = df.groupby("sym")["currency"].first().to_dict()

    for _, row in df.iterrows():
        sym = row["sym"]
        action = row["action"].lower()
        if action == "buy":
            sign = 1
        elif action == "sell":
            sign = -1
        else:
            sign = 1  # split
        units_changes.loc[row["date"], sym] += sign * row["units"]

    units_df = units_changes.cumsum()
    portfolio_value = pd.Series(0.0, index=all_dates)
    split_ratios = _split_ratios_by_symbol(txns)
    for sym in df["sym"].unique():
        prices = price_data[sym]
        if sym_currency[sym] == "USD":
            prices = prices / fx_rates
        # Stored closes are split-adjusted; scale raw units by any split that happens
        # after each date so units and prices share one basis.
        factor = pd.Series(1.0, index=all_dates)
        for sd, ratio in split_ratios.get(sym, ()):
            factor.loc[all_dates < sd] *= ratio
        portfolio_value += units_df[sym] * factor * prices

    # Cumulative cash flow (cost basis) for return calculation
    cash_flow_changes = pd.Series(0.0, index=all_dates)
    for _, row in df.iterrows():
        cash_flow_changes.loc[row["date"]] += row["value"]
    cash_flow = cash_flow_changes.cumsum()
    return_val = portfolio_value - cash_flow

    # Build cash + super timeline (forward-filled from snapshots)
    snapshots = conn.execute("SELECT date, super, cash FROM snapshots WHERE user_id = ? ORDER BY date", (uid,)).fetchall()
    conn.close()

    cash_series = pd.Series(0.0, index=all_dates)
    super_series = pd.Series(0.0, index=all_dates)
    for s_date, s_super, s_cash in snapshots:
        ts = pd.Timestamp(s_date)
        if ts in all_dates:
            cash_series.loc[ts:] = s_cash
            super_series.loc[ts:] = s_super
    # Forward-fill from first snapshot back to start
    if snapshots:
        first_ts = pd.Timestamp(snapshots[0][0])
        cash_series.loc[:first_ts] = snapshots[0][2]
        super_series.loc[:first_ts] = snapshots[0][1]

    net_worth = portfolio_value + cash_series + super_series

    return jsonify({
        "dates": [d.strftime("%Y-%m-%d") for d in all_dates],
        "portfolio": portfolio_value.round(2).tolist(),
        "cash": cash_series.round(2).tolist(),
        "super": super_series.round(2).tolist(),
        "net_worth": net_worth.round(2).tolist(),
        "return_val": return_val.round(2).tolist(),
    })

@app.route("/api/monthly-change", methods=["GET"])
@jwt_required()
def get_monthly_change():
    """Return month-over-month net worth change."""
    uid = current_user_id()
    txns = load_transactions(uid)
    conn = db()
    all_snapshots = conn.execute("SELECT date, super, cash, COALESCE(source,'manual') FROM snapshots WHERE user_id = ? ORDER BY date", (uid,)).fetchall()
    conn.close()

    # One boundary per calendar month, at that month's FIRST snapshot — including the
    # month in progress, whose boundary is its 1st, not its latest reading.
    #
    # Deliberately NOT _pick_monthly_snapshots: that returns the LATEST snapshot for the
    # current month, because it powers the Compounder where that point IS the month-to-date
    # value. Used as a boundary it silently misattributes days — with a 6 Aug snapshot the
    # bar labelled July measured 1 Jul to 6 Aug (36 days, five of them August), and the MTD
    # bar shrank to the two days after it.
    first_of_month: dict = {}
    for s in sorted(all_snapshots, key=lambda r: r[0]):
        ym = pd.Timestamp(s[0]).to_period("M")
        if ym not in first_of_month:
            first_of_month[ym] = s
    snapshots = [first_of_month[k] for k in sorted(first_of_month)]

    if not snapshots:
        return jsonify({"months": [], "change": [], "change_pct": [],
                        "sources": [], "is_mtd": []})

    # Build portfolio value at each snapshot date
    df = pd.DataFrame(txns)
    df["date"] = pd.to_datetime(df["date"])
    df["sym"] = df.apply(lambda r: yf_symbol(r["ticker"], r["exchange"]), axis=1)

    conn = db()
    # Get price data
    price_data = {}
    symbols_to_read = list(df["sym"].unique()) + ["AUDUSD=X"]
    for sym in symbols_to_read:
        rows = conn.execute(
            "SELECT date, close FROM prices WHERE symbol = ? ORDER BY date", (sym,)
        ).fetchall()
        if rows:
            price_data[sym] = {pd.Timestamp(d): c for d, c in rows}
        else:
            price_data[sym] = {}
    conn.close()

    audusd_rates = price_data.get("AUDUSD=X", {})
    sym_currency = df.groupby("sym")["currency"].first().to_dict()
    split_ratios = _split_ratios_by_symbol(txns)

    def nw_at(as_of, s_cash, s_super):
        """Net worth on a date: portfolio valued at that date's prices, plus cash/super."""
        s_ts = pd.Timestamp(as_of)
        portfolio_val = 0.0
        for sym in df["sym"].unique():
            # Get units up to this date
            units = 0.0
            for _, row in df.iterrows():
                if row["date"] > s_ts:
                    break
                if row["sym"] != sym:
                    continue
                action = row["action"].lower()
                if action == "buy":
                    units += row["units"]
                elif action == "split":
                    units += row["units"]
                elif action == "sell":
                    units = max(0, units - row["units"])
            # Stored closes are split-adjusted, so raw pre-split units undervalue the
            # holding by the split ratio. Put both on the same basis.
            units *= _split_adj_factor(split_ratios.get(sym), s_ts)

            # Find closest price on or before snapshot date
            sym_prices = price_data.get(sym, {})
            available_dates = [d for d in sym_prices if d <= s_ts]
            if available_dates:
                closest = max(available_dates)
                price = sym_prices[closest]
                if sym_currency.get(sym, "AUD") == "USD":
                    aud_dates = [d for d in audusd_rates if d <= s_ts]
                    rate = audusd_rates[max(aud_dates)] if aud_dates else 0.65
                    price = price / rate
                portfolio_val += units * price
        return portfolio_val + s_cash + s_super

    # Value the portfolio at each month boundary first, then turn consecutive
    # boundaries into bars. Anchors are the 1st of each month (see _pick_monthly_snapshots).
    anchors = [(s_date, nw_at(s_date, s_cash, s_super), s_source)
               for s_date, s_super, s_cash, s_source in snapshots]

    # The month in progress has no closing boundary yet, so add one at today using the
    # most recent cash/super on file and today's prices. Without this the chart simply
    # stopped at the 1st and showed nothing at all for the current month.
    today = date.today().isoformat()
    mtd = False
    if anchors and today > anchors[-1][0]:
        latest = max(all_snapshots, key=lambda r: r[0])   # newest of ANY date, not just anchors
        anchors.append((today, nw_at(today, latest[2], latest[1]), latest[3]))
        mtd = True

    months, period_end, changes, changes_pct, sources, is_mtd = [], [], [], [], [], []
    # A change between two boundaries happened during the month the earlier one opens,
    # so label it with that month. Labelling by the closing boundary — which is what this
    # did before — reported every month's performance one month late.
    #
    # period_end is returned so the UI can state the span outright instead of relying on
    # the reader's idea of what a month name means. Both conventions are defensible: a
    # spreadsheet recorded on the 1st naturally files this change under the month it was
    # written down in (the closing date), while "May's return" normally means the return
    # during May. Naming the two dates ends the argument, and it cannot be inferred from
    # the labels alone — labelling by the closing date would give the month in progress
    # and the month before it the same name.
    for (start_date, start_nw, start_src), (end_date, end_nw, _) in zip(anchors, anchors[1:]):
        months.append(start_date)
        period_end.append(end_date)
        sources.append(start_src)
        change = end_nw - start_nw
        changes.append(round(change, 2))
        changes_pct.append(round((change / start_nw) * 100, 2) if start_nw > 0 else 0.0)
        is_mtd.append(False)
    if is_mtd:
        is_mtd[-1] = mtd

    return jsonify({"months": months, "period_end": period_end, "change": changes,
                    "change_pct": changes_pct, "sources": sources, "is_mtd": is_mtd})


def _compute_monthly_nw_series(uid):
    """Return month-by-month absolute NW data for a user.

    Each element: {date, nw, portfolio, cash, super, source, change_pct, is_current_month}
    change_pct is None for the first data point.
    """
    txns = load_transactions(uid)
    conn = db()
    all_snapshots = conn.execute(
        "SELECT date, super, cash, COALESCE(source,'manual') FROM snapshots WHERE user_id = ? ORDER BY date",
        (uid,)
    ).fetchall()
    conn.close()

    snapshots, current_month_flags = _pick_monthly_snapshots(all_snapshots)

    if not txns or not snapshots:
        return []

    df = pd.DataFrame(txns)
    df["date"] = pd.to_datetime(df["date"])
    df["sym"] = df.apply(lambda r: yf_symbol(r["ticker"], r["exchange"]), axis=1)
    df = df.sort_values("date")

    conn = db()
    price_data = {}
    symbols_to_read = list(df["sym"].unique()) + ["AUDUSD=X"]
    for sym in symbols_to_read:
        rows = conn.execute(
            "SELECT date, close FROM prices WHERE symbol = ? ORDER BY date", (sym,)
        ).fetchall()
        price_data[sym] = {pd.Timestamp(d): c for d, c in rows} if rows else {}
    conn.close()

    audusd_rates = price_data.get("AUDUSD=X", {})
    sym_currency = df.groupby("sym")["currency"].first().to_dict()
    split_ratios = _split_ratios_by_symbol(txns)

    result = []
    prev_nw = None
    prev_ts = None

    for (s_date, s_super, s_cash, s_source), is_current in zip(snapshots, current_month_flags):
        s_ts = pd.Timestamp(s_date)
        # The in-progress month is meant to read as month-to-date (see
        # _pick_monthly_snapshots), but only its cash and super legs ever did — the
        # portfolio was valued at the snapshot's own date, so it froze on whatever day
        # the last snapshot happened to fall. With snapshots taken on the 1st that left
        # the current FY up to a month behind the market: on 8 Aug it closed the year at
        # 1 Aug's prices and reported −0.94% while the dashboard showed +1.31%, $22,703
        # apart, because a week of gains simply wasn't in it.
        #
        # Valuing the live point at today puts it on the same basis as /api/breakdown,
        # which is what the dashboard shows. Cash and super still come from the latest
        # snapshot — that's the most recent reading there is, and it's the same source
        # the dashboard uses.
        val_ts = pd.Timestamp(date.today()) if is_current else s_ts
        # Cash too: /api/breakdown reads it live from cash_accounts rather than from a
        # snapshot, so the live point has to do the same or it drifts by whatever has
        # moved through the accounts since the last snapshot — $6,758 of it here.
        # Super does come from the latest snapshot in breakdown, so that one carries.
        if is_current:
            s_cash = get_total_cash(uid)
        portfolio_val = 0.0

        for sym in df["sym"].unique():
            units = 0.0
            for _, row in df.iterrows():
                if row["date"] > val_ts:
                    break
                if row["sym"] != sym:
                    continue
                action = row["action"].lower()
                if action == "buy":
                    units += row["units"]
                elif action == "split":
                    units += row["units"]
                elif action == "sell":
                    units = max(0, units - row["units"])
            # Stored closes are split-adjusted, so raw pre-split units undervalue the
            # holding by the split ratio. Put both on the same basis.
            units *= _split_adj_factor(split_ratios.get(sym), val_ts)

            sym_prices = price_data.get(sym, {})
            available_dates = [d for d in sym_prices if d <= val_ts]
            if available_dates:
                closest = max(available_dates)
                price = sym_prices[closest]
                if sym_currency.get(sym, "AUD") == "USD":
                    aud_dates = [d for d in audusd_rates if d <= val_ts]
                    rate = audusd_rates[max(aud_dates)] if aud_dates else 0.65
                    price = price / rate
                portfolio_val += units * price

        total_nw = portfolio_val + s_cash + s_super
        change_pct = None
        period_days = None
        if prev_nw is not None and prev_nw > 0:
            change_pct = round((total_nw - prev_nw) / prev_nw * 100, 2)
            period_days = (pd.Timestamp(s_date) - prev_ts).days

        result.append({
            "date": s_date,
            "nw": round(total_nw, 2),
            "portfolio": round(portfolio_val, 2),
            "cash": round(s_cash, 2),
            "super": round(s_super, 2),
            "source": s_source,
            "change_pct": change_pct,
            "is_current_month": is_current,
            # Days since the previous point. A change is only comparable to other months
            # if this is roughly a month; see get_compounder's rate-stat filter.
            "period_days": period_days,
        })
        prev_nw = total_nw
        prev_ts = pd.Timestamp(s_date)

    return result


@app.route("/api/compounder", methods=["GET"])
@jwt_required()
def get_compounder():
    """FY-grouped net worth analytics for the Compounder tab."""
    uid = current_user_id()
    monthly = _compute_monthly_nw_series(uid)

    if not monthly:
        return jsonify({
            "monthly": [],
            "fy_rows": [],
            "summary": {"peak_nw": 0, "avg_mom": 0, "months_positive": 0, "months_negative": 0},
        })

    # Rate-based stats (avg/best/worst month, pos/neg counts) only count steps that are
    # actually about a month long. Filtering on `is_current_month` instead was wrong: with
    # points anchored to the 1st, the in-progress month's step is usually a clean month
    # (1 Jul -> 1 Aug) and excluding it threw away a valid sample. It becomes genuinely
    # partial only once a mid-month snapshot lands, stretching the step past ~35 days.
    # Level-based figures (nw_end, peak_nw, growth_dollar/pct) always use the latest data.
    def _full_month(m):
        d = m.get("period_days")
        return m["change_pct"] is not None and d is not None and 20 <= d <= 40

    all_changes = [m["change_pct"] for m in monthly if _full_month(m)]
    peak_nw = max(m["nw"] for m in monthly)
    months_positive = sum(1 for c in all_changes if c > 0)
    months_negative = sum(1 for c in all_changes if c < 0)
    avg_mom = round(sum(all_changes) / len(all_changes), 2) if all_changes else 0

    def fy_year(date_str):
        ts = pd.Timestamp(date_str)
        return ts.year + 1 if ts.month >= 7 else ts.year

    # An FY runs 1 July to 30 June, so a row's opening and closing figures are anchored
    # to the July points on either side of it rather than to a calendar bucket of the
    # points. Two separate things made the old grouping report June to June:
    #
    #   - a point holds the net worth AT its date, so FY2026 closes on the 1 Jul 2026
    #     point, not on the 1 Jun 2026 one that ended a Jul-2025..Jun-2026 bucket;
    #   - a point's change_pct describes the month that ENDED there, so the 1 Jul 2025
    #     point carries June 2025's move — FY2025's final month, which was sitting at
    #     the head of FY2026's bucket and skewing its best/worst/average.
    #
    # Treating the July points as boundaries and taking each FY as the half-open window
    # between them — every point after its opening boundary, up to and including its
    # closing one — fixes both at once.
    july_idx = {}
    for i, m in enumerate(monthly):
        if pd.Timestamp(m["date"]).month == 7:
            july_idx[pd.Timestamp(m["date"]).year + 1] = i   # July 2025 opens FY2026

    fy_rows = []
    for fy in sorted({fy_year(m["date"]) for m in monthly}):
        open_i = july_idx.get(fy)                            # None if data starts mid-FY
        # No closing boundary yet means the year is still running, so it closes on the
        # latest point and nw_end reads as the year-to-date figure.
        close_i = july_idx.get(fy + 1, len(monthly) - 1)
        months_in_fy = monthly[(open_i + 1) if open_i is not None else 0 : close_i + 1]
        last = monthly[close_i]

        prior_nw = monthly[open_i]["nw"] if open_i is not None else None
        fy_changes = [m["change_pct"] for m in months_in_fy if _full_month(m)]

        fy_rows.append({
            "fy": f"FY{fy}",
            "nw_end": last["nw"],
            "prior_nw": prior_nw,
            "growth_dollar": round(last["nw"] - prior_nw, 2) if prior_nw is not None else None,
            "growth_pct": round((last["nw"] - prior_nw) / prior_nw * 100, 2) if prior_nw else None,
            "best_month": max(fy_changes) if fy_changes else None,
            "worst_month": min(fy_changes) if fy_changes else None,
            "avg_mom": round(sum(fy_changes) / len(fy_changes), 2) if fy_changes else None,
            "portfolio_end": last["portfolio"],
            "cash_end": last["cash"],
            "port_pct": round(last["portfolio"] / last["nw"] * 100, 2) if last["nw"] else None,
            "months_count": len(months_in_fy),
        })

    return jsonify({
        "monthly": monthly,
        "fy_rows": fy_rows,
        "summary": {
            "peak_nw": round(peak_nw, 2),
            "avg_mom": avg_mom,
            "months_positive": months_positive,
            "months_negative": months_negative,
        },
    })


# On startup: seed snapshots if DB is empty, then start background price sync
try:
    seed_historical_snapshots()
except Exception as e:
    print(f"[backend] seed_historical_snapshots failed (non-fatal, continuing startup): {e}")

# Background price sync — runs automatically after market close (UTC times)
# ASX closes ~06:00 UTC, NYSE/NASDAQ closes ~21:00 UTC
def _scheduled_sync():
    print(f"[scheduler] Auto-sync triggered at {datetime.now().isoformat()}")
    try:
        results = _run_sync()
        ok = sum(1 for r in results if r.get("ok"))
        print(f"[scheduler] Sync complete: {ok}/{len(results)} symbols OK")
    except Exception as e:
        print(f"[scheduler] Sync failed: {e}")

def _is_market_open() -> bool:
    """True if ASX or US market is in its REGULAR session (UTC clock, weekdays only)."""
    now = datetime.utcnow()
    if now.weekday() >= 5:          # Saturday or Sunday
        return False
    h = now.hour + now.minute / 60
    asx_open  = 0.0  <= h < 6.17   # 10:00–16:10 AEST  = 00:00–06:10 UTC
    us_open   = 14.5 <= h < 21.08  # 09:30–16:05 ET    = 14:30–21:05 UTC
    return asx_open or us_open


def _is_extended_hours() -> bool:
    """True during regular OR extended trading (US pre-market and after-hours).

    The refresh used to gate on _is_market_open, which covers regular sessions only. That
    left US pre-market (from 04:00 ET) and after-hours (to 20:00 ET) never refreshing, so
    a move there didn't land until the next scheduled sync — even though the Pre/After
    Market card reads those quotes live. Windows are approximate and ignore holidays;
    a non-trading day just means yfinance returns nothing and the refresh is a no-op.
    """
    now = datetime.utcnow()
    if now.weekday() >= 5:
        return False
    h = now.hour + now.minute / 60
    # ASX pre-open auction 07:00 AEST through close 16:10 AEST = 21:00–06:10 UTC.
    asx = h >= 21.0 or h < 6.17
    # US pre-market 04:00 ET through after-hours 20:00 ET = 08:00–00:00 UTC (EDT).
    us = 8.0 <= h < 24.0
    return asx or us


# Demand-driven price refresh. The scheduler alone can only ever be as fresh as its
# interval, so opening the app used to show prices up to 15 minutes old — or a day old
# outside the regular session. A read now kicks a background refresh when the cache is
# stale, which means no yfinance traffic at all when nobody is looking, and near-live
# numbers when someone is. Fire-and-forget on purpose: a full refresh of 11 symbols takes
# ~1s, but the request returns the current DB values immediately rather than waiting, and
# the frontend's 60s poll picks up the newer figures.
_LIVE_REFRESH = {"at": 0.0, "running": False}
_LIVE_REFRESH_LOCK = threading.Lock()
LIVE_REFRESH_SECONDS = int(os.environ.get("LIVE_REFRESH_SECONDS", "60"))


def _maybe_refresh_prices_async():
    """Kick a background intraday refresh if prices are stale and a market is active."""
    if not _is_extended_hours():
        return
    now = time.time()
    with _LIVE_REFRESH_LOCK:
        if _LIVE_REFRESH["running"]:
            return
        if now - _LIVE_REFRESH["at"] < LIVE_REFRESH_SECONDS:
            return
        _LIVE_REFRESH["at"] = now
        _LIVE_REFRESH["running"] = True

    def _work():
        try:
            _run_intraday_refresh()
        except Exception as e:
            print(f"[live] on-demand refresh failed (non-fatal): {e}")
        finally:
            with _LIVE_REFRESH_LOCK:
                _LIVE_REFRESH["running"] = False

    threading.Thread(target=_work, daemon=True, name="live-price-refresh").start()


def _scheduled_intraday_refresh():
    # Extended hours, not just the regular session — otherwise US pre-market and
    # after-hours moves never reach the DB until the next full sync.
    if not _is_extended_hours():
        return
    try:
        results = _run_intraday_refresh()
        ok = sum(1 for r in results if r.get("ok"))
        print(f"[scheduler] Intraday refresh: {ok}/{len(results)} symbols OK")
    except Exception as e:
        print(f"[scheduler] Intraday refresh failed: {e}")

# Monthly snapshot — cash has no bank connector, so it's only ever as current as the
# last time someone manually updated cash_accounts. This locks in a snapshot on the 1st
# of every month automatically using whatever's on record right then (live cash_accounts
# total + the last known super figure), so a monthly data point always exists without
# depending on remembering to click anything — and always lands on a real, non-future
# date since it only ever fires on the actual 1st.
def _scheduled_monthly_snapshot():
    print(f"[scheduler] Monthly snapshot triggered at {datetime.now().isoformat()}")
    conn = db()
    user_ids = [r[0] for r in conn.execute("SELECT id FROM users").fetchall()]
    conn.close()
    today_str = date.today().isoformat()
    for uid in user_ids:
        try:
            cash_total = get_total_cash(uid)
            conn = db()
            prior = conn.execute(
                "SELECT super FROM snapshots WHERE date <= ? AND user_id = ? ORDER BY date DESC LIMIT 1", (today_str, uid)
            ).fetchone()
            super_val = prior[0] if prior else 0.0
            conn.execute(
                "INSERT OR REPLACE INTO snapshots (date, super, cash, user_id, source) VALUES (?, ?, ?, ?, 'auto')",
                (today_str, super_val, cash_total, uid),
            )
            conn.commit()
            conn.close()
            print(f"[scheduler] Monthly snapshot logged for user {uid}: cash=${cash_total:.2f}, super=${super_val:.2f}")
        except Exception as e:
            print(f"[scheduler] Monthly snapshot failed for user {uid}: {e}")

_scheduler = BackgroundScheduler()
_scheduler.add_job(_scheduled_sync, "cron", hour=6, minute=15, id="asx_close")   # after ASX close
_scheduler.add_job(_scheduled_sync, "cron", hour=21, minute=15, id="us_close")   # after NYSE/NASDAQ close
# Floor that keeps data warm even with nobody using the app. Real freshness comes from
# the demand-driven refresh in /api/portfolio; this only has to stop the cache going
# properly cold, so a modest interval is enough. Tune with INTRADAY_REFRESH_MINUTES.
_scheduler.add_job(_scheduled_intraday_refresh, "interval",
                   minutes=int(os.environ.get("INTRADAY_REFRESH_MINUTES", "5")),
                   id="intraday_refresh")
_scheduler.add_job(_scheduled_monthly_snapshot, "cron", day=1, hour=13, minute=0,
                    timezone="Australia/Melbourne", id="monthly_snapshot")  # 1pm Melbourne time, 1st of month, DST-aware
_scheduler.start()
print(f"[scheduler] Auto-sync 06:15 UTC (ASX) + 21:15 UTC (US); intraday every "
      f"{os.environ.get('INTRADAY_REFRESH_MINUTES', '5')}min during extended hours; "
      f"on-demand refresh throttled to {LIVE_REFRESH_SECONDS}s")

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5050)